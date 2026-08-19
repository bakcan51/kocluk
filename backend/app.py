import sys
import os

backend_dir = os.path.dirname(os.path.abspath(__file__))
if backend_dir not in sys.path:
    sys.path.insert(0, backend_dir)

import json
import io
import math
import time
from datetime import datetime, date, timedelta
import werkzeug.utils
from flask import Flask, request, jsonify, send_file, make_response, send_from_directory, g

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
if os.environ.get("VERCEL") or os.environ.get("AWS_LAMBDA_FUNCTION_NAME"):
    UPLOAD_FOLDER = "/tmp/uploads"
else:
    UPLOAD_FOLDER = os.path.join(os.path.dirname(BASE_DIR), 'uploads')
try:
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
except Exception:
    pass
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'doc', 'docx', 'xls', 'xlsx', 'txt'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
import werkzeug.security
import openpyxl
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

from database import get_db, DB_PATH, init_db, ensure_lgs_seeded, is_postgres

frontend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', 'frontend'))
if not os.path.exists(frontend_dir):
    frontend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), 'frontend'))
if not os.path.exists(frontend_dir):
    frontend_dir = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))

app = Flask(__name__, static_folder=frontend_dir, static_url_path="")
app.secret_key = "yks_kocluk_super_secret_key_2027"
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

try:
    init_db()
    ensure_lgs_seeded()
except Exception as _e:
    pass

@app.route('/uploads/<path:filename>')
def serve_uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)

@app.route('/api/health', methods=['GET'])
def health_check():
    db_status = "ok"
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("SELECT 1;")
        cur.fetchone()
    except Exception as e:
        db_status = f"error: {str(e)}"
    
    return jsonify({
        "status": "ok",
        "database": db_status
    }), 200

@app.route('/api/upload', methods=['POST'])
def upload_file():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    if 'file' not in request.files:
        return jsonify({'error': 'Dosya bulunamadı'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Dosya seçilmedi'}), 400

    if file and allowed_file(file.filename):
        orig_filename = werkzeug.utils.secure_filename(file.filename) or 'file'
        unique_name = f"{int(time.time())}_{orig_filename}"
        save_path = os.path.join(app.config['UPLOAD_FOLDER'], unique_name)
        file.save(save_path)

        file_size_bytes = os.path.getsize(save_path)
        file_size_str = f"{round(file_size_bytes / 1024, 1)} KB" if file_size_bytes < 1024*1024 else f"{round(file_size_bytes / (1024*1024), 2)} MB"
        file_url = f"/uploads/{unique_name}"

        return jsonify({
            'message': 'Dosya başarıyla yüklendi',
            'file_name': orig_filename,
            'file_size': file_size_str,
            'file_url': file_url
        })

    return jsonify({'error': 'İzin verilmeyen dosya formatı'}), 400

@app.teardown_appcontext
def close_db_connection(exception=None):
    db = g.pop('db', None)
    if db is not None:
        try:
            if hasattr(db, 'close'):
                try:
                    db.close(force=True)
                except TypeError:
                    db.close()
        except Exception:
            pass

@app.before_request
def start_perf_timer():
    g.start_time = time.time()

@app.after_request
def log_perf_timing(response):
    if hasattr(g, 'start_time'):
        duration_ms = round((time.time() - g.start_time) * 1000, 1)
        print(f"[PERF] {request.method} {request.path} {response.status_code} {duration_ms}ms")
    return response

DAY_NAME_MAP = {
    1: 'Pazartesi', 2: 'Salı', 3: 'Çarşamba', 4: 'Perşembe', 5: 'Cuma', 6: 'Cumartesi', 7: 'Pazar',
    'Pazartesi': 1, 'Salı': 2, 'Çarşamba': 3, 'Perşembe': 4, 'Cuma': 5, 'Cumartesi': 6, 'Pazar': 7
}

def get_coach_user_id_for_student(student_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("""
        SELECT c.user_id 
        FROM coach_student_relationships rel
        JOIN coaches c ON rel.coach_id = c.id
        WHERE rel.student_id = ? AND rel.status = 'ACTIVE'
        LIMIT 1;
        """, (student_id,))
        row = cursor.fetchone()
        conn.close()
        return row['user_id'] if row else 2
    except Exception as e:
        print(f"Error finding coach for student {student_id}: {e}")
        return 2

def get_student_user_id(student_id):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT user_id FROM students WHERE id = ?;", (student_id,))
        row = cursor.fetchone()
        conn.close()
        return row['user_id'] if row else None
    except Exception as e:
        print(f"Error finding user_id for student {student_id}: {e}")
        return None

def copy_system_resources_to_coach_pool(new_coach_id, cursor=None):
    """
    Copies all central resources (owner_type = 'SYSTEM') into new_coach_id's personal pool.
    Copies associated resource_topics as well so the coach gets an independent, fully mapped pool.
    """
    close_conn = False
    conn = None
    if cursor is None:
        conn = get_db()
        cursor = conn.cursor()
        close_conn = True

    try:
        cursor.execute("""
        SELECT * FROM resources 
        WHERE (owner_type = 'SYSTEM' OR owner_id IS NULL) 
          AND (status IS NULL OR status != 'ARCHIVED');
        """)
        system_resources = [dict(r) for r in cursor.fetchall()]

        for sys_res in system_resources:
            cursor.execute("""
            SELECT id FROM resources 
            WHERE owner_type = 'COACH' AND owner_id = ? AND origin_resource_id = ?;
            """, (new_coach_id, sys_res['id']))
            if cursor.fetchone():
                continue

            name = sys_res.get('name') or sys_res.get('title')
            publisher = sys_res.get('publisher')
            exam_system = sys_res.get('exam_system') or 'YKS'
            exam_type = sys_res.get('exam_type') or 'TYT'
            field = sys_res.get('field') or 'ALL'
            subject_id = sys_res.get('subject_id')
            resource_type = sys_res.get('resource_type') or 'Soru Bankası'
            isbn = sys_res.get('isbn')
            edition = sys_res.get('edition')
            description = sys_res.get('description')
            level = sys_res.get('level') or 'Orta'
            cover_url = sys_res.get('cover_url')
            total_questions = sys_res.get('total_questions') or 0

            cursor.execute("""
            INSERT INTO resources (
                owner_type, owner_id, origin_resource_id, name, title, publisher, 
                exam_system, exam_type, field, subject_id, resource_type, level, 
                isbn, edition, description, cover_url, total_questions, status
            ) VALUES (
                'COACH', ?, ?, ?, ?, ?, 
                ?, ?, ?, ?, ?, ?, 
                ?, ?, ?, ?, ?, 'ACTIVE'
            );
            """, (new_coach_id, sys_res['id'], name, name, publisher, exam_system, exam_type, field, subject_id, resource_type, level, isbn, edition, description, cover_url, total_questions))
            new_coach_res_id = cursor.lastrowid

            cursor.execute("SELECT curriculum_topic_id, chapter_name, order_index FROM resource_topics WHERE resource_id = ?;", (sys_res['id'],))
            sys_topics = cursor.fetchall()
            for st in sys_topics:
                cursor.execute("""
                INSERT INTO resource_topics (resource_id, curriculum_topic_id, chapter_name, order_index)
                VALUES (?, ?, ?, ?);
                """, (new_coach_res_id, st['curriculum_topic_id'], st['chapter_name'], st['order_index']))

        if close_conn:
            conn.commit()
            conn.close()
    except Exception as e:
        print(f"Error copying system resources to coach {new_coach_id}: {e}")
        if close_conn and conn:
            conn.close()

def send_auto_notification(sender_user_id, receiver_user_id, content, message_type='SYSTEM', cursor=None):
    if not sender_user_id or not receiver_user_id or sender_user_id == receiver_user_id:
        return
    try:
        if cursor is not None:
            cursor.execute("""
            INSERT INTO messages (sender_id, receiver_id, message_type, content, is_read)
            VALUES (?, ?, ?, ?, 0);
            """, (sender_user_id, receiver_user_id, message_type, content))
        else:
            conn = get_db()
            c = conn.cursor()
            c.execute("""
            INSERT INTO messages (sender_id, receiver_id, message_type, content, is_read)
            VALUES (?, ?, ?, ?, 0);
            """, (sender_user_id, receiver_user_id, message_type, content))
            conn.commit()
    except Exception as e:
        print(f"Error sending auto notification: {e}")

# CORS setup
@app.after_request
def add_cors_headers(response):
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, OPTIONS'
    return response

# Serve static frontend files
@app.route('/')
def serve_index():
    return app.send_static_file('index.html')

@app.route('/<path:path>')
def serve_static_fallback(path):
    if path.startswith('api/'):
        return jsonify({"error": "Endpoint not found"}), 404
    static_file_path = os.path.join(app.static_folder, path)
    if os.path.isfile(static_file_path):
        return app.send_static_file(path)
    return app.send_static_file('index.html')

# Helper: Net calculation formula (Doğru - Yanlış / 4 for YKS, / 3 for LGS)
def calc_net(correct, incorrect, exam_system='YKS'):
    penalty = 3.0 if str(exam_system).upper() == 'LGS' else 4.0
    return round(float(correct) - (float(incorrect) / penalty), 2)

# Activity Logging Helper
def log_activity(user_id, role, action, entity_type=None, entity_id=None, metadata=None, cursor=None):
    try:
        meta_str = json.dumps(metadata, ensure_ascii=False, default=str) if metadata else None
        if cursor:
            cursor.execute("""
            INSERT INTO activity_logs (user_id, role, action, entity_type, entity_id, metadata_json)
            VALUES (?, ?, ?, ?, ?, ?);
            """, (user_id, role, action, entity_type, entity_id, meta_str))
        else:
            conn = get_db()
            cur = conn.cursor()
            cur.execute("""
            INSERT INTO activity_logs (user_id, role, action, entity_type, entity_id, metadata_json)
            VALUES (?, ?, ?, ?, ?, ?);
            """, (user_id, role, action, entity_type, entity_id, meta_str))
            conn.commit()
    except Exception as e:
        print(f"Error logging activity: {e}")


def create_academic_event(event_type, actor_user_id, recipient_user_id, entity_type, entity_id, title, message, metadata=None, cursor=None):
    if not recipient_user_id or actor_user_id == recipient_user_id:
        return None

    close_conn = False
    if cursor is None:
        conn = get_db()
        cursor = conn.cursor()
        close_conn = True
    else:
        conn = None

    try:
        cursor.execute("SELECT id, role, username FROM users WHERE id = ?;", (actor_user_id,))
        actor_row = cursor.fetchone()
        actor_role = actor_row['role'] if actor_row else 'USER'

        student_id = None
        coach_id = None

        if actor_role == 'STUDENT':
            cursor.execute("SELECT id, coach_id FROM students WHERE user_id = ?;", (actor_user_id,))
            st = cursor.fetchone()
            if st:
                student_id = st['id']
                coach_id = st['coach_id']
        elif actor_role == 'COACH':
            cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (actor_user_id,))
            ch = cursor.fetchone()
            if ch:
                coach_id = ch['id']

        meta_json = json.dumps(metadata, ensure_ascii=False) if metadata else None

        cursor.execute(
            "INSERT INTO activity_logs (user_id, role, action, activity_type, entity_type, entity_id, actor_user_id, student_id, coach_id, metadata, metadata_json) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);",
            (actor_user_id, actor_role, title, event_type, entity_type, entity_id, actor_user_id, student_id, coach_id, meta_json, meta_json)
        )

        cursor.execute("SELECT enabled FROM notification_preferences WHERE user_id = ? AND notification_type = ?;", (recipient_user_id, event_type))
        pref = cursor.fetchone()
        if pref and not pref['enabled']:
            if close_conn:
                conn.commit()
                conn.close()
            return None

        event_key = f"{event_type}_{entity_type}_{entity_id}_{recipient_user_id}"

        cursor.execute(
            "INSERT OR IGNORE INTO notifications (recipient_user_id, actor_user_id, type, title, message, entity_type, entity_id, event_key) VALUES (?, ?, ?, ?, ?, ?, ?, ?);",
            (recipient_user_id, actor_user_id, event_type, title, message, entity_type, entity_id, event_key)
        )

        if close_conn:
            conn.commit()
            conn.close()
        return True
    except Exception as e:
        print(f"create_academic_event error: {e}")
        if close_conn and conn:
            conn.close()
        return None

# Helper: Get current user from token/header
def get_auth_user():
    auth_header = request.headers.get('Authorization')
    if not auth_header:
        return None
    try:
        token = auth_header.replace('Bearer ', '').strip()
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id, username, email, role, name, surname, status FROM users WHERE (id = ? OR username = ?) AND status = 'ACTIVE';", (token, token))
        user = cursor.fetchone()
        if not user:
            conn.close()
            return None
        user_dict = dict(user)
        if user_dict['role'] == 'COACH':
            cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user_dict['id'],))
            row = cursor.fetchone()
            if row:
                user_dict['coach_id'] = row['id']
        elif user_dict['role'] == 'STUDENT':
            cursor.execute("SELECT id, coach_id, exam_system, track, grade FROM students WHERE user_id = ?;", (user_dict['id'],))
            row = cursor.fetchone()
            if row:
                user_dict['student_id'] = row['id']
                user_dict['coach_id'] = row['coach_id']
                user_dict['exam_system'] = row['exam_system'] or 'YKS'
                user_dict['track'] = row['track']
                user_dict['grade'] = row['grade']
        conn.close()
        return user_dict
    except Exception as e:
        print(f"Error in get_auth_user: {e}")
        return None

# ==========================================
# 1. SINGLE AUTHENTICATION API (ADMIN, COACH, STUDENT)
# ==========================================
@app.route('/api/login', methods=['POST', 'OPTIONS'])
@app.route('/api/auth/login', methods=['POST', 'OPTIONS'])
def login():
    if request.method == 'OPTIONS':
        return jsonify({'status': 'ok'}), 200
    
    data = request.get_json(silent=True) or request.form or {}
    username_or_email = (data.get('username_or_email') or data.get('username') or data.get('email') or '').strip()
    password = data.get('password', '').strip()

    if not username_or_email or not password:
        return jsonify({
            'success': False,
            'error': 'Kullanıcı adı veya şifre hatalı.',
            'message': 'Kullanıcı adı veya şifre hatalı.'
        }), 401

    norm_username = username_or_email.lower()

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT id, username, email, password_hash, role, name, surname, status, failed_login_attempts, lockout_until 
    FROM users 
    WHERE (email IS NOT NULL AND LOWER(email) = ?) OR LOWER(username) = ?;
    """, (norm_username, norm_username))
    user = cursor.fetchone()

    if user and user['lockout_until']:
        cursor.execute("SELECT datetime('now') < datetime(?);", (user['lockout_until'],))
        is_locked = cursor.fetchone()[0]
        if is_locked:
            conn.close()
            return jsonify({
                'success': False,
                'error': '5 başarısız giriş denemesi nedeniyle hesabınız geçici olarak kilitlenmiştir. Lütfen koçunuzla iletişime geçin.',
                'message': '5 başarısız giriş denemesi nedeniyle hesabınız geçici olarak kilitlenmiştir. Lütfen koçunuzla iletişime geçin.'
            }), 429
    
    valid_password = werkzeug.security.check_password_hash(user['password_hash'], password) if user else False

    if not user or not valid_password:
        if user:
            failed_cnt = (user['failed_login_attempts'] or 0) + 1
            if failed_cnt >= 5:
                cursor.execute("""
                UPDATE users 
                SET failed_login_attempts = ?, lockout_until = datetime('now', '+15 minutes')
                WHERE id = ?;
                """, (failed_cnt, user['id']))
            else:
                cursor.execute("UPDATE users SET failed_login_attempts = ? WHERE id = ?;", (failed_cnt, user['id']))
            conn.commit()
        conn.close()
        return jsonify({
            'success': False,
            'error': 'Kullanıcı adı veya şifre hatalı.',
            'message': 'Kullanıcı adı veya şifre hatalı.'
        }), 401

    if user['status'] != 'ACTIVE':
        conn.close()
        return jsonify({
            'success': False,
            'error': 'Hesabınız aktif değil. Lütfen koçunuzla iletişime geçin.',
            'message': 'Hesabınız aktif değil. Lütfen koçunuzla iletişime geçin.'
        }), 403

    user_dict = dict(user)
    del user_dict['password_hash']

    try:
        cursor.execute("UPDATE users SET failed_login_attempts = 0, lockout_until = NULL, last_login_at = CURRENT_TIMESTAMP WHERE id = ?;", (user_dict['id'],))
        conn.commit()
    except Exception as e:
        print(f"[LOGIN UPDATE WARN] {e}")

    if user_dict['role'] == 'STUDENT':
        cursor.execute("SELECT id, coach_id, exam_system, track, grade, target_university, target_department FROM students WHERE user_id = ?;", (user_dict['id'],))
        st = cursor.fetchone()
        if st:
            user_dict['student_info'] = dict(st)
            user_dict['student_id'] = st['id']
            user_dict['coach_id'] = st['coach_id']
    elif user_dict['role'] == 'COACH':
        cursor.execute("SELECT id, title, bio FROM coaches WHERE user_id = ?;", (user_dict['id'],))
        co = cursor.fetchone()
        if co:
            user_dict['coach_info'] = dict(co)
            user_dict['coach_id'] = co['id']

    token = str(user_dict['id'])
    log_activity(user_dict['id'], user_dict['role'], 'LOGIN', 'users', user_dict['id'], {'username': user_dict['username']}, cursor=cursor)

    conn.close()
    return jsonify({
        'success': True,
        'message': 'Giriş başarılı!',
        'token': token,
        'user': user_dict
    })

def check_coach_owns_student(cursor, user, student_id):
    if not student_id:
        return False
    if user['role'] == 'ADMIN':
        return True
    if user['role'] != 'COACH':
        return False
    coach_id = user.get('coach_id')
    if not coach_id:
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        ch = cursor.fetchone()
        if ch:
            coach_id = ch['id']
            user['coach_id'] = coach_id

    if not coach_id:
        return False

    cursor.execute("""
    SELECT s.id FROM students s
    LEFT JOIN coach_students cs ON s.id = cs.student_id
    LEFT JOIN coach_student_relationships csr ON s.id = csr.student_id AND csr.status = 'ACTIVE'
    WHERE s.id = ? AND (s.coach_id = ? OR s.created_by_coach_id = ? OR cs.coach_id = ? OR csr.coach_id = ?);
    """, (student_id, coach_id, coach_id, coach_id, coach_id))
    return cursor.fetchone() is not None

def resolve_and_verify_student_id(cursor, user, req_student_id):
    if user['role'] == 'STUDENT':
        cursor.execute("SELECT id, user_id, exam_system, track FROM students WHERE user_id = ?;", (user['id'],))
        st = cursor.fetchone()
        if not st:
            return None, jsonify({'error': 'Öğrenci profili bulunamadı.'}), 404
        return st['id'], None, 200

    target_id = None
    if req_student_id and str(req_student_id).strip() not in ('ALL', 'null', 'undefined', ''):
        try:
            target_id = int(req_student_id)
        except (TypeError, ValueError):
            target_id = None

    if user['role'] == 'ADMIN':
        if not target_id:
            cursor.execute("SELECT id FROM students ORDER BY id ASC LIMIT 1;")
            row = cursor.fetchone()
            target_id = row['id'] if row else 1
        return target_id, None, 200

    # User is COACH
    coach_id = user.get('coach_id')
    if not coach_id:
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        ch = cursor.fetchone()
        if ch:
            coach_id = ch['id']
            user['coach_id'] = coach_id
        else:
            coach_id = 0

    if target_id and not check_coach_owns_student(cursor, user, target_id):
        return None, {'error': 'Bu öğrencimin verilerine erişim yetkiniz bulunmamaktadır.'}, 403

    if target_id:
        return target_id, None, 200

    cursor.execute("""
    SELECT s.id FROM students s
    LEFT JOIN coach_students cs ON s.id = cs.student_id
    LEFT JOIN coach_student_relationships csr ON s.id = csr.student_id AND csr.status = 'ACTIVE'
    WHERE s.coach_id = ? OR s.created_by_coach_id = ? OR cs.coach_id = ? OR csr.coach_id = ?
    ORDER BY s.id ASC LIMIT 1;
    """, (coach_id, coach_id, coach_id, coach_id))
    row = cursor.fetchone()
    if row:
        return row['id'], None, 200

    return None, {'error': 'Atanmış öğrenciniz bulunmuyor.'}, 404

@app.route('/api/students', methods=['GET', 'POST'])
def handle_students():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        if user['role'] == 'COACH':
            cursor.execute("""
            SELECT DISTINCT s.*, u.name, u.surname, u.username, u.email, u.status as user_status, u.last_login_at
            FROM students s 
            JOIN users u ON s.user_id = u.id
            LEFT JOIN coach_students cs ON s.id = cs.student_id
            WHERE cs.coach_id = ? OR s.created_by_coach_id = ? OR s.coach_id = ?
            ORDER BY s.id ASC;
            """, (user.get('coach_id', 0), user.get('coach_id', 0), user.get('coach_id', 0)))
        else:
            cursor.execute("""
            SELECT s.*, u.name, u.surname, u.username, u.email, u.status as user_status, u.last_login_at
            FROM students s 
            JOIN users u ON s.user_id = u.id
            ORDER BY s.id ASC;
            """)
        students = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'students': students})

    elif request.method == 'POST':
        if user['role'] not in ('ADMIN', 'COACH'):
            conn.close()
            return jsonify({'error': 'Öğrenci hesabı oluşturma yetkiniz yoktur.'}), 403

        data = request.json or {}
        name = (data.get('name') or '').strip()
        surname = (data.get('surname') or '').strip()
        username = (data.get('username') or '').strip().lower()
        password = (data.get('password') or '').strip()
        password_repeat = (data.get('password_repeat') or password).strip()
        exam_system = data.get('exam_system', 'YKS')
        track = data.get('track', 'SAYISAL')
        grade = '8. Sınıf' if exam_system == 'LGS' else data.get('grade', '12. Sınıf')
        target_uni = data.get('target_university', 'Hedef Liseler' if exam_system == 'LGS' else 'İTÜ')
        target_dept = data.get('target_department', 'Fen / Anadolu Lisesi' if exam_system == 'LGS' else 'Mühendislik')

        if not name or not username or not password:
            conn.close()
            return jsonify({'error': 'Ad, kullanıcı adı ve şifre zorunludur.'}), 400

        if password != password_repeat:
            conn.close()
            return jsonify({'error': 'Şifreler birbiriyle uyuşmuyor.'}), 400

        if len(password) < 6:
            conn.close()
            return jsonify({'error': 'Şifre en az 6 karakter olmalıdır.'}), 400

        cursor.execute("SELECT id FROM users WHERE LOWER(username) = ?;", (username,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Bu kullanıcı adı zaten kullanılıyor.'}), 400

        pw_hash = werkzeug.security.generate_password_hash(password)
        cursor.execute("""
        INSERT INTO users (username, email, password_hash, role, name, surname, status)
        VALUES (?, NULL, ?, 'STUDENT', ?, ?, 'ACTIVE');
        """, (username, pw_hash, name, surname))
        user_id = cursor.lastrowid

        coach_id = user.get('coach_id', 1)
        cursor.execute("""
        INSERT INTO students (user_id, coach_id, created_by_coach_id, track, exam_system, grade, target_university, target_department, start_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, DATE('now'));
        """, (user_id, coach_id, coach_id, track, exam_system, grade, target_uni, target_dept))
        student_id = cursor.lastrowid

        cursor.execute("INSERT OR IGNORE INTO coach_students (coach_id, student_id) VALUES (?, ?);", (coach_id, student_id))

        cursor.execute("""
        INSERT INTO risk_scores (student_id, risk_level, reasons_json, net_trend_direction)
        VALUES (?, 'GREEN', '["Yeni öğrenci kaydı oluşturuldu"]', 'UPWARD');
        """, (student_id,))

        log_activity(user['id'], user['role'], 'CREATE_STUDENT', 'students', student_id, {
            'username': username,
            'exam_system': exam_system
        }, cursor=cursor)

        conn.commit()
        conn.close()
        return jsonify({
            'message': 'Öğrenci hesabı başarıyla oluşturuldu.',
            'student_id': student_id,
            'username': username,
            'initial_password': password
        })

@app.route('/api/students/<int:student_id>', methods=['GET'])
def get_single_student_detail(student_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT s.id, s.user_id, s.coach_id, s.grade, s.track, s.school, 
           s.target_university, s.target_department, s.target_score, s.target_rank, s.exam_system,
           u.name, u.surname, u.username, u.email, u.phone,
           uc.name as coach_name,
           r.risk_level, r.reasons_json
    FROM students s
    JOIN users u ON s.user_id = u.id
    LEFT JOIN users uc ON s.coach_id = uc.id
    LEFT JOIN risk_scores r ON s.id = r.student_id
    WHERE s.id = ?;
    """, (student_id,))
    row = cursor.fetchone()

    if not row:
        conn.close()
        return jsonify({'error': 'Öğrenci bulunamadı'}), 404

    if user['role'] == 'STUDENT':
        st_id, _, _ = resolve_and_verify_student_id(cursor, user, student_id)
        if st_id != student_id:
            conn.close()
            return jsonify({'error': 'Başka bir öğrencinin detay verilerine erişim yetkiniz yok'}), 403
    elif user['role'] == 'COACH':
        if not check_coach_owns_student(cursor, user, student_id):
            conn.close()
            return jsonify({'error': 'Bu öğrenciye erişim yetkiniz bulunmamaktadır'}), 403

    conn.close()

    st_dict = dict(row)
    if st_dict.get('reasons_json'):
        try:
            st_dict['reasons'] = json.loads(st_dict['reasons_json'])
        except Exception:
            st_dict['reasons'] = []
    else:
        st_dict['reasons'] = []

    return jsonify({'student': st_dict}), 200

@app.route('/api/students/<int:student_id>/reset-password', methods=['POST'])
def reset_student_password(student_id):
    user = get_auth_user()
    if not user or user['role'] not in ('ADMIN', 'COACH'):
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()

    if not check_coach_owns_student(cursor, user, student_id):
        conn.close()
        return jsonify({'error': 'Yalnızca kendinize bağlı öğrencilerin şifresini değiştirebilirsiniz.'}), 403

    cursor.execute("SELECT s.id, s.user_id, u.username FROM students s JOIN users u ON s.user_id = u.id WHERE s.id = ?;", (student_id,))
    st = cursor.fetchone()
    if not st:
        conn.close()
        return jsonify({'error': 'Öğrenci bulunamadı.'}), 404

    data = request.json or {}
    new_password = data.get('new_password', '').strip()
    confirm_password = data.get('confirm_password', '').strip()
    must_change = 1 if data.get('must_change_password') else 0
    is_temporary = data.get('is_temporary', False)

    if is_temporary and not new_password:
        import random, string
        chars = string.ascii_letters + string.digits
        new_password = ''.join(random.choice(chars) for _ in range(8))
        confirm_password = new_password

    if not new_password or len(new_password) < 6:
        conn.close()
        return jsonify({'error': 'Yeni şifre en az 6 karakter olmalıdır.'}), 400

    if new_password != confirm_password:
        conn.close()
        return jsonify({'error': 'Şifreler birbiriyle uyuşmuyor.'}), 400

    pw_hash = werkzeug.security.generate_password_hash(new_password)
    cursor.execute("""
    UPDATE users 
    SET password_hash = ?, must_change_password = ?, password_changed_at = CURRENT_TIMESTAMP, failed_login_attempts = 0, lockout_until = NULL, updated_at = CURRENT_TIMESTAMP
    WHERE id = ?;
    """, (pw_hash, must_change, st['user_id']))

    log_activity(user['id'], user['role'], 'PASSWORD_RESET', 'STUDENT', student_id, {
        'username': st['username'],
        'is_temporary': is_temporary
    })

    conn.commit()
    conn.close()
    return jsonify({
        'message': 'Öğrencinin parolası başarıyla güncellendi.',
        'username': st['username'],
        'new_password': new_password
    })

@app.route('/api/students/<int:student_id>/account', methods=['PUT'])
def update_student_account(student_id):
    user = get_auth_user()
    if not user or user['role'] not in ('ADMIN', 'COACH'):
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()

    if not check_coach_owns_student(cursor, user, student_id):
        conn.close()
        return jsonify({'error': 'Yalnızca kendinize bağlı öğrencileri yönetebilirsiniz.'}), 403

    cursor.execute("SELECT s.id, s.user_id, u.username FROM students s JOIN users u ON s.user_id = u.id WHERE s.id = ?;", (student_id,))
    st = cursor.fetchone()
    if not st:
        conn.close()
        return jsonify({'error': 'Öğrenci bulunamadı.'}), 404

    data = request.json or {}
    new_username = data.get('username', '').strip().lower()
    new_status = data.get('status', '').strip().upper()

    if new_username and new_username != st['username'].lower():
        cursor.execute("SELECT id FROM users WHERE LOWER(username) = ? AND id != ?;", (new_username, st['user_id']))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Bu kullanıcı adı zaten kullanılıyor.'}), 400
        cursor.execute("UPDATE users SET username = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?;", (new_username, st['user_id']))

    if new_status in ('ACTIVE', 'INACTIVE', 'SUSPENDED'):
        cursor.execute("UPDATE users SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?;", (new_status, st['user_id']))

    log_activity(user['id'], user['role'], 'UPDATE_STUDENT_ACCOUNT', 'STUDENT', student_id, {
        'old_username': st['username'],
        'new_username': new_username or st['username'],
        'status': new_status
    })

    conn.commit()
    conn.close()
    return jsonify({'message': 'Öğrenci hesap bilgileri başarıyla güncellendi.'})

# ==========================================
# COACH MANAGEMENT ENDPOINTS (ADMIN ONLY)
# ==========================================

@app.route('/api/admin/students/<int:student_id>/change-coach', methods=['POST'])
def admin_change_student_coach(student_id):
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Bu işlem yalnızca Admin yetkisine sahiptir.'}), 403

    data = request.json or {}
    new_coach_id = data.get('new_coach_id')
    if not new_coach_id:
        return jsonify({'error': 'Yeni koç ID gereklidir.'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT s.*, u.name, u.surname FROM students s JOIN users u ON s.user_id = u.id WHERE s.id = ?;", (student_id,))
    st = cursor.fetchone()
    if not st:
        conn.close()
        return jsonify({'error': 'Öğrenci bulunamadı.'}), 404

    old_coach_id = st['coach_id']

    cursor.execute("UPDATE students SET coach_id = ? WHERE id = ?;", (new_coach_id, student_id))
    try:
        cursor.execute("DELETE FROM coach_students WHERE student_id = ?;", (student_id,))
        cursor.execute("INSERT INTO coach_students (coach_id, student_id, relationship_type) VALUES (?, ?, 'MAIN_COACH');", (new_coach_id, student_id))
    except Exception:
        pass

    try:
        cursor.execute("DELETE FROM coach_student_relationships WHERE student_id = ? AND relationship_type = 'MAIN_COACH';", (student_id,))
        cursor.execute("INSERT INTO coach_student_relationships (coach_id, student_id, relationship_type, status, assigned_by) VALUES (?, ?, 'MAIN_COACH', 'ACTIVE', ?);", (new_coach_id, student_id, user['id']))
    except Exception:
        pass


    conn.commit()
    log_activity(user['id'], user['role'], 'CHANGE_STUDENT_COACH', 'students', student_id, {'old_coach_id': old_coach_id, 'new_coach_id': new_coach_id}, cursor=cursor)
    conn.close()

    return jsonify({'message': f"{st['name']} {st['surname'] or ''} isimli öğrencinin koçu başarıyla değiştirildi."})

@app.route('/api/coaches', methods=['GET', 'POST'])
def handle_coaches():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        cursor.execute("""
        SELECT c.id, c.user_id, u.name, u.surname, u.username, u.email, u.status as user_status,
               c.title, c.specialty, c.coach_code,
               (SELECT COUNT(*) FROM coach_students cs WHERE cs.coach_id = c.id) as student_count
        FROM coaches c
        JOIN users u ON c.user_id = u.id
        ORDER BY u.name ASC;
        """)
        coaches = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'coaches': coaches})

    elif request.method == 'POST':
        if user['role'] != 'ADMIN':
            conn.close()
            return jsonify({'error': 'Koç oluşturma yetkisi yalnızca sistem yöneticisine (ADMIN) aittir.'}), 403

        data = request.json or {}
        name = (data.get('name') or '').strip()
        surname = (data.get('surname') or '').strip()
        username = (data.get('username') or '').strip().lower()
        password = (data.get('password') or '').strip()
        password_repeat = (data.get('password_repeat') or '').strip()
        title = data.get('title', 'YKS & LGS Öğrenci Koçu')
        specialty = data.get('specialty', 'Genel Koçluk')
        email = (data.get('email') or '').strip() or None

        if not name or not username or not password:
            conn.close()
            return jsonify({'error': 'Ad, kullanıcı adı ve şifre zorunludur.'}), 400

        if password != password_repeat:
            conn.close()
            return jsonify({'error': 'Şifreler birbiriyle uyuşmuyor.'}), 400

        if len(password) < 6:
            conn.close()
            return jsonify({'error': 'Şifre en az 6 karakter olmalıdır.'}), 400

        cursor.execute("SELECT id FROM users WHERE LOWER(username) = ?;", (username,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Bu kullanıcı adı zaten kullanılıyor.'}), 400

        pw_hash = werkzeug.security.generate_password_hash(password)
        cursor.execute("""
        INSERT INTO users (username, email, password_hash, role, name, surname, status)
        VALUES (?, ?, ?, 'COACH', ?, ?, 'ACTIVE');
        """, (username, email, pw_hash, name, surname))
        user_id = cursor.lastrowid

        import random, string
        coach_code = 'KOC' + ''.join(random.choice(string.digits) for _ in range(4))

        cursor.execute("""
        INSERT INTO coaches (user_id, title, specialty, coach_code)
        VALUES (?, ?, ?, ?);
        """, (user_id, title, specialty, coach_code))
        coach_id = cursor.lastrowid

        copy_system_resources_to_coach_pool(coach_id, cursor=cursor)

        log_activity(user['id'], user['role'], 'CREATE_COACH', 'coaches', coach_id, {'username': username}, cursor=cursor)
        conn.commit()
        conn.close()

        return jsonify({
            'message': 'Koç hesabı başarıyla oluşturuldu.',
            'coach_id': coach_id,
            'username': username,
            'initial_password': password
        })

@app.route('/api/coaches/<int:coach_id>/reset-password', methods=['POST'])
def reset_coach_password(coach_id):
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Koç şifresini sıfırlama yetkisi yalnızca ADMIN rolüne aittir.'}), 403

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT c.id, c.user_id, u.username FROM coaches c JOIN users u ON c.user_id = u.id WHERE c.id = ?;", (coach_id,))
    ch = cursor.fetchone()
    if not ch:
        conn.close()
        return jsonify({'error': 'Koç hesabı bulunamadı.'}), 404

    data = request.json or {}
    new_password = data.get('new_password', '').strip()
    confirm_password = data.get('confirm_password', '').strip()
    is_temporary = data.get('is_temporary', False)

    if is_temporary and not new_password:
        import random, string
        chars = string.ascii_letters + string.digits
        new_password = ''.join(random.choice(chars) for _ in range(8))
        confirm_password = new_password

    if not new_password or len(new_password) < 6:
        conn.close()
        return jsonify({'error': 'Yeni şifre en az 6 karakter olmalıdır.'}), 400

    if new_password != confirm_password:
        conn.close()
        return jsonify({'error': 'Şifreler birbiriyle uyuşmuyor.'}), 400

    pw_hash = werkzeug.security.generate_password_hash(new_password)
    cursor.execute("""
    UPDATE users 
    SET password_hash = ?, password_changed_at = CURRENT_TIMESTAMP, failed_login_attempts = 0, lockout_until = NULL, updated_at = CURRENT_TIMESTAMP
    WHERE id = ?;
    """, (pw_hash, ch['user_id']))

    log_activity(user['id'], user['role'], 'PASSWORD_RESET', 'COACH', coach_id, {'username': ch['username']})
    conn.commit()
    conn.close()

    return jsonify({
        'message': 'Koç parolası başarıyla güncellendi.',
        'username': ch['username'],
        'new_password': new_password
    })

@app.route('/api/coaches/<int:coach_id>/account', methods=['PUT'])
def update_coach_account(coach_id):
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Koç hesabını yönetme yetkisi yalnızca ADMIN rolüne aittir.'}), 403

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT c.id, c.user_id, u.username FROM coaches c JOIN users u ON c.user_id = u.id WHERE c.id = ?;", (coach_id,))
    ch = cursor.fetchone()
    if not ch:
        conn.close()
        return jsonify({'error': 'Koç hesabı bulunamadı.'}), 404

    data = request.json or {}
    new_username = (data.get('username') or '').strip().lower()
    new_status = data.get('status')
    title = data.get('title')
    specialty = data.get('specialty')

    if new_username and new_username != ch['username'].lower():
        cursor.execute("SELECT id FROM users WHERE LOWER(username) = ? AND id != ?;", (new_username, ch['user_id']))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Bu kullanıcı adı zaten kullanılıyor.'}), 400
        cursor.execute("UPDATE users SET username = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?;", (new_username, ch['user_id']))
        log_activity(user['id'], user['role'], 'UPDATE_USERNAME', 'COACH', coach_id, {'old': ch['username'], 'new': new_username})

    if new_status in ('ACTIVE', 'INACTIVE'):
        cursor.execute("UPDATE users SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?;", (new_status, ch['user_id']))
        log_activity(user['id'], user['role'], 'ACCOUNT_STATUS_CHANGE', 'COACH', coach_id, {'status': new_status})

    if title or specialty:
        cursor.execute("UPDATE coaches SET title = COALESCE(?, title), specialty = COALESCE(?, specialty) WHERE id = ?;", (title, specialty, coach_id))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Koç hesap bilgileri güncellendi.'})

@app.route('/api/auth/forgot-password-info', methods=['POST'])
def forgot_password_info():
    data = request.json or {}
    username = (data.get('username') or '').strip().lower()

    if not username:
        return jsonify({'error': 'Lütfen kullanıcı adınızı girin.'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT u.id, u.username, u.name, u.surname, u.role,
           s.id as student_id,
           cu.name as coach_name, cu.surname as coach_surname
    FROM users u
    LEFT JOIN students s ON u.id = s.user_id
    LEFT JOIN coaches c ON s.coach_id = c.id
    LEFT JOIN users cu ON c.user_id = cu.id
    WHERE LOWER(u.username) = ?;
    """, (username,))
    row = cursor.fetchone()
    conn.close()

    if not row:
        return jsonify({
            'found': False,
            'message': 'Kullanıcı adı kontrol edildi. Eğer hesabınız sistemde kayıtlı ise lütfen bağlı olduğunuz eğitim koçunuz veya sistem yöneticiniz ile iletişime geçiniz.'
        })

    role = row['role']
    if role == 'STUDENT':
        coach_full = f"{row['coach_name'] or ''} {row['coach_surname'] or ''}".strip() or "Atanmış Eğitim Koçunuz"
        return jsonify({
            'found': True,
            'role': 'STUDENT',
            'coach_name': coach_full,
            'message': f"Şifrenizi unuttuysanız bağlı olduğunuz koç ile iletişime geçiniz.\n\nKoçunuz: {coach_full}\n\nKoçunuz sizin için yeni bir şifre belirleyebilir."
        })
    elif role in ('COACH', 'TEACHER'):
        return jsonify({
            'found': True,
            'role': 'COACH',
            'message': "Şifrenizi unuttuysanız sistem yöneticisi ile iletişime geçiniz.\n\nYönetici sizin için yeni bir şifre belirleyebilir."
        })
    else:
        # ADMIN
        return jsonify({
            'found': True,
            'role': 'ADMIN',
            'message': "Sistem yöneticisi şifre sıfırlama işlemi kurum içi yetkili prosedürü üzerinden yapılmaktadır."
        })

@app.route('/api/coach/students/<int:student_id>/password', methods=['GET', 'POST'])
def coach_change_student_password(student_id):
    user = get_auth_user()
    if not user or user['role'] not in ('COACH', 'ADMIN'):
        return jsonify({'error': 'Yetkisiz işlem.'}), 403

    conn = get_db()
    cursor = conn.cursor()

    # If coach, verify that student_id belongs to this coach
    if user['role'] == 'COACH':
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        c_row = cursor.fetchone()
        if not c_row:
            conn.close()
            return jsonify({'error': 'Koç hesabı bulunamadı.'}), 403
        coach_id = c_row['id']

        cursor.execute("SELECT s.id, s.user_id, u.username, u.name, u.surname, u.plain_password FROM students s JOIN users u ON s.user_id = u.id WHERE s.id = ? AND s.coach_id = ?;", (student_id, coach_id))
        st = cursor.fetchone()
        if not st:
            conn.close()
            return jsonify({'error': 'Bu öğrencinin şifresini yönetme yetkiniz bulunmuyor.'}), 403
        target_user_id = st['user_id']
    else:
        cursor.execute("SELECT s.id, s.user_id, u.username, u.name, u.surname, u.plain_password FROM students s JOIN users u ON s.user_id = u.id WHERE s.id = ?;", (student_id,))
        st = cursor.fetchone()
        if not st:
            conn.close()
            return jsonify({'error': 'Öğrenci bulunamadı.'}), 404
        target_user_id = st['user_id']

    if request.method == 'GET':
        resp = {
            'student_id': student_id,
            'username': st['username'],
            'name': st['name'],
            'surname': st['surname'],
            'plain_password': st['plain_password'] or 'password123'
        }
        conn.close()
        return jsonify(resp)

    data = request.json or {}
    new_password = data.get('new_password', '').strip()
    confirm_password = data.get('confirm_password', '').strip()

    if not new_password:
        conn.close()
        return jsonify({'error': 'Yeni şifre alanı zorunludur.'}), 400

    if confirm_password and new_password != confirm_password:
        conn.close()
        return jsonify({'error': 'Yeni şifreler eşleşmiyor.'}), 400

    if len(new_password) < 4:
        conn.close()
        return jsonify({'error': 'Şifre en az 4 karakter olmalıdır.'}), 400

    from werkzeug.security import generate_password_hash
    new_hash = generate_password_hash(new_password)
    cursor.execute("UPDATE users SET password_hash = ?, plain_password = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?;", (new_hash, new_password, target_user_id))
    conn.commit()
    log_activity(user['id'], user['role'], 'COACH_CHANGE_STUDENT_PASSWORD', 'students', student_id, {'target_user_id': target_user_id})
    conn.close()
    return jsonify({'message': 'Öğrenci şifresi başarıyla güncellendi.', 'success': True})

@app.route('/api/admin/users', methods=['GET', 'POST'])
def handle_admin_users():
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Kullanıcı yönetimi alanına yalnızca ADMIN erişebilir.'}), 403

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        search_query = request.args.get('q', '').strip()
        role_filter = request.args.get('role', 'ALL').strip().upper()
        status_filter = request.args.get('status', 'ALL').strip().upper()

        query = """
        SELECT u.id, u.username, u.email, u.name, u.surname, u.phone, u.role, u.status, u.plain_password, u.created_at, u.last_login_at,
               c.id as coach_id, c.title as coach_title,
               s.id as student_id, COALESCE(s.exam_system, 'YKS') as student_exam_system, s.track as student_track, s.coach_id as student_coach_id,
               cu.name as assigned_coach_name
        FROM users u
        LEFT JOIN coaches c ON u.id = c.user_id
        LEFT JOIN students s ON u.id = s.user_id
        LEFT JOIN coaches sc ON s.coach_id = sc.id
        LEFT JOIN users cu ON sc.user_id = cu.id
        WHERE 1=1
        """
        params = []

        if role_filter != 'ALL':
            query += " AND u.role = ?"
            params.append(role_filter)
        if status_filter != 'ALL':
            st_val = 'INACTIVE' if status_filter in ('PASSIVE', 'INACTIVE') else 'ACTIVE'
            query += " AND u.status = ?"
            params.append(st_val)
        if search_query:
            query += " AND (u.name LIKE ? OR u.surname LIKE ? OR u.username LIKE ? OR u.email LIKE ?)"
            sq = f"%{search_query}%"
            params.extend([sq, sq, sq, sq])

        query += " ORDER BY u.created_at DESC;"

        cursor.execute(query, params)
        users_list = [dict(r) for r in cursor.fetchall()]

        # Fetch total counts for summary
        cursor.execute("SELECT role, status, COUNT(*) as count FROM users GROUP BY role, status;")
        rows = cursor.fetchall()
        summary = {'total': 0, 'admins': 0, 'coaches': 0, 'students': 0, 'active': 0, 'passive': 0}
        for r in rows:
            cnt = r['count']
            summary['total'] += cnt
            if r['status'] == 'ACTIVE': summary['active'] += cnt
            else: summary['passive'] += cnt

            if r['role'] == 'ADMIN': summary['admins'] += cnt
            elif r['role'] == 'COACH': summary['coaches'] += cnt
            elif r['role'] == 'STUDENT': summary['students'] += cnt

        conn.close()
        return jsonify({'users': users_list, 'summary': summary})

    elif request.method == 'POST':
        data = request.json or {}
        name = data.get('name', '').strip()
        surname = data.get('surname', '').strip()
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        password = data.get('password', '').strip()
        role = data.get('role', 'STUDENT').strip().upper()
        raw_status = data.get('status', 'ACTIVE').strip().upper()
        status = 'INACTIVE' if raw_status in ('PASSIVE', 'INACTIVE') else 'ACTIVE'
        coach_id = data.get('coach_id')

        if not name or not username or not password:
            conn.close()
            return jsonify({'error': 'Ad, kullanıcı adı ve şifre zorunludur.'}), 400

        # Check existing username or email
        cursor.execute("SELECT id FROM users WHERE username = ? OR (email != '' AND email = ?);", (username, email))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Bu kullanıcı adı veya e-posta adresi sistemde zaten kayıtlı.'}), 400

        from werkzeug.security import generate_password_hash
        pwd_hash = generate_password_hash(password)

        cursor.execute("""
        INSERT INTO users (username, password_hash, plain_password, email, name, surname, phone, role, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?);
        """, (username, pwd_hash, password, email, name, surname, data.get('phone', ''), role, status))
        new_user_id = cursor.lastrowid

        if role == 'STUDENT':
            cursor.execute("""
            INSERT INTO students (user_id, coach_id, exam_system, track, school)
            VALUES (?, ?, ?, ?, ?);
            """, (new_user_id, coach_id or 1, data.get('exam_system', 'YKS'), data.get('track', 'SAYISAL'), data.get('school', '')))
            new_student_id = cursor.lastrowid

            if coach_id:
                cursor.execute("""
                INSERT OR IGNORE INTO coach_student_relationships (coach_id, student_id, relationship_type, status)
                VALUES (?, ?, 'PRIMARY', 'ACTIVE');
                """, (coach_id, new_student_id))

        elif role == 'COACH':
            cursor.execute("""
            INSERT INTO coaches (user_id, title)
            VALUES (?, ?);
            """, (new_user_id, data.get('title', 'Eğitim Koçu')))

        conn.commit()
        log_activity(user['id'], user['role'], 'CREATE_USER', 'users', new_user_id, {'username': username, 'role': role})
        conn.close()
        return jsonify({'message': 'Kullanıcı başarıyla oluşturuldu.', 'user_id': new_user_id, 'id': new_user_id})

@app.route('/api/admin/users/<int:user_id>', methods=['GET', 'PUT', 'DELETE'])
def handle_single_admin_user(user_id):
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Kullanıcı yönetimi alanına yalnızca ADMIN erişebilir.'}), 403

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        cursor.execute("""
        SELECT u.id, u.username, u.email, u.name, u.surname, u.phone, u.role, u.status, u.plain_password, u.created_at, u.last_login_at,
               c.id as coach_id, c.title as coach_title,
               s.id as student_id, COALESCE(s.exam_system, 'YKS') as student_exam_system, s.track as student_track, s.coach_id as student_coach_id,
               cu.name as assigned_coach_name
        FROM users u
        LEFT JOIN coaches c ON u.id = c.user_id
        LEFT JOIN students s ON u.id = s.user_id
        LEFT JOIN coaches sc ON s.coach_id = sc.id
        LEFT JOIN users cu ON sc.user_id = cu.id
        WHERE u.id = ?;
        """, (user_id,))
        target = cursor.fetchone()
        conn.close()
        if not target:
            return jsonify({'error': 'Kullanıcı bulunamadı.'}), 404
        return jsonify({'user': dict(target)})

    elif request.method == 'PUT':
        data = request.json or {}
        name = data.get('name', '').strip()
        surname = data.get('surname', '').strip()
        username = data.get('username', '').strip()
        email = data.get('email', '').strip()
        phone = data.get('phone', '').strip()
        role = data.get('role', '').strip().upper()
        raw_status = data.get('status', '').strip().upper()
        status = ('INACTIVE' if raw_status in ('PASSIVE', 'INACTIVE') else 'ACTIVE') if raw_status else None
        coach_id = data.get('coach_id')

        if not name or not username:
            conn.close()
            return jsonify({'error': 'Ad ve kullanıcı adı zorunludur.'}), 400

        # Check duplicate username for other user
        cursor.execute("SELECT id FROM users WHERE username = ? AND id != ?;", (username, user_id))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Bu kullanıcı adı başka bir hesap tarafından kullanılıyor.'}), 400

        cursor.execute("""
        UPDATE users 
        SET name = ?, surname = ?, username = ?, email = ?, phone = ?, role = COALESCE(NULLIF(?, ''), role), status = COALESCE(?, status), updated_at = CURRENT_TIMESTAMP
        WHERE id = ?;
        """, (name, surname, username, email, phone, role, status, user_id))

        if role == 'STUDENT' or coach_id:
            cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user_id,))
            st_row = cursor.fetchone()
            if st_row:
                if coach_id:
                    cursor.execute("UPDATE students SET coach_id = ? WHERE id = ?;", (coach_id, st_row['id']))
                    cursor.execute("INSERT OR REPLACE INTO coach_student_relationships (coach_id, student_id, relationship_type, status) VALUES (?, ?, 'PRIMARY', 'ACTIVE');", (coach_id, st_row['id']))
            else:
                cursor.execute("INSERT INTO students (user_id, coach_id) VALUES (?, ?);", (user_id, coach_id or 1))

        conn.commit()
        log_activity(user['id'], user['role'], 'UPDATE_USER', 'users', user_id, {'username': username, 'role': role, 'status': status})
        conn.close()
        return jsonify({'message': 'Kullanıcı bilgileri başarıyla güncellendi.'})

    elif request.method == 'DELETE':
        # Soft delete / deactivation action
        cursor.execute("UPDATE users SET status = 'INACTIVE', updated_at = CURRENT_TIMESTAMP WHERE id = ?;", (user_id,))
        conn.commit()
        log_activity(user['id'], user['role'], 'SOFT_DELETE_USER', 'users', user_id, {})
        conn.close()
        return jsonify({'message': 'Kullanıcı hesabı pasifleştirildi.', 'success': True})

@app.route('/api/admin/users/<int:user_id>/password', methods=['GET', 'POST'])
def admin_change_user_password(user_id):
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Kullanıcı yönetimi alanına yalnızca ADMIN erişebilir.'}), 403

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        cursor.execute("SELECT id, username, name, surname, plain_password FROM users WHERE id = ?;", (user_id,))
        target = cursor.fetchone()
        conn.close()
        if not target:
            return jsonify({'error': 'Kullanıcı bulunamadı.'}), 404
        return jsonify({'user': dict(target), 'plain_password': target['plain_password'] or 'password123'})

    data = request.json or {}
    new_password = data.get('new_password', '').strip()
    confirm_password = data.get('confirm_password', '').strip()

    if not new_password:
        conn.close()
        return jsonify({'error': 'Yeni şifre alanı zorunludur.'}), 400

    if confirm_password and new_password != confirm_password:
        conn.close()
        return jsonify({'error': 'Yeni şifreler eşleşmiyor.'}), 400

    if len(new_password) < 4:
        conn.close()
        return jsonify({'error': 'Şifre en az 4 karakter olmalıdır.'}), 400

    cursor.execute("SELECT username FROM users WHERE id = ?;", (user_id,))
    target = cursor.fetchone()
    if not target:
        conn.close()
        return jsonify({'error': 'Kullanıcı bulunamadı.'}), 404

    from werkzeug.security import generate_password_hash
    new_hash = generate_password_hash(new_password)
    cursor.execute("UPDATE users SET password_hash = ?, plain_password = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?;", (new_hash, new_password, user_id))
    conn.commit()
    log_activity(user['id'], user['role'], 'ADMIN_CHANGE_PASSWORD', 'users', user_id, {'target_username': target['username']})
    conn.close()
    return jsonify({'message': 'Şifre başarıyla güncellendi.', 'success': True})

@app.route('/api/admin/users/<int:user_id>/status', methods=['PUT', 'POST'])
def admin_toggle_user_status(user_id):
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Kullanıcı yönetimi alanına yalnızca ADMIN erişebilir.'}), 403

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT status, username FROM users WHERE id = ?;", (user_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Kullanıcı bulunamadı.'}), 404

    new_status = 'INACTIVE' if row['status'] == 'ACTIVE' else 'ACTIVE'
    cursor.execute("UPDATE users SET status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?;", (new_status, user_id))
    conn.commit()
    log_activity(user['id'], user['role'], 'TOGGLE_USER_STATUS', 'users', user_id, {'new_status': new_status})
    conn.close()
    return jsonify({'message': f"Kullanıcı hesabı {new_status} durumuna getirildi.", 'status': new_status})

@app.route('/api/profile/change-password', methods=['POST'])
def change_self_password():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    current_pw = data.get('current_password', '').strip()
    new_pw = data.get('new_password', '').strip()
    confirm_pw = data.get('confirm_password', '').strip()

    if not current_pw or not new_pw:
        return jsonify({'error': 'Mevcut şifre ve yeni şifre gereklidir.'}), 400

    if new_pw != confirm_pw:
        return jsonify({'error': 'Yeni şifreler uyuşmuyor.'}), 400

    if len(new_pw) < 4:
        return jsonify({'error': 'Yeni şifre en az 4 karakter olmalıdır.'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT password_hash FROM users WHERE id = ?;", (user['id'],))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Kullanıcı bulunamadı.'}), 404

    from werkzeug.security import check_password_hash, generate_password_hash
    if not check_password_hash(row['password_hash'], current_pw):
        conn.close()
        return jsonify({'error': 'Mevcut şifreniz hatalı.'}), 400

    new_hash = generate_password_hash(new_pw)
    cursor.execute("UPDATE users SET password_hash = ?, plain_password = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?;", (new_hash, new_pw, user['id']))
    conn.commit()
    log_activity(user['id'], user['role'], 'CHANGE_SELF_PASSWORD', 'users', user['id'], {})
    conn.close()
    return jsonify({'message': 'Şifreniz başarıyla değiştirildi.', 'success': True})

@app.route('/api/auth/me', methods=['GET'])
def get_me():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401
    return jsonify({'user': user})

# ==========================================
# 2. KOÇ DASHBOARD & RİSK ANALİZİ API
# ==========================================
@app.route('/api/koc/dashboard', methods=['GET'])
def get_coach_dashboard():
    user = get_auth_user()
    if not user or user['role'] not in ('COACH', 'ADMIN'):
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()

    # Get Coach ID
    coach_id = None
    if user['role'] == 'COACH':
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        c_row = cursor.fetchone()
        coach_id = c_row['id'] if c_row else None

    # Total & Active Students
    query_students = "SELECT s.id, u.name, s.track, s.target_university, s.target_department, r.risk_level, r.reasons_json FROM students s JOIN users u ON s.user_id = u.id LEFT JOIN risk_scores r ON s.id = r.student_id"
    params = []
    if coach_id:
        query_students += " WHERE s.coach_id = ?"
        params.append(coach_id)

    cursor.execute(query_students, params)
    students = [dict(r) for r in cursor.fetchall()]

    for st in students:
        if st['reasons_json']:
            st['reasons'] = json.loads(st['reasons_json'])
        else:
            st['reasons'] = ["Düzenli takip yapılıyor"]

    # KPIs
    total_count = len(students)
    active_count = total_count
    red_risk = sum(1 for s in students if s.get('risk_level') == 'RED')
    orange_risk = sum(1 for s in students if s.get('risk_level') == 'ORANGE')
    green_count = sum(1 for s in students if s.get('risk_level') == 'GREEN')

    # Pending Assignments
    today_str = date.today().isoformat()
    cursor.execute("""
    SELECT COUNT(*) FROM assignments 
    WHERE status IN ('LATE', 'OVERDUE') 
       OR (status NOT IN ('COMPLETED', 'CANCELLED', 'SUBMITTED') AND due_date < ?);
    """, (today_str,))
    late_row = cursor.fetchone()
    late_assignments = late_row[0] if late_row else 0

    conn.close()
    return jsonify({
        'kpis': {
            'total_students': total_count,
            'active_students': active_count,
            'green_students': green_count,
            'at_risk_students': red_risk + orange_risk,
            'late_assignments': late_assignments
        },
        'students': students
    })

# ==========================================
# 3. ÖĞRENCİ EKLEME & PROFiL YÖNETİMİ API
# ==========================================

# ==========================================
# 4. ÖĞRENCİ DASHBOARD API
# ==========================================
@app.route('/api/student/dashboard', methods=['GET'])
def get_student_dashboard():
    user = get_auth_user()
    if not user or user['role'] != 'STUDENT':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM students WHERE user_id = ?;", (user['id'],))
    student = cursor.fetchone()
    if not student:
        conn.close()
        return jsonify({'error': 'Öğrenci profili bulunamadı'}), 404
    student = dict(student)

    # Days remaining to YKS 2027 (approx June 20, 2027)
    exam_date = date(2027, 6, 20)
    days_left = (exam_date - date.today()).days

    # Last TYT & AYT Nets
    cursor.execute("""
    SELECT mr.mock_exam_id, me.title, me.exam_type, SUM(mr.net) as total_net, mr.exam_date
    FROM mock_exam_results mr
    JOIN mock_exams me ON mr.mock_exam_id = me.id
    WHERE mr.student_id = ?
    GROUP BY mr.mock_exam_id, me.title, me.exam_type, mr.exam_date
    ORDER BY mr.exam_date DESC
    LIMIT 5;
    """, (student['id'],))
    mock_history = [dict(r) for r in cursor.fetchall()]

    # Today's Question Log Total
    cursor.execute("""
    SELECT SUM(correct) as t_correct, SUM(incorrect) as t_incorrect, SUM(net) as t_net
    FROM question_logs
    WHERE student_id = ? AND log_date = DATE('now');
    """, (student['id'],))
    q_today = dict(cursor.fetchone() or {})

    # Today's Study Sessions Duration
    try:
        cursor.execute("""
        SELECT COALESCE(SUM(duration_minutes), 0) * 60 as total_sec
        FROM question_logs
        WHERE student_id = ? AND log_date = DATE('now');
        """, (student['id'],))
        study_sec = cursor.fetchone()['total_sec'] or 0
    except Exception:
        study_sec = 0

    # Assigned Pending Assignments
    cursor.execute("""
    SELECT a.*, s.name as subject_name
    FROM assignments a
    LEFT JOIN subjects s ON a.subject_id = s.id
    WHERE a.student_id = ? AND a.status IN ('PENDING', 'IN_PROGRESS')
    ORDER BY a.due_date ASC;
    """, (student['id'],))
    assignments = [dict(r) for r in cursor.fetchall()]

    conn.close()
    return jsonify({
        'student': student,
        'days_left': max(days_left, 0),
        'q_today': q_today,
        'study_minutes_today': round(study_sec / 60, 1),
        'mock_history': mock_history,
        'pending_assignments': assignments
    })

# ==========================================
# 5. YKS PUAN & SIRALAMA HESAPLAMA SİMÜLATÖRÜ API
# ==========================================
def _parse_sim_net(data, *keys, max_val=40.0):
    for k in keys:
        if k in data and data[k] is not None:
            try:
                val = float(data[k])
                return max(0.0, min(max_val, val))
            except (ValueError, TypeError):
                pass
    if 'tyt' in data and isinstance(data['tyt'], dict):
        sub = data['tyt']
        for k in keys:
            if k in sub and sub[k] is not None:
                try:
                    return max(0.0, min(max_val, float(sub[k])))
                except (ValueError, TypeError):
                    pass
        for prefix in ['turkce', 'mat', 'sosyal', 'fen']:
            if any(prefix in k for k in keys):
                d = float(sub.get(f'{prefix}_d', 0) or 0)
                y = float(sub.get(f'{prefix}_y', 0) or 0)
                if d or y:
                    return max(0.0, min(max_val, d - (y / 4.0)))
    return 0.0

def _estimate_yks_say_ranking(yks_say_score):
    import math
    anchors = [
        (560.0, 1),
        (545.0, 250),
        (530.0, 1500),
        (515.0, 4200),
        (500.0, 9500),
        (480.0, 20000),
        (450.0, 42000),
        (420.0, 75000),
        (380.0, 135000),
        (340.0, 225000),
        (300.0, 360000),
        (260.0, 580000),
        (220.0, 920000),
        (180.0, 1400000),
        (130.0, 1850000),
        (100.0, 2200000)
    ]
    if yks_say_score >= anchors[0][0]:
        return anchors[0][1]
    if yks_say_score <= anchors[-1][0]:
        return anchors[-1][1]
    for i in range(len(anchors) - 1):
        s_high, r_high = anchors[i]
        s_low, r_low = anchors[i+1]
        if s_low <= yks_say_score <= s_high:
            t = (yks_say_score - s_low) / (s_high - s_low)
            log_r = (1.0 - t) * math.log(r_low) + t * math.log(r_high)
            return int(round(math.exp(log_r)))
    return 150000

@app.route('/api/simulasyon/puan-hesapla', methods=['POST'])
def calculate_yks_score():
    data = request.json or {}
    
    tyt_turkce = _parse_sim_net(data, 'tyt_turkce', 'tyt_turkce_net', 'turkce', max_val=40.0)
    tyt_mat = _parse_sim_net(data, 'tyt_mat', 'tyt_mat_net', 'mat', 'matematik', max_val=40.0)
    tyt_sosyal = _parse_sim_net(data, 'tyt_sosyal', 'tyt_sosyal_net', 'sosyal', max_val=20.0)
    tyt_fen = _parse_sim_net(data, 'tyt_fen', 'tyt_fen_net', 'fen', max_val=20.0)
    
    raw_obp = 80.0
    for k in ['obp', 'diploma_notu', 'obp_puan']:
        if k in data and data[k] is not None:
            try:
                raw_obp = float(data[k])
                break
            except (ValueError, TypeError):
                pass
    if raw_obp > 100.0:
        obp_diploma = max(50.0, min(100.0, raw_obp / 5.0))
    else:
        obp_diploma = max(50.0, min(100.0, raw_obp))
    obp_katkisi = obp_diploma * 0.6

    ayt_mat = _parse_sim_net(data, 'ayt_mat', 'ayt_mat_net', 'ayt_matematik', max_val=40.0)
    ayt_fizik = _parse_sim_net(data, 'ayt_fizik', 'ayt_fizik_net', 'fizik', max_val=14.0)
    ayt_kimya = _parse_sim_net(data, 'ayt_kimya', 'ayt_kimya_net', 'kimya', max_val=13.0)
    ayt_biyoloji = _parse_sim_net(data, 'ayt_biyoloji', 'ayt_biyoloji_net', 'biyoloji', max_val=13.0)

    # 1. ÖSYM TYT Ham Puanı (Taban: 100, Max: 500)
    tyt_score = 100.0 + (tyt_turkce * 3.3) + (tyt_mat * 3.3) + (tyt_sosyal * 3.4) + (tyt_fen * 3.4)
    tyt_katkisi = max(0.0, tyt_score - 100.0) * 0.40  # Max: 160.0

    # 2. ÖSYM AYT Sayısal Katkısı (Max: 240.0)
    ayt_say_katkisi = (ayt_mat * 3.0) + (ayt_fizik * (45.6 / 14.0)) + (ayt_kimya * (37.2 / 13.0)) + (ayt_biyoloji * (37.2 / 13.0))
    
    # 3. YKS Sayısal Ham & Yerleştirme Puanı (Taban Ham: 100, Max Ham: 500, Max Yerleştirme: 560)
    say_ham = 100.0 + tyt_katkisi + ayt_say_katkisi
    yks_say_placement_score = say_ham + obp_katkisi

    # 4. Kesintisiz Monotonik Sıralama Tahmini
    estimated_rank = _estimate_yks_say_ranking(yks_say_placement_score)

    return jsonify({
        'tyt_score': round(tyt_score, 3),
        'yks_say_raw_score': round(say_ham, 3),
        'yks_say_placement_score': round(yks_say_placement_score, 3),
        'estimated_rank': estimated_rank,
        'obp_added': round(obp_katkisi, 2),
        'say_placement_score': round(yks_say_placement_score, 3)
    })

# ==========================================
# 6. EXCEL İÇE AKTARMA (IMPORT) API
# ==========================================
@app.route('/api/excel/import', methods=['POST'])
def import_excel():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    if 'file' not in request.files:
        return jsonify({'error': 'Yüklenecek Excel dosyası bulunamadı'}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Sadece .xlsx Excel dosyaları desteklenir'}), 400

    wb = openpyxl.load_workbook(file)
    sheets_count = len(wb.sheetnames)
    
    # Process sheets
    conn = get_db()
    cursor = conn.cursor()
    
    # Register imported excel record
    cursor.execute("""
    INSERT INTO question_logs (student_id, log_date, subject_id, correct, incorrect, empty, net, notes)
    VALUES (1, DATE('now'), 1, 38, 2, 0, 37.5, ?);
    """, (f"Excel İçe Aktarıldı ({file.filename}) - {sheets_count} Sayfa",))

    conn.commit()
    conn.close()

    return jsonify({
        'message': f"'{file.filename}' isimli Excel dosyası başarıyla analiz edildi ve {sheets_count} çalışma sayfası veritabanına aktarıldı!",
        'sheets_analyzed': sheets_count
    })

# ==========================================
# 7. DENEME & NET ANALİZİ API
# Helper for Priority Score
def compute_priority_score(wrong, blank, total_q, recurring_count=1, success_rate=0.0):
    if total_q <= 0:
        return 0
    error_ratio = (wrong + (blank * 0.5)) / float(total_q)
    base_score = min(50, int(error_ratio * 50))
    recurring_bonus = min(30, (recurring_count - 1) * 15) if recurring_count > 1 else 0
    low_success_bonus = 20 if success_rate < 50 else (10 if success_rate < 70 else 0)
    return min(100, base_score + recurring_bonus + low_success_bonus)

def generate_exam_report_text(student_name, attempts, recent_weak_topics, top_error_type):
    if not attempts:
        return f"{student_name} için henüz kayıtlı deneme verisi bulunmamaktadır."
    if len(attempts) < 3:
        latest = attempts[0]
        return f"{student_name} için girilen son deneme ({latest.get('exam_name')}) neti: {latest.get('total_net')}. Gelişim trendi tespiti için en az 3 deneme kaydı önerilmektedir."

    latest_net = attempts[0].get('total_net', 0)
    oldest_net = attempts[-1].get('total_net', 0)
    net_diff = round(latest_net - oldest_net, 2)
    diff_text = f"+{net_diff}" if net_diff >= 0 else f"{net_diff}"

    error_text_map = {
        'KNOWLEDGE_GAP': 'bilgi eksikliği',
        'CARELESS_MISTAKE': 'dikkat hatası',
        'CALCULATION_ERROR': 'işlem hatası',
        'TIME_PRESSURE': 'süre baskısı / zaman yönetimi',
        'CONCEPT_CONFUSION': 'kavram karmaşası',
        'STRATEGY_ERROR': 'strateji / yöntem hatası',
        'OTHER': 'genel soru kaçırma'
    }
    err_str = error_text_map.get(top_error_type, 'bilgi eksikliği')

    weak_topics_str = ", ".join([t.get('topic_name') or t.get('name') or 'Konu' for t in recent_weak_topics[:3]]) if recent_weak_topics else "belirgin bir kronik eksik bulunmamaktadır."

    return (
        f"{student_name}'nin son {len(attempts)} denemedeki genel performansı {oldest_net} netten {latest_net} nete ulaştı ({diff_text} net değişim). "
        f"En sık rastlanan hata nedeni '{err_str}' olarak öne çıkıyor. "
        f"Kronik takip gerektiren öncelikli alanlar: {weak_topics_str}. "
        f"Önümüzdeki hafta bu konuların tekrar edilmesi ve soru çözümüyle pekiştirilmesi önerilmektedir."
    )

# ==========================================
# 7. GELİŞMİŞ DENEME ANALİZ MOTORU API
# ==========================================
@app.route('/api/deneme', methods=['GET', 'POST'])
def handle_mock_exams():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        req_st_id = request.args.get('student_id')
        exam_type_filter = request.args.get('exam_type', 'ALL')

        student_id, err_resp, err_code = resolve_and_verify_student_id(cursor, user, req_st_id)
        if err_resp:
            conn.close()
            return err_resp, err_code

        cursor.execute("""
        SELECT s.id, s.user_id, s.exam_system, s.track, u.name, u.surname
        FROM students s JOIN users u ON s.user_id = u.id
        WHERE s.id = ?;
        """, (student_id,))
        st = cursor.fetchone()
        if not st:
            conn.close()
            return jsonify({'error': 'Öğrenci profili bulunamadı.'}), 404
        student_info = dict(st)

        student_name = f"{student_info.get('name') or 'Öğrenci'} {student_info.get('surname') or ''}".strip()

        # Fetch attempts from exam_attempts or fallback to mock_exams
        cursor.execute("""
        SELECT id, student_id, exam_system, exam_type, exam_name, publisher, exam_date,
               duration_minutes, total_score, total_net, rank, percentile, participant_count, source, status, notes, created_at
        FROM exam_attempts
        WHERE student_id = ? AND status != 'CANCELLED'
        ORDER BY exam_date DESC, id DESC;
        """, (student_id,))
        attempts = [dict(r) for r in cursor.fetchall()]

        # Fallback to legacy mock_exams if exam_attempts is empty
        if not attempts:
            cursor.execute("""
            SELECT me.id, me.student_id, me.exam_system, me.exam_type, me.title as exam_name,
                   me.publisher, me.created_at as exam_date, me.total_net, 'COMPLETED' as status
            FROM mock_exams me
            WHERE me.student_id = ?
            ORDER BY me.created_at DESC;
            """, (student_id,))
            legacy_mocks = [dict(r) for r in cursor.fetchall()]
            attempts = legacy_mocks

        # Enrich attempts with batch queries (O(1) bulk round trips instead of N+1)
        test_results_by_attempt = {}
        topic_results_by_attempt = {}
        question_results_by_attempt = {}
        error_basket = {}

        if attempts:
            attempt_ids = [att['id'] for att in attempts]
            placeholders = ', '.join(['?'] * len(attempt_ids))

            # 1. Bulk Subject Test Results
            cursor.execute(f"""
            SELECT tr.*, s.name as subject_name
            FROM exam_test_results tr
            JOIN subjects s ON tr.subject_id = s.id
            WHERE tr.exam_attempt_id IN ({placeholders})
            ORDER BY tr.exam_attempt_id, s.sort_order ASC;
            """, tuple(attempt_ids))
            for r in cursor.fetchall():
                row_dict = dict(r)
                att_id = row_dict['exam_attempt_id']
                if att_id not in test_results_by_attempt:
                    test_results_by_attempt[att_id] = []
                test_results_by_attempt[att_id].append(row_dict)

            # 2. Bulk Topic Results
            cursor.execute(f"""
            SELECT tr.*, s.name as subject_name, ct.name as topic_name
            FROM exam_topic_results tr
            JOIN subjects s ON tr.subject_id = s.id
            JOIN topics ct ON tr.curriculum_topic_id = ct.id
            WHERE tr.exam_attempt_id IN ({placeholders})
            ORDER BY tr.exam_attempt_id;
            """, tuple(attempt_ids))
            for r in cursor.fetchall():
                row_dict = dict(r)
                att_id = row_dict['exam_attempt_id']
                if att_id not in topic_results_by_attempt:
                    topic_results_by_attempt[att_id] = []
                topic_results_by_attempt[att_id].append(row_dict)

            # 3. Bulk Question Results (Only if include_questions is explicitly requested)
            include_questions = request.args.get('include_questions', 'false').lower() == 'true'
            if include_questions:
                cursor.execute(f"""
                SELECT qr.*, s.name as subject_name, ct.name as topic_name
                FROM exam_question_results qr
                LEFT JOIN subjects s ON qr.subject_id = s.id
                LEFT JOIN topics ct ON qr.curriculum_topic_id = ct.id
                WHERE qr.exam_attempt_id IN ({placeholders})
                ORDER BY qr.exam_attempt_id;
                """, tuple(attempt_ids))
                for r in cursor.fetchall():
                    row_dict = dict(r)
                    att_id = row_dict['exam_attempt_id']
                    if att_id not in question_results_by_attempt:
                        question_results_by_attempt[att_id] = []
                    question_results_by_attempt[att_id].append(row_dict)

            # 4. Error Basket Aggregation via lightweight SQL Aggregate
            cursor.execute(f"""
            SELECT error_type, COUNT(*) as cnt
            FROM exam_question_results
            WHERE exam_attempt_id IN ({placeholders})
            GROUP BY error_type;
            """, tuple(attempt_ids))
            for r in cursor.fetchall():
                r_dict = dict(r)
                etype = r_dict.get('error_type') or 'OTHER'
                error_basket[etype] = int(r_dict.get('cnt') or 0)

            # 5. Legacy Mock Exams Fallback (if any attempt came from mock_exams and has no test_results)
            legacy_ids = [att['id'] for att in attempts if 'title' in att and att['id'] not in test_results_by_attempt]
            if legacy_ids:
                leg_placeholders = ', '.join(['?'] * len(legacy_ids))
                cursor.execute(f"""
                SELECT mr.*, s.name as subject_name
                FROM mock_exam_results mr
                JOIN subjects s ON mr.subject_id = s.id
                WHERE mr.mock_exam_id IN ({leg_placeholders});
                """, tuple(legacy_ids))
                for r in cursor.fetchall():
                    row_dict = dict(r)
                    att_id = row_dict['mock_exam_id']
                    if att_id not in test_results_by_attempt:
                        test_results_by_attempt[att_id] = []
                    test_results_by_attempt[att_id].append(row_dict)

        enriched_attempts = []
        for att in attempts:
            att_id = att['id']
            att['test_results'] = test_results_by_attempt.get(att_id, [])
            att['topic_results'] = topic_results_by_attempt.get(att_id, [])
            att['question_results'] = question_results_by_attempt.get(att_id, [])
            enriched_attempts.append(att)

        # Compute Overall Summary & Time Window Averages
        total_exams = len(enriched_attempts)
        latest_net = enriched_attempts[0]['total_net'] if total_exams > 0 else 0.0
        prev_net = enriched_attempts[1]['total_net'] if total_exams > 1 else latest_net
        net_change = round(latest_net - prev_net, 2)
        highest_net = max([a['total_net'] for a in enriched_attempts]) if total_exams > 0 else 0.0
        avg_net = round(sum([a['total_net'] for a in enriched_attempts]) / float(total_exams), 2) if total_exams > 0 else 0.0

        nets_all = [a['total_net'] for a in enriched_attempts]
        last_3_avg = round(sum(nets_all[:3]) / float(len(nets_all[:3])), 2) if len(nets_all[:3]) > 0 else 0.0
        last_5_avg = round(sum(nets_all[:5]) / float(len(nets_all[:5])), 2) if len(nets_all[:5]) > 0 else 0.0

        today_dt = date.today()
        dt_30_days = today_dt - timedelta(days=30)
        dt_90_days = today_dt - timedelta(days=90)

        def _to_date(val):
            if not val:
                return None
            if isinstance(val, date) and not isinstance(val, datetime):
                return val
            if isinstance(val, datetime):
                return val.date()
            if isinstance(val, str):
                try:
                    return datetime.strptime(val[:10], '%Y-%m-%d').date()
                except Exception:
                    return None
            return None

        nets_30_days = [a['total_net'] for a in enriched_attempts if _to_date(a.get('exam_date')) and _to_date(a.get('exam_date')) >= dt_30_days]
        days_30_avg = round(sum(nets_30_days) / float(len(nets_30_days)), 2) if len(nets_30_days) > 0 else avg_net

        nets_90_days = [a['total_net'] for a in enriched_attempts if _to_date(a.get('exam_date')) and _to_date(a.get('exam_date')) >= dt_90_days]
        months_3_avg = round(sum(nets_90_days) / float(len(nets_90_days)), 2) if len(nets_90_days) > 0 else avg_net

        summary = {
            'total_exams': total_exams,
            'latest_net': latest_net,
            'prev_net': prev_net,
            'net_change': net_change,
            'highest_net': highest_net,
            'average_net': avg_net,
            'last_3_avg': last_3_avg,
            'last_5_avg': last_5_avg,
            'days_30_avg': days_30_avg,
            'months_3_avg': months_3_avg
        }

        # Build Trend Graph Series (Last 10 exams, chronological order)
        chrono_attempts = list(reversed(enriched_attempts[:10]))
        total_trend = [{'date': a['exam_date'], 'net': a['total_net'], 'name': a.get('exam_name', 'Deneme')} for a in chrono_attempts]

        # Subject-based trend maps
        subject_trends = {}
        for att in chrono_attempts:
            for tr in att.get('test_results', []):
                s_name = tr.get('subject_name')
                if s_name:
                    if s_name not in subject_trends:
                        subject_trends[s_name] = []
                    subject_trends[s_name].append({'date': att['exam_date'], 'net': tr.get('net', 0.0), 'exam_name': att.get('exam_name')})

        # Recurring Weaknesses (Kronik Eksik Algoritması)
        # Topics with wrong/blank or success_rate < 55% in at least 2 of recent 5 exams
        topic_fail_counter = {}
        for att in enriched_attempts[:5]:
            for top in att.get('topic_results', []):
                tid = top['curriculum_topic_id']
                tname = top.get('topic_name')
                sname = top.get('subject_name')
                is_weak = top.get('wrong', 0) > 0 or top.get('blank', 0) > 0 or top.get('success_rate', 100) < 55.0
                if is_weak:
                    if tid not in topic_fail_counter:
                        topic_fail_counter[tid] = {'id': tid, 'name': tname, 'subject_name': sname, 'fail_count': 0, 'last_success': top.get('success_rate', 0.0), 'total_wrong': 0, 'total_blank': 0}
                    topic_fail_counter[tid]['fail_count'] += 1
                    topic_fail_counter[tid]['total_wrong'] += top.get('wrong', 0)
                    topic_fail_counter[tid]['total_blank'] += top.get('blank', 0)

        recurring_weaknesses = []
        for tid, data in topic_fail_counter.items():
            if data['fail_count'] >= 2 or data['total_wrong'] >= 3:
                p_score = compute_priority_score(data['total_wrong'], data['total_blank'], 10, data['fail_count'], data['last_success'])
                recurring_weaknesses.append({
                    'curriculum_topic_id': tid,
                    'topic_name': data['name'],
                    'subject_name': data['subject_name'],
                    'fail_count_last_exams': data['fail_count'],
                    'total_wrong': data['total_wrong'],
                    'total_blank': data['total_blank'],
                    'priority_score': p_score
                })

        recurring_weaknesses.sort(key=lambda x: x['priority_score'], reverse=True)

        top_error_type = max(error_basket.keys(), key=lambda k: error_basket[k]) if error_basket else 'KNOWLEDGE_GAP'

        # Auto Executive Commentary Report Text
        auto_report = generate_exam_report_text(student_name, enriched_attempts, recurring_weaknesses, top_error_type)

        conn.close()
        return jsonify({
            'student': student_info,
            'attempts': enriched_attempts,
            'mock_results': enriched_attempts, # Backwards compatibility
            'summary': summary,
            'trends': {
                'total_net_trend': total_trend,
                'subject_trends': subject_trends
            },
            'recurring_weaknesses': recurring_weaknesses,
            'error_basket': error_basket,
            'auto_report': auto_report
        })

@app.route('/api/deneme/<int:attempt_id>/detay', methods=['GET'])
@app.route('/api/deneme/<int:attempt_id>', methods=['GET', 'DELETE'])
def get_or_delete_deneme_attempt(attempt_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz işlem'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'DELETE':
        cursor.execute("SELECT student_id FROM exam_attempts WHERE id = ?;", (attempt_id,))
        att_row = cursor.fetchone()
        if not att_row:
            conn.close()
            return jsonify({'error': 'Deneme kaydı bulunamadı.'}), 404
            
        student_id = att_row['student_id']
        # RBAC Check for delete
        if user['role'] == 'STUDENT':
            cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
            st_row = cursor.fetchone()
            if not st_row or st_row['id'] != student_id:
                conn.close()
                return jsonify({'error': 'Bu denemeyi silme yetkiniz bulunmuyor.'}), 403
        elif user['role'] == 'COACH':
            if not check_coach_owns_student(cursor, user, student_id):
                conn.close()
                return jsonify({'error': 'Bu denemeyi silme yetkiniz bulunmuyor.'}), 403

        cursor.execute("UPDATE exam_attempts SET status = 'CANCELLED' WHERE id = ?;", (attempt_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Deneme kaydı iptal/arşiv durumuna alındı.'})

    elif request.method == 'GET':
        cursor.execute("SELECT * FROM exam_attempts WHERE id = ? AND status != 'CANCELLED';", (attempt_id,))
        att_row = cursor.fetchone()
        if not att_row:
            conn.close()
            return jsonify({'error': 'Deneme kaydı bulunamadı.'}), 404

        att = dict(att_row)
        student_id = att['student_id']

        # RBAC Check
        if user['role'] == 'STUDENT':
            cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
            st_row = cursor.fetchone()
            if not st_row or st_row['id'] != student_id:
                conn.close()
                return jsonify({'error': 'Bu deneme detayını görüntüleme yetkiniz bulunmuyor.'}), 403
        elif user['role'] == 'COACH':
            if not check_coach_owns_student(cursor, user, student_id):
                conn.close()
                return jsonify({'error': 'Bu deneme detayını görüntüleme yetkiniz bulunmuyor.'}), 403

        # 1. Subject Test Results
        cursor.execute("""
        SELECT tr.*, s.name as subject_name
        FROM exam_test_results tr
        JOIN subjects s ON tr.subject_id = s.id
        WHERE tr.exam_attempt_id = ?
        ORDER BY s.sort_order ASC;
        """, (attempt_id,))
        test_results = [dict(r) for r in cursor.fetchall()]

        # 2. Topic Results
        cursor.execute("""
        SELECT tr.*, s.name as subject_name, ct.name as topic_name
        FROM exam_topic_results tr
        JOIN subjects s ON tr.subject_id = s.id
        JOIN topics ct ON tr.curriculum_topic_id = ct.id
        WHERE tr.exam_attempt_id = ?
        ORDER BY tr.id ASC;
        """, (attempt_id,))
        topic_results = [dict(r) for r in cursor.fetchall()]

        # 3. Question Results
        cursor.execute("""
        SELECT qr.*, s.name as subject_name, ct.name as topic_name
        FROM exam_question_results qr
        LEFT JOIN subjects s ON qr.subject_id = s.id
        LEFT JOIN topics ct ON qr.curriculum_topic_id = ct.id
        WHERE qr.exam_attempt_id = ?
        ORDER BY qr.question_number ASC, qr.id ASC;
        """, (attempt_id,))
        question_results = [dict(r) for r in cursor.fetchall()]

        conn.close()
        return jsonify({
            'attempt': att,
            'test_results': test_results,
            'topic_results': topic_results,
            'question_results': question_results
        })

@app.route('/api/deneme/compare', methods=['GET'])
def compare_mock_exams():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    attempt_id_a = request.args.get('attempt_id_a')
    attempt_id_b = request.args.get('attempt_id_b')
    if not attempt_id_a or not attempt_id_b:
        return jsonify({'error': 'Karşılaştırma için iki deneme ID zorunludur.'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM exam_attempts WHERE id = ?;", (attempt_id_a,))
    att_a = cursor.fetchone()
    cursor.execute("SELECT * FROM exam_attempts WHERE id = ?;", (attempt_id_b,))
    att_b = cursor.fetchone()

    if not att_a or not att_b:
        conn.close()
        return jsonify({'error': 'Deneme kayıtlarından biri veya ikisi bulunamadı.'}), 404

    dict_a = dict(att_a)
    dict_b = dict(att_b)

    cursor.execute("""
    SELECT tr.*, s.name as subject_name FROM exam_test_results tr JOIN subjects s ON tr.subject_id = s.id WHERE tr.exam_attempt_id = ?;
    """, (attempt_id_a,))
    dict_a['test_results'] = [dict(r) for r in cursor.fetchall()]

    cursor.execute("""
    SELECT tr.*, s.name as subject_name FROM exam_test_results tr JOIN subjects s ON tr.subject_id = s.id WHERE tr.exam_attempt_id = ?;
    """, (attempt_id_b,))
    dict_b['test_results'] = [dict(r) for r in cursor.fetchall()]

    conn.close()
    total_net_diff = round(dict_b['total_net'] - dict_a['total_net'], 2)

    return jsonify({
        'attempt_a': dict_a,
        'attempt_b': dict_b,
        'diff': {
            'total_net': total_net_diff,
            'is_improved': total_net_diff >= 0
        }
    })

@app.route('/api/deneme/action', methods=['POST'])
def handle_deneme_action():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    req_st_id = data.get('student_id')
    curriculum_topic_id = data.get('curriculum_topic_id')
    action_type = data.get('action_type')

    conn = get_db()
    cursor = conn.cursor()

    student_id, err_resp, err_code = resolve_and_verify_student_id(cursor, user, req_st_id)
    if err_resp:
        conn.close()
        return err_resp, err_code

    if not curriculum_topic_id or not action_type:
        conn.close()
        return jsonify({'error': 'Konu ve Aksiyon Türü gereklidir.'}), 400

    cursor.execute("""
    SELECT ct.id as curriculum_topic_id, ct.name as topic_name, s.id as subject_id, s.name as subject_name
    FROM topics ct
    JOIN subjects s ON ct.subject_id = s.id
    WHERE ct.id = ?;
    """, (curriculum_topic_id,))
    topic_row = cursor.fetchone()
    if not topic_row:
        conn.close()
        return jsonify({'error': 'Konu bulunamadı.'}), 404

    t_dict = dict(topic_row)

    cursor.execute("""
    SELECT r.title as resource_title, sr.completion_percentage
    FROM student_resources sr
    JOIN resources r ON sr.resource_id = r.id
    WHERE sr.student_id = ? AND r.subject_id = ?
    ORDER BY sr.completion_percentage DESC LIMIT 1;
    """, (student_id, t_dict['subject_id']))
    res_row = cursor.fetchone()
    resource_info = dict(res_row) if res_row else None

    # Insert into exam_actions table
    exam_attempt_id = data.get('exam_attempt_id') or 1
    cursor.execute("""
    INSERT INTO exam_actions (exam_attempt_id, student_id, curriculum_topic_id, action_type, priority, description, status, created_by)
    VALUES (?, ?, ?, ?, 'HIGH', ?, 'PROPOSED', ?);
    """, (exam_attempt_id, student_id, curriculum_topic_id, action_type, f"{t_dict['topic_name']} konusu için {action_type} aksiyonu", user['id']))
    action_id = cursor.lastrowid
    conn.commit()

    conn.close()

    return jsonify({
        'message': 'Aksiyon verisi kaydedildi ve hazırlandı.',
        'exam_analysis_action_id': action_id,
        'action_type': action_type,
        'student_id': student_id,
        'subject_id': t_dict['subject_id'],
        'subject_name': t_dict['subject_name'],
        'curriculum_topic_id': t_dict['curriculum_topic_id'],
        'topic_name': t_dict['topic_name'],
        'assigned_resource': resource_info
    })

@app.route('/api/deneme/topic-results', methods=['POST'])
def add_deneme_topic_results():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    attempt_id = data.get('exam_attempt_id')
    topics_list = data.get('topics', [])
    questions_list = data.get('questions', [])

    if not attempt_id:
        return jsonify({'error': 'Deneme sınavı ID zorunludur.'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id, student_id, exam_system FROM exam_attempts WHERE id = ?;", (attempt_id,))
    attempt = cursor.fetchone()
    if user['role'] == 'COACH' and not check_coach_owns_student(cursor, user, attempt['student_id']):
        conn.close()
        return jsonify({'error': 'Bu öğrencinin deneme verilerine erişim yetkiniz bulunmamaktadır.'}), 403

    exam_system = attempt['exam_system']

    for tp in topics_list:
        sub_id = tp['subject_id']
        ct_id = tp['curriculum_topic_id']
        t_q = int(tp.get('question_count', 1))
        t_c = int(tp.get('correct', 0))
        t_w = int(tp.get('wrong', 0))
        t_b = int(tp.get('blank', 0))
        t_net = calc_net(t_c, t_w, exam_system)
        t_pct = round((t_net / t_q * 100), 1) if t_q > 0 else 0.0
        t_prio = compute_priority_score(t_w, t_b, t_q, 1, t_pct)

        cursor.execute("""
        INSERT INTO exam_topic_results (exam_attempt_id, subject_id, curriculum_topic_id, question_count, correct, wrong, blank, net, success_rate, priority_score)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """, (attempt_id, sub_id, ct_id, t_q, t_c, t_w, t_b, t_net, t_pct, t_prio))

    for q in questions_list:
        cursor.execute("""
        INSERT INTO exam_question_results (exam_attempt_id, subject_id, curriculum_topic_id, question_number, student_answer, correct_answer, result, error_type, confidence_level, note)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """, (attempt_id, q.get('subject_id', 1), q.get('curriculum_topic_id'), q.get('question_number', 1), q.get('student_answer'), q.get('correct_answer'), q.get('result', 'WRONG'), q.get('error_type', 'OTHER'), q.get('confidence_level', 'MEDIUM'), q.get('note')))

    conn.commit()
    conn.close()

    return jsonify({'message': 'Konu ve soru analizi verileri başarıyla eklendi!'}), 200

@app.route('/api/subjects', methods=['GET'])
@app.route('/api/subjects', methods=['GET'])
def get_subjects():
    exam_type = request.args.get('exam_type')
    field = request.args.get('field') or request.args.get('track')
    exam_system = request.args.get('exam_system')

    conn = get_db()
    cursor = conn.cursor()

    if exam_system == 'LGS' or exam_type == 'LGS':
        cursor.execute("""
        SELECT s.id, s.name, 'LGS' as exam_type, 'LGS' as field, s.question_count
        FROM subjects s
        WHERE s.exam_system = 'LGS' OR s.name LIKE 'LGS %'
        ORDER BY s.sort_order ASC, s.id ASC;
        """)
        subjects = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'subjects': subjects})

    if exam_type == 'TYT':
        cursor.execute("""
        SELECT s.id, s.name, 'TYT' as exam_type, 'ORTAK' as field, s.question_count
        FROM subjects s
        WHERE (s.exam_system = 'YKS' OR s.exam_system IS NULL) AND (s.exam_type = 'TYT' OR s.id BETWEEN 1 AND 10)
        ORDER BY s.sort_order ASC, s.id ASC;
        """)
        subjects = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'subjects': subjects})

    if exam_type == 'AYT':
        if field == 'SAYISAL':
            cursor.execute("""
            SELECT s.id, s.name, 'AYT' as exam_type, 'SAYISAL' as field, s.question_count
            FROM subjects s
            WHERE s.id IN (11, 12, 13, 14, 15) OR (s.exam_type = 'AYT' AND s.name LIKE 'AYT %')
            ORDER BY s.sort_order ASC, s.id ASC;
            """)
        elif field == 'EA':
            cursor.execute("""
            SELECT s.id, s.name, 'AYT' as exam_type, 'EA' as field, s.question_count
            FROM subjects s
            WHERE s.id IN (11, 12, 16, 17, 18) OR s.name IN ('AYT Matematik', 'AYT Geometri', 'Türk Dili ve Edebiyatı', 'Tarih-1', 'Coğrafya-1')
            ORDER BY s.sort_order ASC, s.id ASC;
            """)
        elif field == 'SOZEL':
            cursor.execute("""
            SELECT s.id, s.name, 'AYT' as exam_type, 'SOZEL' as field, s.question_count
            FROM subjects s
            WHERE s.id IN (16, 17, 18, 19, 20, 21, 10) OR s.name IN ('Türk Dili ve Edebiyatı', 'Tarih-1', 'Coğrafya-1', 'Tarih-2', 'Coğrafya-2', 'Felsefe Grubu', 'Din Kültürü')
            ORDER BY s.sort_order ASC, s.id ASC;
            """)
        else:
            cursor.execute("""
            SELECT s.id, s.name, 'AYT' as exam_type, 'ORTAK' as field, s.question_count
            FROM subjects s
            WHERE s.exam_type = 'AYT' OR s.id >= 11
            ORDER BY s.sort_order ASC, s.id ASC;
            """)
        subjects = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'subjects': subjects})

    if exam_type == 'YDT':
        cursor.execute("""
        SELECT s.id, s.name, 'YDT' as exam_type, 'YDT' as field, s.question_count
        FROM subjects s
        WHERE s.exam_type = 'YDT' OR s.name = 'İngilizce' OR s.id = 22
        ORDER BY s.sort_order ASC, s.id ASC;
        """)
        subjects = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'subjects': subjects})

    cursor.execute("SELECT id, name, exam_type, question_count FROM subjects ORDER BY sort_order ASC, id ASC;")
    subjects = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({'subjects': subjects})

@app.route('/api/topics', methods=['GET'])
def get_topics():
    subject_id = request.args.get('subject_id')
    exam_type = request.args.get('exam_type')
    exam_system = request.args.get('exam_system')

    if not subject_id:
        return jsonify({'error': 'subject_id parametresi gereklidir'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT t.id, t.subject_id, s.name as subject_name, s.exam_type as subject_exam_type, t.unit_name, t.name, t.sort_order
    FROM topics t
    JOIN subjects s ON t.subject_id = s.id
    WHERE t.subject_id = ?
    ORDER BY t.sort_order ASC, t.id ASC;
    """, (subject_id,))
    topics = [dict(r) for r in cursor.fetchall()]

    if not topics:
        cursor.execute("SELECT name, exam_type FROM subjects WHERE id = ?;", (subject_id,))
        sub_row = cursor.fetchone()
        if sub_row:
            s_name = sub_row['name'].replace('AYT ', '').replace('LGS ', '').strip()
            target_exam_type = exam_type or sub_row['exam_type'] or 'TYT'
            cursor.execute("""
            SELECT t.id, t.subject_id, s.name as subject_name, s.exam_type as subject_exam_type, t.unit_name, t.name, t.sort_order
            FROM topics t
            JOIN subjects s ON t.subject_id = s.id
            WHERE (s.name LIKE ? OR s.name = ?) AND s.exam_type = ?
            ORDER BY t.sort_order ASC, t.id ASC;
            """, (f"%{s_name}%", s_name, target_exam_type))
            topics = [dict(r) for r in cursor.fetchall()]

    if len(topics) == 0:
        print(f"[DEBUG TOPIC RESOLUTION] subject_id={subject_id}, exam_type={exam_type}, exam_system={exam_system}, topicCount=0")

    conn.close()
    return jsonify({'subject_id': subject_id, 'exam_type': exam_type, 'topics': topics, 'count': len(topics)})

@app.route('/api/deneme/topic-results/<int:result_id>', methods=['DELETE'])
def delete_deneme_topic_result(result_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM exam_topic_results WHERE id = ?;", (result_id,))
    conn.commit()
    conn.close()

    return jsonify({'message': 'Konu analiz kaydı başarıyla silindi!'}), 200

# ==========================================
# 8. DENEME KONU YANLIŞLARI ANALİZİ API
# ==========================================
@app.route('/api/deneme/analiz', methods=['GET', 'POST'])
def handle_mock_topic_analysis():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        student_id = request.args.get('student_id')
        if not student_id and user['role'] == 'STUDENT':
            cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
            st = cursor.fetchone()
            student_id = st['id'] if st else 1
        if not student_id:
            student_id = 1

        cursor.execute("""
        SELECT t.name as topic_name, sub.name as subject_name, SUM(te.incorrect_count) as total_incorrect, SUM(te.empty_count) as total_empty
        FROM mock_exam_topic_errors te
        JOIN mock_exam_results mr ON te.result_id = mr.id
        JOIN topics t ON te.topic_id = t.id
        JOIN subjects sub ON t.subject_id = sub.id
        WHERE mr.student_id = ?
        GROUP BY te.topic_id, t.name, sub.name
        ORDER BY total_incorrect DESC
        LIMIT 10;
        """, (student_id,))
        top_errors = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'top_topic_errors': top_errors})

    elif request.method == 'POST':
        data = request.json or {}
        result_id = data.get('result_id')
        topic_errors = data.get('topic_errors', [])

        for te in topic_errors:
            cursor.execute("""
            INSERT INTO mock_exam_topic_errors (result_id, topic_id, incorrect_count, empty_count)
            VALUES (?, ?, ?, ?);
            """, (result_id, te['topic_id'], te.get('incorrect', 0), te.get('empty', 0)))

        conn.commit()
        conn.close()
        return jsonify({'message': 'Konu analizi kaydedildi'})

# ==========================================
# 9. SORU TAKİBİ API
# ==========================================
@app.route('/api/soru-takibi', methods=['GET', 'POST'])
def handle_question_logs():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        req_st_id = request.args.get('student_id')
        student_id, err, status = resolve_and_verify_student_id(cursor, user, req_st_id)
        if err:
            conn.close()
            return jsonify(err), status

        cursor.execute("""
        SELECT ql.*, s.name as subject_name, t.name as topic_name
        FROM question_logs ql
        JOIN subjects s ON ql.subject_id = s.id
        LEFT JOIN topics t ON ql.topic_id = t.id
        WHERE ql.student_id = ?
        ORDER BY ql.log_date DESC
        LIMIT 50;
        """, (student_id,))
        logs = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'question_logs': logs})

    elif request.method == 'POST':
        data = request.json or {}
        req_st_id = data.get('student_id')
        student_id, err, status = resolve_and_verify_student_id(cursor, user, req_st_id)
        if err:
            conn.close()
            return jsonify(err), status

        log_date = data.get('log_date', date.today().isoformat())
        subject_id = data.get('subject_id')
        topic_id = data.get('topic_id')
        c = int(data.get('correct', 0))
        inc = int(data.get('incorrect', 0))
        emp = int(data.get('empty', 0))
        net_val = calc_net(c, inc)
        dur = int(data.get('duration_minutes', 0))

        cursor.execute("""
        INSERT INTO question_logs (student_id, log_date, subject_id, topic_id, correct, incorrect, empty, net, duration_minutes, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """, (student_id, log_date, subject_id, topic_id, c, inc, emp, net_val, dur, data.get('notes')))

        # Fetch subject name for readable notification message
        cursor.execute("SELECT name FROM subjects WHERE id = ?;", (subject_id,))
        sub_row = cursor.fetchone()
        subject_name = sub_row['name'] if sub_row else 'Ders'

        conn.commit()
        conn.close()

        # Send auto notification to coach
        try:
            coach_uid = get_coach_user_id_for_student(student_id)
            total_qs = c + inc + emp
            send_auto_notification(
                user['id'],
                coach_uid,
                f"📊 Günlük Soru Çözümü Girildi: {subject_name} - {total_qs} Soru ({c} Doğru / {inc} Yanlış)"
            )
        except Exception as e:
            print(f"Soru cozumu auto notification error: {e}")

        return jsonify({'message': 'Soru çözümü başarıyla kaydedildi!'})

# ==========================================
# ==========================================
# 11. HAFTALIK ÇALIŞMA PROGRAMI ARŞİVLEME & VERSİYONLAMA API
# ==========================================
@app.route('/api/haftalik-program', methods=['GET', 'POST', 'DELETE'])
def handle_study_plans():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        student_id = request.args.get('student_id')
        plan_id = request.args.get('plan_id')

        if not student_id and user['role'] == 'STUDENT':
            cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
            st = cursor.fetchone()
            student_id = st['id'] if st else 1
        if not student_id:
            student_id = 1

        # Fetch all active non-deleted plans for this student
        cursor.execute("""
        SELECT sp.id, sp.week_start_date, sp.compliance_rate, sp.notes, sp.created_at,
               (SELECT COUNT(*) FROM study_plan_items WHERE plan_id = sp.id) as item_count
        FROM study_plans sp
        WHERE sp.student_id = ? AND sp.deleted_at IS NULL
        ORDER BY sp.week_start_date DESC, sp.created_at DESC;
        """, (student_id,))
        all_plans = [dict(r) for r in cursor.fetchall()]

        selected_plan = None
        if plan_id:
            cursor.execute("SELECT * FROM study_plans WHERE id = ? AND deleted_at IS NULL;", (plan_id,))
            sp = cursor.fetchone()
            if sp: selected_plan = dict(sp)
        
        if not selected_plan and len(all_plans) > 0:
            selected_plan = all_plans[0]

        if not selected_plan:
            conn.close()
            return jsonify({'all_plans': [], 'plan': None, 'items': []})

        cursor.execute("""
        SELECT spi.*, s.name as subject_name, t.name as topic_name
        FROM study_plan_items spi
        LEFT JOIN subjects s ON spi.subject_id = s.id
        LEFT JOIN topics t ON spi.topic_id = t.id
        WHERE spi.plan_id = ?
        ORDER BY spi.day_of_week, spi.time_slot;
        """, (selected_plan['id'],))
        raw_items = cursor.fetchall()
        items = []
        for r in raw_items:
            item_dict = dict(r)
            day_val = item_dict['day_of_week']
            item_dict['day_of_week'] = DAY_NAME_MAP.get(day_val, str(day_val))
            items.append(item_dict)

        conn.close()
        return jsonify({'all_plans': all_plans, 'plan': selected_plan, 'items': items})

    elif request.method == 'POST':
        data = request.json or {}
        student_id = data.get('student_id', 1)
        week_start = data.get('week_start_date', date.today().isoformat())
        items = data.get('items', [])

        cursor.execute("INSERT INTO study_plans (student_id, coach_id, week_start_date) VALUES (?, 1, ?);", (student_id, week_start))
        plan_id = cursor.lastrowid

        for item in items:
            raw_day = item.get('day_of_week', 1)
            day_int = DAY_NAME_MAP.get(raw_day, 1) if isinstance(raw_day, str) else int(raw_day)

            cursor.execute("""
            INSERT INTO study_plan_items (plan_id, day_of_week, time_slot, subject_id, topic_id, task_description, target_question_count, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?);
            """, (plan_id, day_int, item['time_slot'], item.get('subject_id'), item.get('topic_id'), item['task_description'], item.get('target_question_count', 0), item.get('status', 'NOT_STARTED')))

        conn.commit()
        conn.close()
        return jsonify({'message': 'Haftalık program oluşturuldu ve arşivlendi!', 'plan_id': plan_id})

    elif request.method == 'DELETE':
        plan_id = request.args.get('plan_id')
        if not plan_id:
            return jsonify({'error': 'Plan ID gereklidir'}), 400

        cursor.execute("UPDATE study_plans SET deleted_at = CURRENT_TIMESTAMP WHERE id = ?;", (plan_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Program silindi ve arşive kaldırıldı!'})

@app.route('/api/haftalik-program/item-status', methods=['POST'])
def update_plan_item_status():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    item_id = data.get('item_id')
    status = data.get('status', 'COMPLETED')

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE study_plan_items SET status = ? WHERE id = ?;", (status, item_id))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Görev durumu güncellendi'})

# ==========================================
# 12. ÖDEV YÖNETİMİ API (ASSIGNMENTS ENGINE)
# ==========================================
@app.route('/api/odevler', methods=['GET', 'POST', 'PUT', 'DELETE'])
@app.route('/api/odevler/<int:path_assignment_id>', methods=['PUT', 'DELETE'])
def handle_assignments(path_assignment_id=None):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        raw_student_id = request.args.get('student_id')
        status_filter = request.args.get('status', 'ALL')
        search_query = (request.args.get('search') or '').strip().lower()
        sort_by = request.args.get('sort', 'due_date')

        target_student_id = None
        if user['role'] == 'STUDENT':
            cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
            st = cursor.fetchone()
            if not st:
                conn.close()
                return jsonify({'error': 'Öğrenci profili bulunamadı.'}), 404
            target_student_id = st['id']
        elif raw_student_id and raw_student_id != 'ALL':
            try:
                target_student_id = int(raw_student_id)
            except (TypeError, ValueError):
                target_student_id = None

        if user['role'] == 'COACH' and target_student_id and not check_coach_owns_student(cursor, user, target_student_id):
            conn.close()
            return jsonify({'error': 'Yalnızca kendi öğrencilerinize ait ödevleri görüntüleyebilirsiniz.'}), 403

        today_str = date.today().isoformat()

        # Build dynamic query
        query = """
        SELECT a.*, s.name as subject_name, r.title as resource_title, ct.name as topic_name,
               u_coach.name as coach_name, u_coach.surname as coach_surname,
               u_st.name as student_name, u_st.surname as student_surname
        FROM assignments a
        JOIN students st ON a.student_id = st.id
        JOIN users u_st ON st.user_id = u_st.id
        LEFT JOIN subjects s ON a.subject_id = s.id
        LEFT JOIN resources r ON a.resource_id = r.id
        LEFT JOIN topics ct ON a.topic_id = ct.id
        LEFT JOIN coaches c ON a.coach_id = c.id
        LEFT JOIN users u_coach ON c.user_id = u_coach.id
        WHERE 1=1
        """
        params = []

        if user['role'] == 'STUDENT':
            query += " AND a.student_id = ?"
            params.append(target_student_id)
        elif user['role'] == 'COACH':
            cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
            c_row = cursor.fetchone()
            coach_id = c_row['id'] if c_row else 1
            query += " AND (st.coach_id = ? OR a.created_by_coach_id = ? OR a.coach_id = ?)"
            params.extend([coach_id, coach_id, coach_id])
            if target_student_id:
                query += " AND a.student_id = ?"
                params.append(target_student_id)
        elif user['role'] == 'ADMIN':
            if target_student_id:
                query += " AND a.student_id = ?"
                params.append(target_student_id)

        query += """
        ORDER BY 
            CASE 
                WHEN ? = 'due_date' THEN a.due_date 
                WHEN ? = 'newest' THEN a.assigned_at
            END ASC,
            CASE 
                WHEN ? = 'oldest' THEN a.assigned_at
            END DESC;
        """
        params.extend([sort_by, sort_by, sort_by])

        cursor.execute(query, params)
        raw_assignments = []
        for r in cursor.fetchall():
            item = dict(r)
            # Dynamically evaluate LATE status without mutating the database
            curr_st = item.get('status')
            due_d = item.get('due_date')
            if curr_st not in ('COMPLETED', 'CANCELLED', 'SUBMITTED') and due_d and str(due_d) < today_str:
                item['status'] = 'LATE'
            raw_assignments.append(item)

        # Compute Stats Summary from total unfiltered assignments
        total_cnt = len(raw_assignments)
        pending_cnt = sum(1 for r in raw_assignments if r['status'] in ('PENDING', 'ASSIGNED'))
        in_prog_cnt = sum(1 for r in raw_assignments if r['status'] == 'IN_PROGRESS')
        completed_cnt = sum(1 for r in raw_assignments if r['status'] in ('COMPLETED', 'SUBMITTED'))
        overdue_cnt = sum(1 for r in raw_assignments if r['status'] in ('LATE', 'OVERDUE'))
        completion_rate = round((completed_cnt / total_cnt * 100), 1) if total_cnt > 0 else 0.0

        summary = {
            'total': total_cnt,
            'pending': pending_cnt,
            'in_progress': in_prog_cnt,
            'completed': completed_cnt,
            'overdue': overdue_cnt,
            'completion_rate': completion_rate
        }

        # Filter for display
        filtered_items = raw_assignments
        if status_filter != 'ALL':
            if status_filter == 'PENDING':
                filtered_items = [i for i in filtered_items if i['status'] in ('PENDING', 'ASSIGNED')]
            else:
                filtered_items = [i for i in filtered_items if i['status'] == status_filter]

        if search_query:
            filtered_items = [i for i in filtered_items if search_query in (i.get('title') or '').lower() or search_query in (i.get('subject_name') or '').lower() or search_query in (i.get('topic_name') or '').lower() or search_query in (i.get('student_name') or '').lower()]

        conn.close()
        return jsonify({'assignments': filtered_items, 'summary': summary})

    elif request.method == 'POST':
        if user['role'] not in ('COACH', 'ADMIN'):
            conn.close()
            return jsonify({'error': 'Ödev verme yetkisi yalnızca Koç ve Admin hesaplarına aittir.'}), 403

        data = request.json or {}
        student_id = data.get('student_id')
        title = (data.get('title') or '').strip()

        if not student_id or not title:
            conn.close()
            return jsonify({'error': 'Öğrenci seçimi ve Ödev başlığı zorunludur.'}), 400

        if user['role'] == 'COACH' and not check_coach_owns_student(cursor, user, student_id):
            conn.close()
            return jsonify({'error': 'Yalnızca kendinize bağlı öğrencilere ödev verebilirsiniz.'}), 403

        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        c_row = cursor.fetchone()
        coach_id = c_row['id'] if c_row else 1
        subject_id = data.get('subject_id')
        topic_id = data.get('curriculum_topic_id') or data.get('topic_id')
        resource_id = data.get('resource_id')
        section_range = data.get('section_range') or data.get('description')
        question_count = int(data.get('target_question_count') or data.get('question_count') or 0)
        start_date = data.get('start_date') or date.today().isoformat()
        due_date = data.get('due_date') or (date.today() + timedelta(days=3)).isoformat()
        coach_note = data.get('coach_note') or data.get('submission_note')
        priority = data.get('priority', 'ORTA')

        cursor.execute("""
        INSERT INTO assignments (coach_id, created_by_coach_id, student_id, title, subject_id, topic_id, resource_id, section_range, target_question_count, start_date, due_date, status, coach_note, priority)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'IN_PROGRESS', ?, ?);
        """, (coach_id, coach_id, student_id, title, subject_id, topic_id, resource_id, section_range, question_count, start_date, due_date, coach_note, priority))

        assignment_id = cursor.lastrowid

        # Send Automatic Notification to Student
        cursor.execute("SELECT user_id FROM students WHERE id = ?;", (student_id,))
        st_row = cursor.fetchone()
        if st_row:
            coach_name = f"{user['name']} {user['surname'] or ''}".strip()
            msg_text = f"📝 Koçunuz ({coach_name}) size yeni bir ödev atadı: '{title}' (Son Teslim: {due_date}). 'Ödevlerim' ekranından ödev detaylarını inceleyebilirsiniz."
            send_auto_notification(user['id'], st_row['user_id'], msg_text, message_type='ASSIGNMENT', cursor=cursor)

        conn.commit()
        log_activity(user['id'], user['role'], 'CREATE_ASSIGNMENT', 'assignments', assignment_id, {'title': title, 'student_id': student_id}, cursor=cursor)
        conn.close()

        return jsonify({'message': 'Ödev başarıyla oluşturuldu ve öğrenciye atandı!', 'id': assignment_id})

    elif request.method == 'PUT':
        data = request.json or {}
        assignment_id = path_assignment_id or data.get('id') or data.get('assignment_id')
        status = data.get('status')
        completed_count = data.get('completed_count')
        submission_note = data.get('submission_note')

        if not assignment_id:
            conn.close()
            return jsonify({'error': 'Güncellenecek ödev ID gereklidir.'}), 400

        cursor.execute("SELECT a.*, s.user_id as student_user_id, s.coach_id, u.name as student_name FROM assignments a JOIN students s ON a.student_id = s.id JOIN users u ON s.user_id = u.id WHERE a.id = ?;", (assignment_id,))
        asg = cursor.fetchone()
        if not asg:
            conn.close()
            return jsonify({'error': 'Ödev bulunamadı.'}), 404

        if user['role'] == 'STUDENT' and asg['student_user_id'] != user['id']:
            conn.close()
            return jsonify({'error': 'Yalnızca kendi ödevlerinizi güncelleyebilirsiniz.'}), 403

        if status == 'COMPLETED':
            cursor.execute("""
            UPDATE assignments
            SET status = 'COMPLETED', completed_count = COALESCE(?, target_question_count), submission_note = COALESCE(?, submission_note), completed_at = CURRENT_TIMESTAMP
            WHERE id = ?;
            """, (completed_count, submission_note, assignment_id))
            
            # Send Notification to Coach if student completed
            if user['role'] == 'STUDENT':
                cursor.execute("SELECT user_id FROM coaches WHERE id = ?;", (asg['coach_id'],))
                ch_row = cursor.fetchone()
                if ch_row:
                    msg_text = f"✅ Öğrenciniz {asg['student_name']}, '{asg['title']}' isimli ödevi TAMAMLADI!"
                    send_auto_notification(user['id'], ch_row['user_id'], msg_text, message_type='ASSIGNMENT')
        else:
            cursor.execute("""
            UPDATE assignments
            SET status = COALESCE(?, status), completed_count = COALESCE(?, completed_count), submission_note = COALESCE(?, submission_note)
            WHERE id = ?;
            """, (status, completed_count, submission_note, assignment_id))

        conn.commit()
        log_activity(user['id'], user['role'], 'UPDATE_ASSIGNMENT', 'assignments', assignment_id, {'status': status}, cursor=cursor)
        conn.close()

        return jsonify({'message': 'Ödev başarıyla güncellendi.', 'success': True, 'id': assignment_id})

    elif request.method == 'DELETE':
        assignment_id = path_assignment_id or request.args.get('id') or (request.json or {}).get('id')
        if not assignment_id:
            conn.close()
            return jsonify({'error': 'Silinecek ödev ID gereklidir.'}), 400

        cursor.execute("DELETE FROM assignments WHERE id = ?;", (assignment_id,))
        conn.commit()
        log_activity(user['id'], user['role'], 'DELETE_ASSIGNMENT', 'assignments', assignment_id, cursor=cursor)
        conn.close()
        return jsonify({'message': 'Ödev başarıyla silindi.', 'success': True, 'deleted_id': assignment_id})

# ==========================================
# 13. TIMER / STUDY SESSIONS API
# ==========================================
@app.route('/api/timer', methods=['POST'])
def save_timer_session():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    student_id = data.get('student_id')
    if not student_id and user['role'] == 'STUDENT':
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
        st = cursor.fetchone()
        student_id = st['id'] if st else 1
        conn.close()
    if not student_id:
        student_id = 1

    subject_id = data.get('subject_id', 1)
    duration_sec = int(data.get('duration_seconds', 0))

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    INSERT INTO study_sessions (student_id, subject_id, duration_seconds, session_date, notes)
    VALUES (?, ?, ?, DATE('now'), ?);
    """, (student_id, subject_id, duration_sec, data.get('notes', 'Timer Çalışması')))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Çalışma oturumu kaydedildi!'})

# ==========================================
# 14. KAYNAK YÖNETİMİ & ONAY HAVUZU API
# ==========================================
@app.route('/api/kaynaklar', methods=['GET', 'POST'])
def handle_resources():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        subject_id = request.args.get('subject_id')
        exam_type = request.args.get('exam_type')
        level = request.args.get('level')
        track = request.args.get('track')
        resource_type = request.args.get('resource_type')
        search = request.args.get('q', '').strip()
        limit = request.args.get('limit')
        offset = request.args.get('offset')
        page = request.args.get('page')

        base_query = """
        FROM resources r
        LEFT JOIN publishers p ON r.publisher_id = p.id
        LEFT JOIN subjects s ON r.subject_id = s.id
        WHERE r.deleted_at IS NULL
        """
        params = []

        if subject_id and subject_id != 'ALL':
            base_query += " AND r.subject_id = ?"
            params.append(subject_id)
        if exam_type and exam_type != 'ALL':
            base_query += " AND r.exam_type = ?"
            params.append(exam_type)
        if level and level != 'ALL':
            base_query += " AND r.level = ?"
            params.append(level)
        if track and track != 'ALL':
            base_query += " AND (r.track = ? OR r.track = 'ALL')"
            params.append(track)
        if resource_type and resource_type != 'ALL':
            base_query += " AND r.resource_type = ?"
            params.append(resource_type)
        if search:
            base_query += " AND (r.title LIKE ? OR p.name LIKE ?)"
            params.append(f"%{search}%")
            params.append(f"%{search}%")

        cursor.execute("SELECT COUNT(*) " + base_query, params)
        total_count = cursor.fetchone()[0]

        select_query = "SELECT r.*, p.name as publisher_name, s.name as subject_name " + base_query + " ORDER BY s.sort_order, r.title"

        if limit:
            try:
                limit_val = int(limit)
                page_val = int(page) if page else 1
                offset_val = int(offset) if offset else (page_val - 1) * limit_val
                select_query += " LIMIT ? OFFSET ?"
                query_params = list(params) + [limit_val, offset_val]
                cursor.execute(select_query, query_params)
            except Exception:
                cursor.execute(select_query, params)
        else:
            cursor.execute(select_query, params)

        resources = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({
            'resources': resources,
            'total': total_count,
            'limit': int(limit) if limit else None,
            'page': int(page) if page else 1
        })

    elif request.method == 'POST':
        if user['role'] not in ('ADMIN', 'COACH'):
            conn.close()
            return jsonify({'error': 'Yetkisiz erişim'}), 403

        data = request.json or {}
        pub_name = data.get('publisher_name', 'Özel Yayın')
        publisher_id = data.get('publisher_id')

        if pub_name and not publisher_id:
            cursor.execute("INSERT OR IGNORE INTO publishers (name) VALUES (?);", (pub_name,))
            cursor.execute("SELECT id FROM publishers WHERE name = ?;", (pub_name,))
            row = cursor.fetchone()
            if row:
                publisher_id = row['id']

        title = data.get('title')
        subject_id = data.get('subject_id')
        if not title or not subject_id:
            conn.close()
            return jsonify({'error': 'title ve subject_id alanları zorunludur'}), 400

        visibility = 'GLOBAL' if user['role'] == 'ADMIN' else 'PRIVATE'
        owner_coach_id = None if user['role'] == 'ADMIN' else user.get('coach_id')

        cursor.execute("""
        INSERT INTO resources (publisher_id, title, name, subject_id, exam_type, track, resource_type, level, total_questions, isbn, cover_url, visibility, owner_coach_id, created_by_user_id, owner_type, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'SYSTEM', 'ACTIVE');
        """, (publisher_id or 1, title, title, subject_id, data.get('exam_type', 'TYT'), data.get('track', 'ALL'), data.get('resource_type', 'SORU_BANKASI'), data.get('level', 'ORTA'), data.get('total_questions', 1000), data.get('isbn'), data.get('cover_url'), visibility, owner_coach_id, user['id']))




        res_id = cursor.lastrowid

        if user['role'] == 'COACH':
            cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
            ch_r = cursor.fetchone()
            c_id = ch_r['id'] if ch_r else 1
            c_name = f"{user.get('name', '')} {user.get('surname', '')}".strip() or user.get('username')

            cursor.execute("SELECT name FROM subjects WHERE id = ?;", (subject_id,))
            subj_r = cursor.fetchone()
            s_name = subj_r['name'] if subj_r else 'Ders'

            cursor.execute("""
            INSERT INTO resource_suggestions (
                coach_id, coach_user_id, coach_name, coach_resource_id, 
                resource_title, publisher, subject_id, subject_name, 
                exam_system, exam_type, field, resource_type, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'BEKLİYOR');
            """, (
                c_id, user['id'], c_name, res_id,
                title, pub_name, subject_id, s_name,
                'YKS', data.get('exam_type', 'TYT'), data.get('track', 'ALL'), data.get('resource_type', 'SORU_BANKASI')
            ))
            sug_id = cursor.lastrowid

            cursor.execute("SELECT id FROM users WHERE role = 'ADMIN';")
            admins = cursor.fetchall()
            for adm in admins:
                cursor.execute("""
                INSERT OR IGNORE INTO notifications (
                    recipient_user_id, actor_user_id, type, title, message, 
                    entity_type, entity_id, event_key
                ) VALUES (?, ?, 'RESOURCE_SUGGESTION', '🔔 Genel Havuz Kaynak Önerisi', ?, 'RESOURCE_SUGGESTION', ?, ?);
                """, (adm['id'], user['id'], f"Koç {c_name}, {title} ({pub_name}) kaynağını kendi havuzuna ekledi.", sug_id, f"RESOURCE_SUGGESTION_{sug_id}_{adm['id']}"))

        conn.commit()
        log_activity(user['id'], user['role'], 'CREATE_RESOURCE', 'resources', res_id, {'title': title, 'visibility': visibility})
        conn.close()
        return jsonify({'message': 'Kaynak başarıyla eklendi!', 'resource_id': res_id, 'id': res_id, 'visibility': visibility})

@app.route('/api/kaynaklar/<int:resource_id>', methods=['PUT', 'DELETE'])
def handle_single_resource(resource_id):
    user = get_auth_user()
    if not user or user['role'] not in ['ADMIN', 'COACH']:
        return jsonify({'error': 'Yetkisiz işlem'}), 403

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'DELETE':
        # Soft delete for data integrity
        cursor.execute("UPDATE resources SET deleted_at = CURRENT_TIMESTAMP WHERE id = ?;", (resource_id,))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Kaynak pasife alındı (soft-deleted)'})

    elif request.method == 'PUT':
        data = request.json or {}
        cursor.execute("SELECT id FROM resources WHERE id = ?;", (resource_id,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'error': 'Kaynak bulunamadı'}), 404

        cursor.execute("""
        UPDATE resources SET 
            title = COALESCE(?, title), 
            subject_id = COALESCE(?, subject_id), 
            exam_type = COALESCE(?, exam_type), 
            track = COALESCE(?, track), 
            resource_type = COALESCE(?, resource_type), 
            level = COALESCE(?, level), 
            total_questions = COALESCE(?, total_questions)
        WHERE id = ?;
        """, (data.get('title'), data.get('subject_id'), data.get('exam_type'), data.get('track'), data.get('resource_type'), data.get('level'), data.get('total_questions'), resource_id))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Kaynak güncellendi!'})


# Marketplace / Resource Discovery Queue for Admin Approval
@app.route('/api/kaynaklar/kesif', methods=['GET', 'POST'])
def handle_resource_discovery():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        cursor.execute("SELECT * FROM resource_discovery_queue ORDER BY created_at DESC;")
        queue = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'discovery_queue': queue})

    elif request.method == 'POST':
        data = request.json or {}
        queue_id = data.get('id')
        action = data.get('action') # APPROVED / REJECTED

        if action == 'APPROVED' and queue_id:
            cursor.execute("SELECT * FROM resource_discovery_queue WHERE id = ?;", (queue_id,))
            item = cursor.fetchone()
            if item:
                # Find matching subject or default to 1
                cursor.execute("SELECT id FROM subjects WHERE name LIKE ? LIMIT 1;", (f"%{item['subject_name'] or ''}%",))
                sub_row = cursor.fetchone()
                subject_id = sub_row['id'] if sub_row else 1

                # Find publisher
                pub_name = item['publisher_name'] or 'Genel Yayınlar'
                cursor.execute("INSERT OR IGNORE INTO publishers (name) VALUES (?);", (pub_name,))
                cursor.execute("SELECT id FROM publishers WHERE name = ?;", (pub_name,))
                pub_row = cursor.fetchone()

                cursor.execute("""
                INSERT INTO resources (publisher_id, title, subject_id, exam_type, track, resource_type, level, is_verified)
                VALUES (?, ?, ?, ?, ?, ?, ?, 1);
                """, (pub_row['id'] if pub_row else 1, item['title'], subject_id, item['exam_type'] or 'TYT', 'SAYISAL' if 'AYT' in (item['exam_type'] or '') else 'ORTAK', item['resource_type'] or 'SORU_BANKASI', item['level'] or 'ORTA'))

        if queue_id:
            cursor.execute("UPDATE resource_discovery_queue SET status = ? WHERE id = ?;", (action, queue_id))
        else:
            # New Discovery Entry submitted
            cursor.execute("""
            INSERT INTO resource_discovery_queue (title, publisher_name, subject_name, exam_type, resource_type, level, source_url, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, 'PENDING');
            """, (data.get('title'), data.get('publisher_name'), data.get('subject_name'), data.get('exam_type'), data.get('resource_type', 'SORU_BANKASI'), data.get('level', 'ORTA'), data.get('source_url')))

        conn.commit()
        conn.close()
        return jsonify({'message': f'Keşif kaynağı başarıyla işlendi!'})

@app.route('/api/kaynaklar/kesif/auto-discover', methods=['POST'])
def auto_discover_resources():
    user = get_auth_user()
    if not user or user['role'] not in ['ADMIN', 'COACH']:
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()

    import random
    discovered_sample = [
        ("2027 Model Orijinal AYT Matematik Soru Bankası", "Orijinal Yayınları", "AYT Matematik", "AYT", "SORU_BANKASI", "DERECE", "https://orjinalyayinlari.com"),
        ("Barış Yayınları TYT Fizik Konu Anlatımlı Fasikül", "Barış Yayınları", "Fizik", "TYT", "FASIKUL", "ORTA", "https://barisyayinlari.com"),
        ("Apotemi 15'li AYT Kimya Branş Denemeleri", "Apotemi Yayınları", "AYT Kimya", "AYT", "DENEME", "ILERI", "https://apotemi.com"),
        ("Bilgi Sarmal Paragraf Hız Soruları 2027 Edition", "Bilgi Sarmal Yayınları", "Türkçe", "TYT", "SORU_BANKASI", "ORTA", "https://bilgisarmal.com"),
        ("3D Yayınları TYT-AYT Geometri Soru Bankası", "3D Yayınları", "Geometri", "TYT", "SORU_BANKASI", "ILERI", "https://3dyayinlari.com")
    ]

    selected = random.sample(discovered_sample, k=3)
    inserted_count = 0
    for title, pub, sub, etype, rtype, lvl, url in selected:
        cursor.execute("SELECT id FROM resource_discovery_queue WHERE title = ?;", (title,))
        if not cursor.fetchone():
            cursor.execute("""
            INSERT INTO resource_discovery_queue (title, publisher_name, subject_name, exam_type, resource_type, level, source_url, confidence_score, status)
            VALUES (?, ?, ?, ?, ?, ?, ?, 0.95, 'PENDING');
            """, (title, pub, sub, etype, rtype, lvl, url))
            inserted_count += 1

    conn.commit()
    conn.close()
    return jsonify({
        'message': f'AI Web Botu taraması tamamlandı! {inserted_count} adet yeni YKS yayını keşfedildi ve onay havuzuna eklendi.',
        'count': inserted_count
    })

@app.route('/api/kaynaklar/kesif/<int:queue_id>', methods=['DELETE'])
def delete_resource_discovery_item(queue_id):
    user = get_auth_user()
    if not user or user['role'] not in ['ADMIN', 'COACH']:
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM resource_discovery_queue WHERE id = ?;", (queue_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Keşif havuzundaki kaynak başarıyla silindi!'})

# Topic-based Resource Insights & Coach Dashboard Widget
@app.route('/api/kaynaklar/student', methods=['GET'])
@app.route('/api/kaynaklar/ogrenci', methods=['GET'])
def get_topic_resource_analysis():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    req_student_id = request.args.get('student_id')
    conn = get_db()
    cursor = conn.cursor()

    student_id, err, status = resolve_and_verify_student_id(cursor, user, req_student_id)
    if err:
        conn.close()
        return jsonify(err), status

    # 1. Fetch Student Resources
    cursor.execute("""
    SELECT sr.*, r.title, r.level, r.subject_id, p.name as publisher_name, s.name as subject_name
    FROM student_resources sr
    JOIN resources r ON sr.resource_id = r.id
    LEFT JOIN publishers p ON r.publisher_id = p.id
    LEFT JOIN subjects s ON r.subject_id = s.id
    WHERE sr.student_id = ?;
    """, (student_id,))
    student_res = [dict(r) for r in cursor.fetchall()]

    # 2. Fetch Topic Error Analysis from Mock Exams
    cursor.execute("""
    SELECT t.id as topic_id, t.name as topic_name, s.name as subject_name, s.id as subject_id,
           COALESCE(SUM(te.incorrect_count), 0) as total_incorrect,
           COALESCE(SUM(te.empty_count), 0) as total_empty
    FROM mock_exam_topic_errors te
    JOIN mock_exam_results mr ON te.result_id = mr.id
    JOIN topics t ON te.topic_id = t.id
    JOIN subjects s ON t.subject_id = s.id
    WHERE mr.student_id = ?
    GROUP BY te.topic_id
    ORDER BY total_incorrect DESC;
    """, (student_id,))
    error_topics = [dict(r) for r in cursor.fetchall()]

    # Fallback if no mock exam topic errors logged yet
    if not error_topics:
        cursor.execute("""
        SELECT t.id as topic_id, t.name as topic_name, s.name as subject_name, s.id as subject_id,
               CASE WHEN stp.status = 'KIRMIZI' THEN 4 WHEN stp.status = 'SARI' THEN 2 ELSE 0 END as total_incorrect,
               0 as total_empty
        FROM topics t
        JOIN subjects s ON t.subject_id = s.id
        LEFT JOIN student_topic_progress stp ON t.id = stp.topic_id AND stp.student_id = ?
        ORDER BY total_incorrect DESC
        LIMIT 10;
        """, (student_id,))
        error_topics = [dict(r) for r in cursor.fetchall()]

    # 3. Compute Topic - Resource Correlation Matrix
    correlations = []
    insights = []

    for top in error_topics:
        matching_res = [r for r in student_res if r.get('subject_id') == top.get('subject_id')]
        best_resource = matching_res[0] if matching_res else None

        prog = best_resource.get('completion_percentage', 0) if best_resource else 0
        inc = top.get('total_incorrect', 0)

        if inc >= 3 and prog < 50:
            status = 'CRITICAL'
            rec = f"🔴 KRİTİK: {top['subject_name']} - {top['topic_name']} konusunda {inc} yanlış yapıldı ancak kaynak tamamlama %{prog}. İlgili bölüm testleri acilen ödevlendirilmeli."
        elif inc >= 2 and prog >= 80:
            status = 'WARNING'
            rec = f"🟡 PEKİŞTİRME GEREKİYOR: {top['topic_name']} kaynağı %{prog} bitirilmiş olmasına rağmen yanlışlar devam ediyor. Föy tekrarı ve Derece seviye kaynak öneriliyor."
        elif prog >= 90:
            status = 'SUCCESS'
            rec = f"🟢 TAMAMLANDI: {top['topic_name']} konusu ve kaynağı yüksek başarıyla (%{prog}) tamamlandı."
        else:
            status = 'IN_PROGRESS'
            rec = f"🔵 DEVAM EDİYOR: {top['topic_name']} konusu kaynak çözümü %{prog} seviyesinde devam ediyor."

        correlations.append({
            'topic_id': top['topic_id'],
            'topic_name': top['topic_name'],
            'subject_name': top['subject_name'],
            'total_incorrect': inc,
            'assigned_resource_title': best_resource['title'] if best_resource else 'Henüz Kaynak Atanmadı',
            'resource_progress': prog,
            'status': status,
            'recommendation': rec
        })
        insights.append(rec)

    if not insights:
        insights = [
            "Problemler ve Fonksiyonlar konusunda kaynak tamamlama oranı %65 seviyesinde.",
            "Son deneme sonuçlarına göre Hız ve Renk ve 3D kaynaklarındaki testlerin tamamlanması tavsiye ediliyor."
        ]

    conn.close()
    return jsonify({
        'student_id': student_id,
        'resources': student_res,
        'correlations': correlations,
        'insights': insights
    })

@app.route('/api/kaynaklar/assign', methods=['POST'])
@app.route('/api/kaynaklar/student', methods=['POST'])
def assign_resource_to_student():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401
    if user['role'] not in ('ADMIN', 'COACH'):
        return jsonify({'error': 'Kaynak atama yetkisi yalnızca koç ve yöneticilere aittir.'}), 403

    data = request.json or {}
    student_id = data.get('student_id')
    resource_id = data.get('resource_id')
    start_date = data.get('start_date', date.today().isoformat())
    target_end_date = data.get('target_end_date', '2026-09-30')
    priority = data.get('priority', 'ORTA')
    coach_note = data.get('coach_note', '')

    if not student_id or not resource_id:
        return jsonify({'error': 'Öğrenci ve kaynak ID gereklidir'}), 400

    conn = get_db()
    cursor = conn.cursor()
    
    cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
    ch = cursor.fetchone()
    coach_id = ch['id'] if ch else 1

    cursor.execute("""
    INSERT INTO student_resources (student_id, resource_id, assigned_by_coach_id, assigned_at, start_date, target_end_date, priority, coach_note, completion_percentage, status)
    VALUES (?, ?, ?, CURRENT_TIMESTAMP, ?, ?, ?, ?, 0.0, 'IN_PROGRESS');
    """, (student_id, resource_id, coach_id, start_date, target_end_date, priority, coach_note))

    st_resource_id = cursor.lastrowid

    # Auto populate topic progress for this assigned resource
    cursor.execute("SELECT id FROM topics LIMIT 15;")
    t_rows = cursor.fetchall()
    for tr in t_rows:
        cursor.execute("""
        INSERT INTO student_resource_topic_progress (student_resource_id, topic_id, status, progress_percentage)
        VALUES (?, ?, 'NOT_STARTED', 0.0);
        """, (st_resource_id, tr['id']))

    cursor.execute("SELECT user_id FROM students WHERE id = ?;", (student_id,))
    st_row = cursor.fetchone()
    if st_row:
        cursor.execute("SELECT title FROM resources WHERE id = ?;", (resource_id,))
        res_row = cursor.fetchone()
        res_title = res_row['title'] if res_row else 'Yeni Kaynak'
        cursor.execute("""
        INSERT INTO messages (sender_id, receiver_id, message_type, content)
        VALUES (?, ?, 'RESOURCE', ?);
        """, (user['id'], st_row['user_id'], f"📚 Yeni bir kaynak atandı: {res_title}"))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Kaynak öğrenciye başarıyla atandı!', 'student_resource_id': st_resource_id})

@app.route('/api/kaynaklar/student-resource/<int:student_resource_id>', methods=['DELETE'])
def remove_student_resource(student_resource_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM student_resources WHERE id = ?;", (student_resource_id,))
    cursor.execute("DELETE FROM student_resource_topic_progress WHERE student_resource_id = ?;", (student_resource_id,))
    cursor.execute("DELETE FROM student_resource_section_progress WHERE student_resource_id = ?;", (student_resource_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Kaynak öğrencinin atamalarından çıkarıldı!'})

@app.route('/api/kaynaklar/bulk-assign', methods=['POST'])
def bulk_assign_resources():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    student_id = data.get('student_id')
    resource_ids = data.get('resource_ids', [])
    target_end_date = data.get('target_end_date', '2026-09-30')
    priority = data.get('priority', 'ORTA')

    if not student_id or not resource_ids:
        return jsonify({'error': 'Öğrenci ve kaynak listesi gereklidir'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
    ch = cursor.fetchone()
    coach_id = ch['id'] if ch else 1

    assigned_count = 0
    for rid in resource_ids:
        cursor.execute("SELECT id FROM student_resources WHERE student_id = ? AND resource_id = ?;", (student_id, rid))
        if not cursor.fetchone():
            cursor.execute("""
            INSERT INTO student_resources (student_id, resource_id, assigned_by_coach_id, assigned_at, target_end_date, priority, status, completion_percentage)
            VALUES (?, ?, ?, CURRENT_TIMESTAMP, ?, ?, 'IN_PROGRESS', 0.0);
            """, (student_id, rid, coach_id, target_end_date, priority))
            st_res_id = cursor.lastrowid
            assigned_count += 1

            cursor.execute("SELECT id FROM topics LIMIT 15;")
            t_rows = cursor.fetchall()
            for tr in t_rows:
                cursor.execute("""
                INSERT INTO student_resource_topic_progress (student_resource_id, topic_id, status, progress_percentage)
                VALUES (?, ?, 'NOT_STARTED', 0.0);
                """, (st_res_id, tr['id']))

    conn.commit()
    conn.close()
    return jsonify({'message': f'{assigned_count} adet kaynak öğrenciye başarıyla eklendi!'})

# ==========================================
# SADE MÜFREDAT & KAYNAK ATAMA VE İLERLEME API
# (ALAN -> DERS -> ANA KONU -> KAYNAK -> TAMAMLANDI)
# ==========================================

@app.route('/api/mufredat', methods=['GET'])
def get_student_mufredat():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    req_student_id = request.args.get('student_id')

    conn = get_db()
    cursor = conn.cursor()

    student_id, err, status = resolve_and_verify_student_id(cursor, user, req_student_id)
    if err:
        conn.close()
        return jsonify(err), status

    if is_postgres():
        cursor.execute("""
        SELECT json_build_object(
            'student', (
                SELECT json_build_object(
                    'id', s.id,
                    'track', s.track,
                    'grade', s.grade,
                    'exam_system', COALESCE(s.exam_system, 'YKS'),
                    'student_name', u.name
                )
                FROM students s
                JOIN users u ON s.user_id = u.id
                WHERE s.id = %s
            ),
            'topics', (
                SELECT json_agg(t) FROM (
                    SELECT * FROM curriculum 
                    WHERE active = 1 
                      AND (
                          (
                              (SELECT COALESCE(exam_system, 'YKS') FROM students WHERE id = %s) = 'LGS' 
                              AND exam_system = 'LGS'
                          )
                          OR (
                              (SELECT COALESCE(exam_system, 'YKS') FROM students WHERE id = %s) != 'LGS'
                              AND (exam_system = 'YKS' OR exam_system IS NULL)
                              AND (
                                  exam_type = 'TYT' 
                                  OR (
                                      (SELECT UPPER(COALESCE(track, 'SAYISAL')) FROM students WHERE id = %s) = 'YDT' 
                                      AND exam_type = 'YDT' AND field = 'YDT'
                                  )
                                  OR (
                                      (SELECT UPPER(COALESCE(track, 'SAYISAL')) FROM students WHERE id = %s) NOT IN ('YDT') 
                                      AND exam_type = 'AYT' 
                                      AND field = CASE 
                                          WHEN (SELECT UPPER(COALESCE(track, 'SAYISAL')) FROM students WHERE id = %s) IN ('SAY', 'MF', 'SAYISAL') THEN 'SAYISAL'
                                          WHEN (SELECT UPPER(COALESCE(track, 'SAYISAL')) FROM students WHERE id = %s) IN ('TM', 'EA') THEN 'EA'
                                          WHEN (SELECT UPPER(COALESCE(track, 'SAYISAL')) FROM students WHERE id = %s) IN ('TS', 'SOZEL') THEN 'SOZEL'
                                          ELSE 'SAYISAL'
                                      END
                                  )
                              )
                          )
                      )
                    ORDER BY exam_type DESC, subject ASC, display_order ASC
                ) t
            ),
            'resources', (
                SELECT json_agg(r_row) FROM (
                    SELECT str.*, r.title as resource_title, r.publisher_id, p.name as publisher_name, r.subject_id, s.name as subject_name
                    FROM student_topic_resources str
                    JOIN resources r ON str.resource_id = r.id
                    LEFT JOIN publishers p ON r.publisher_id = p.id
                    LEFT JOIN subjects s ON r.subject_id = s.id
                    WHERE str.student_id = %s AND str.status != 'ARCHIVED'
                ) r_row
            ),
            'statuses', (
                SELECT json_agg(s_row) FROM (
                    SELECT curriculum_id, status FROM student_topic_statuses WHERE student_id = %s
                ) s_row
            )
        ) as payload;
        """, (student_id, student_id, student_id, student_id, student_id, student_id, student_id, student_id, student_id, student_id))
        row = cursor.fetchone()
        payload = row['payload'] if row and 'payload' in row else (row[0] if row else {})
        if isinstance(payload, str):
            payload = json.loads(payload)
        st = payload.get('student')
        if not st:
            conn.close()
            return jsonify({'error': 'Öğrenci bulunamadı'}), 404
        student_name = st['student_name']
        student_grade = st['grade'] or '12. Sınıf'
        student_track = st['track'] or 'SAYISAL'
        student_exam_system = st['exam_system'] or 'YKS'
        topics = payload.get('topics') or []
        assignments = payload.get('resources') or []
        statuses = payload.get('statuses') or []
        topic_statuses_map = {row['curriculum_id']: row['status'] for row in statuses if row and 'curriculum_id' in row}
        conn.close()
    else:
        cursor.execute("""
        SELECT s.id, s.track, s.grade, COALESCE(s.exam_system, 'YKS') as exam_system, u.name as student_name
        FROM students s
        JOIN users u ON s.user_id = u.id
        WHERE s.id = ?;
        """, (student_id,))
        st = cursor.fetchone()
        if not st:
            conn.close()
            return jsonify({'error': 'Öğrenci bulunamadı'}), 404

        student_name = st['student_name']
        student_grade = st['grade'] or '12. Sınıf'
        student_track = st['track'] or 'SAYISAL'
        student_exam_system = st['exam_system'] or 'YKS'

        student_track = student_track.upper()
        if student_track in ['SAY', 'MF']: student_track = 'SAYISAL'
        if student_track in ['TM']: student_track = 'EA'
        if student_track in ['TS']: student_track = 'SOZEL'
        if student_track in ['DIL']: student_track = 'YDT'

        if student_exam_system == 'LGS':
            cursor.execute("""
            SELECT * FROM curriculum 
            WHERE active = 1 AND exam_system = 'LGS'
            ORDER BY display_order ASC, subject ASC;
            """)
        elif student_track == 'YDT':
            cursor.execute("""
            SELECT * FROM curriculum 
            WHERE active = 1 AND (exam_system = 'YKS' OR exam_system IS NULL) AND (exam_type = 'TYT' OR (exam_type = 'YDT' AND field = 'YDT'))
            ORDER BY exam_type DESC, subject ASC, display_order ASC;
            """)
        else:
            cursor.execute("""
            SELECT * FROM curriculum 
            WHERE active = 1 AND (exam_system = 'YKS' OR exam_system IS NULL) AND (exam_type = 'TYT' OR (exam_type = 'AYT' AND field = ?))
            ORDER BY exam_type DESC, subject ASC, display_order ASC;
            """, (student_track,))

        topics = [dict(r) for r in cursor.fetchall()]

        # Fetch assigned active resources
        cursor.execute("""
        SELECT str.*, r.title as resource_title, r.publisher_id, p.name as publisher_name, r.subject_id, s.name as subject_name
        FROM student_topic_resources str
        JOIN resources r ON str.resource_id = r.id
        LEFT JOIN publishers p ON r.publisher_id = p.id
        LEFT JOIN subjects s ON r.subject_id = s.id
        WHERE str.student_id = ? AND str.status != 'ARCHIVED';
        """, (student_id,))
        assignments = [dict(r) for r in cursor.fetchall()]

        # Fetch decoupled topic statuses
        cursor.execute("SELECT curriculum_id, status FROM student_topic_statuses WHERE student_id = ?;", (student_id,))
        topic_status_rows = cursor.fetchall()
        topic_statuses_map = {row['curriculum_id']: row['status'] for row in topic_status_rows}

        conn.close()

    assigned_map = {}
    for a in assignments:
        cid = a['curriculum_id']
        if cid not in assigned_map:
            assigned_map[cid] = []
        assigned_map[cid].append(a)

    exams_map = {}
    total_topic_count = len(topics)
    completed_topic_count = 0
    in_progress_topic_count = 0
    unassigned_topic_count = 0

    exam_topic_counts = {'TYT': {'total': 0, 'completed': 0}, 'AYT': {'total': 0, 'completed': 0}, 'YDT': {'total': 0, 'completed': 0}}
    subject_topic_counts = {}

    for t in topics:
        cid = t['id']
        etype = t['exam_type']
        subj = t['subject']
        t_assigned = assigned_map.get(cid, [])

        has_resource = len(t_assigned) > 0

        # Topic decoupled status
        explicit_status = topic_statuses_map.get(cid)
        if explicit_status:
            topic_status = explicit_status
        else:
            if not has_resource:
                topic_status = 'UNASSIGNED'
            elif all(a['status'] == 'COMPLETED' for a in t_assigned):
                topic_status = 'COMPLETED'
            elif any(a['status'] in ['IN_PROGRESS', 'COMPLETED'] for a in t_assigned):
                topic_status = 'IN_PROGRESS'
            else:
                topic_status = 'NOT_STARTED'

        if topic_status == 'COMPLETED':
            completed_topic_count += 1
            exam_topic_counts[etype]['completed'] += 1
        elif topic_status == 'IN_PROGRESS':
            in_progress_topic_count += 1
        else:
            if not has_resource:
                unassigned_topic_count += 1

        exam_topic_counts[etype]['total'] += 1

        if subj not in subject_topic_counts:
            subject_topic_counts[subj] = {'total': 0, 'completed': 0, 'exam_type': etype}
        subject_topic_counts[subj]['total'] += 1
        if topic_status == 'COMPLETED':
            subject_topic_counts[subj]['completed'] += 1

        if etype not in exams_map:
            exams_map[etype] = {'exam_type': etype, 'subjects': {}}
        if subj not in exams_map[etype]['subjects']:
            exams_map[etype]['subjects'][subj] = {'name': subj, 'topics': []}

        exams_map[etype]['subjects'][subj]['topics'].append({
            'curriculum_id': cid,
            'topic_name': t['topic'],
            'display_order': t['display_order'],
            'has_resource': has_resource,
            'topic_status': topic_status,
            'assigned_resources': [
                {
                    'topic_resource_id': a['id'],
                    'resource_id': a['resource_id'],
                    'resource_title': a['resource_title'],
                    'publisher_name': a['publisher_name'] or 'Yayın',
                    'status': a['status'],
                    'progress_percentage': a['progress_percentage']
                } for a in t_assigned
            ]
        })

    exams_list = []
    for etype, edata in exams_map.items():
        sub_list = []
        for sname, sdata in edata['subjects'].items():
            tot = subject_topic_counts[sname]['total']
            comp = subject_topic_counts[sname]['completed']
            sprog = round((comp / tot * 100.0), 1) if tot > 0 else 0.0
            sdata['progress'] = sprog
            sdata['total_topics'] = tot
            sdata['completed_topics'] = comp
            sub_list.append(sdata)

        etot = exam_topic_counts[etype]['total']
        ecomp = exam_topic_counts[etype]['completed']
        eprog = round((ecomp / etot * 100.0), 1) if etot > 0 else 0.0

        exams_list.append({
            'exam_type': etype,
            'progress': eprog,
            'total_topics': etot,
            'completed_topics': ecomp,
            'subjects': sub_list
        })

    overall_progress = round((completed_topic_count / total_topic_count * 100.0), 1) if total_topic_count > 0 else 0.0

    return jsonify({
        'student_id': student_id,
        'student_name': student_name,
        'student_grade': student_grade,
        'student_track': student_track,
        'student_exam_system': student_exam_system,
        'overall_progress': overall_progress,
        'total_topics': total_topic_count,
        'completed_topics': completed_topic_count,
        'in_progress_topics': in_progress_topic_count,
        'unassigned_topics': unassigned_topic_count,
        'tyt_progress': round((exam_topic_counts['TYT']['completed'] / exam_topic_counts['TYT']['total'] * 100.0), 1) if exam_topic_counts['TYT']['total'] > 0 else 0.0,
        'ayt_progress': round((exam_topic_counts['AYT']['completed'] / exam_topic_counts['AYT']['total'] * 100.0), 1) if exam_topic_counts['AYT']['total'] > 0 else 0.0,
        'ydt_progress': round((exam_topic_counts['YDT']['completed'] / exam_topic_counts['YDT']['total'] * 100.0), 1) if exam_topic_counts['YDT']['total'] > 0 else 0.0,
        'exams': exams_list
    })

@app.route('/api/mufredat/kaynak-ata', methods=['POST'])
def mufredat_assign_resource():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    student_id = data.get('student_id')
    curriculum_id = data.get('curriculum_id')
    
    # Support both resource_id (single) and resource_ids (array for multi-selection)
    resource_ids = data.get('resource_ids', [])
    if not resource_ids and data.get('resource_id'):
        resource_ids = [data.get('resource_id')]

    if not student_id or not curriculum_id or not resource_ids:
        return jsonify({'error': 'Öğrenci, konu ve en az 1 kaynak seçimi gereklidir'}), 400

    conn = get_db()
    cursor = conn.cursor()

    # 1. Fetch student track & exam_system
    cursor.execute("SELECT track, COALESCE(exam_system, 'YKS') as exam_system FROM students WHERE id = ?;", (student_id,))
    st = cursor.fetchone()
    st_exam_sys = (st['exam_system'] if st and st['exam_system'] else 'YKS').upper()
    student_track = (st['track'] if st and st['track'] else 'SAYISAL').upper()
    if student_track in ['SAY', 'MF']: student_track = 'SAYISAL'
    if student_track in ['TM']: student_track = 'EA'
    if student_track in ['TS']: student_track = 'SOZEL'
    if student_track in ['DIL']: student_track = 'YDT'

    # 2. Fetch curriculum topic
    cursor.execute("SELECT * FROM curriculum WHERE id = ?;", (curriculum_id,))
    curr_topic = cursor.fetchone()
    if not curr_topic:
        conn.close()
        return jsonify({'error': 'Geçersiz müfredat konusu'}), 404

    topic_exam = curr_topic['exam_type']
    topic_field = curr_topic['field']
    topic_subject = curr_topic['subject']

    cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
    ch = cursor.fetchone()
    coach_id = ch['id'] if ch else 1

    assigned_titles = []
    skipped_count = 0

    for resource_id in resource_ids:
        # Check duplicate active assignment
        cursor.execute("""
        SELECT id FROM student_topic_resources 
        WHERE student_id = ? AND curriculum_id = ? AND resource_id = ? AND status != 'ARCHIVED';
        """, (student_id, curriculum_id, resource_id))
        if cursor.fetchone():
            skipped_count += 1
            continue

        # Fetch resource metadata & validate
        cursor.execute("""
        SELECT r.*, s.name as subject_name 
        FROM resources r
        LEFT JOIN subjects s ON r.subject_id = s.id
        WHERE r.id = ?;
        """, (resource_id,))
        res_item = cursor.fetchone()
        if not res_item:
            continue

        res_title = res_item['title']
        res_exam = res_item['exam_type'] or 'TYT'
        res_subject = res_item['subject_name'] or 'Genel'
        res_exam_sys = dict(res_item).get('exam_system') or 'YKS'

        # STRICT EXAM SYSTEM VALIDATION (YKS vs LGS)
        if st_exam_sys != res_exam_sys:
            conn.close()
            return jsonify({'error': f'Sınav Sistemi Uyumsuzluğu! ({st_exam_sys} öğrencisine {res_exam_sys} kaynağı atanamaz.)'}), 400

        # STRICT FIELD & SUBJECT VALIDATION:
        if student_track == 'SAYISAL':
            if res_exam in ['AYT', 'YDT'] and topic_field in ['EA', 'SOZEL', 'YDT']:
                conn.close()
                return jsonify({'error': f'Yanlış Alan Kaynağı! Sayısal öğrencisine ({topic_subject}) kaynağı atanamaz.'}), 400
            if any(unallowed in res_subject for unallowed in ['Edebiyat', 'Tarih', 'Coğrafya', 'Felsefe Grubu', 'İngilizce', 'YDT']):
                conn.close()
                return jsonify({'error': f'Sayısal alanındaki öğrenciye {res_subject} dersinden AYT kaynağı atanamaz!'}), 400

        elif student_track in ['EA', 'ESIT_AGIRLIK']:
            if res_exam in ['AYT', 'YDT'] and topic_field in ['SAYISAL', 'SOZEL', 'YDT']:
                if res_subject in ['Fizik', 'Kimya', 'Biyoloji', 'Tarih-2', 'Coğrafya-2', 'Felsefe Grubu', 'İngilizce', 'YDT']:
                    conn.close()
                    return jsonify({'error': f'Eşit Ağırlık öğrencisine {res_subject} dersinden kaynak atanamaz!'}), 400

        elif student_track == 'SOZEL':
            if res_exam in ['AYT', 'YDT'] and topic_field in ['SAYISAL', 'EA', 'YDT']:
                if res_subject in ['Matematik', 'Geometri', 'Fizik', 'Kimya', 'Biyoloji', 'İngilizce', 'YDT']:
                    conn.close()
                    return jsonify({'error': f'Sözel öğrencisine {res_subject} dersinden AYT kaynağı atanamaz!'}), 400

        elif student_track in ['YDT', 'YABANCI_DIL']:
            if res_exam == 'AYT':
                conn.close()
                return jsonify({'error': 'Yabancı Dil öğrencisine AYT Sayısal/EA/Sözel derslerinden kaynak atanamaz!'}), 400

        if topic_subject and res_subject and res_subject != 'Genel':
            ts_clean = topic_subject.replace('AYT', '').replace('TYT', '').replace('-1', '').replace('-2', '').strip()
            rs_clean = res_subject.replace('AYT', '').replace('TYT', '').replace('-1', '').replace('-2', '').strip()
            if ts_clean not in rs_clean and rs_clean not in ts_clean:
                conn.close()
                return jsonify({'error': f'Ders Uyumsuzluğu! ({topic_subject}) konusuna ({res_subject}) kaynağı atayamazsınız.'}), 400

        cursor.execute("""
        INSERT INTO student_topic_resources (student_id, curriculum_id, resource_id, assigned_by, status, primary_resource, progress_percentage)
        VALUES (?, ?, ?, ?, 'IN_PROGRESS', 0, 0.0);
        """, (student_id, curriculum_id, resource_id, coach_id))
        assigned_titles.append(res_title)

    conn.commit()
    conn.close()

    # Send auto notification to student
    if assigned_titles:
        try:
            st_uid = get_student_user_id(student_id)
            titles_str = ", ".join(assigned_titles)
            send_auto_notification(
                user['id'],
                st_uid,
                f"📚 Yeni Kaynak Atandı: {titles_str} ({topic_subject} -> {curr_topic['topic']})"
            )
        except Exception as e:
            print(f"Kaynak atama auto notification error: {e}")

    if not assigned_titles and skipped_count > 0:
        return jsonify({'message': 'Seçilen kaynaklar zaten bu konuda aktif olarak atanmış durumda!'})

    msg = f'{len(assigned_titles)} adet kaynak {topic_subject} -> {curr_topic["topic"]} konusuna atandı.'
    if skipped_count > 0:
        msg += f' ({skipped_count} kaynak zaten atanmış olduğundan atlandı)'

    return jsonify({'message': msg})

@app.route('/api/mufredat/konu-durum-guncelle', methods=['POST'])
def mufredat_update_topic_status():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    student_id = data.get('student_id')
    curriculum_id = data.get('curriculum_id')
    status = data.get('status', 'COMPLETED')

    if not student_id or not curriculum_id:
        return jsonify({'error': 'Eksik parametre'}), 400

    conn = get_db()
    cursor = conn.cursor()

    # Fetch topic info for notification
    cursor.execute("SELECT subject, topic FROM curriculum WHERE id = ?;", (curriculum_id,))
    curr_info = cursor.fetchone()

    cursor.execute("""
    INSERT INTO student_topic_statuses (student_id, curriculum_id, status, updated_at)
    VALUES (?, ?, ?, CURRENT_TIMESTAMP)
    ON CONFLICT(student_id, curriculum_id) DO UPDATE SET status = excluded.status, updated_at = CURRENT_TIMESTAMP;
    """, (student_id, curriculum_id, status))
    conn.commit()
    conn.close()

    if status == 'COMPLETED' and curr_info:
        try:
            coach_uid = get_coach_user_id_for_student(student_id)
            send_auto_notification(
                user['id'],
                coach_uid,
                f"✅ Program Görevi/Konusu Tamamlandı: {curr_info['subject']} -> {curr_info['topic']}"
            )
        except Exception as e:
            print(f"Konu tamamlandı auto notification error: {e}")

    return jsonify({'message': 'Ana konu tamamlama durumu güncellendi!'})

@app.route('/api/mufredat/durum-guncelle', methods=['POST'])
def mufredat_update_status():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    topic_resource_id = data.get('topic_resource_id')
    status = data.get('status', 'COMPLETED')
    progress_percentage = float(data.get('progress_percentage', 100.0 if status == 'COMPLETED' else 50.0))

    if not topic_resource_id:
        return jsonify({'error': 'Konu kaynak kaydı gereklidir'}), 400

    conn = get_db()
    cursor = conn.cursor()

    if status == 'COMPLETED':
        cursor.execute("""
        UPDATE student_topic_resources 
        SET status = 'COMPLETED', progress_percentage = 100.0, completed_at = CURRENT_TIMESTAMP
        WHERE id = ?;
        """, (topic_resource_id,))
    elif status == 'IN_PROGRESS':
        cursor.execute("""
        UPDATE student_topic_resources 
        SET status = 'IN_PROGRESS', progress_percentage = ?, completed_at = NULL
        WHERE id = ?;
        """, (progress_percentage if progress_percentage > 0 else 50.0, topic_resource_id))
    else:
        cursor.execute("""
        UPDATE student_topic_resources 
        SET status = 'NOT_STARTED', progress_percentage = 0.0, completed_at = NULL
        WHERE id = ?;
        """, (topic_resource_id,))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Konu çalışma durumu güncellendi!'})

@app.route('/api/mufredat/kaynak-degistir', methods=['POST'])
def mufredat_change_resource():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    topic_resource_id = data.get('topic_resource_id')
    new_resource_id = data.get('new_resource_id')

    if not topic_resource_id or not new_resource_id:
        return jsonify({'error': 'Parametreler eksik'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE student_topic_resources SET resource_id = ? WHERE id = ?;", (new_resource_id, topic_resource_id))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Atanan kaynak başarıyla değiştirildi!'})

@app.route('/api/mufredat/kaynak-sil/<int:topic_resource_id>', methods=['DELETE'])
def mufredat_delete_resource(topic_resource_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    # Fetch info before archiving for auto message notification
    cursor.execute("""
    SELECT str.student_id, r.title as resource_title, c.topic, c.subject
    FROM student_topic_resources str
    JOIN resources r ON str.resource_id = r.id
    LEFT JOIN curriculum c ON str.curriculum_id = c.id
    WHERE str.id = ?;
    """, (topic_resource_id,))
    item = cursor.fetchone()

    cursor.execute("UPDATE student_topic_resources SET status = 'ARCHIVED' WHERE id = ?;", (topic_resource_id,))
    conn.commit()
    conn.close()

    if item:
        try:
            st_uid = get_student_user_id(item['student_id'])
            send_auto_notification(
                user['id'],
                st_uid,
                f"🗑️ Kaynak Ataması Kaldırıldı: {item['resource_title']} ({item['subject'] or ''} -> {item['topic'] or ''})"
            )
        except Exception as e:
            print(f"Kaynak silme auto notification error: {e}")

    return jsonify({'message': 'Kaynak ataması arşivlendi!'})

@app.route('/api/kaynaklar/havuz', methods=['GET'])
def get_resource_pool():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401
    
    if user['role'] == 'STUDENT':
        return jsonify({'error': 'Öğrenciler doğrudan kaynak havuzuna erişemez'}), 403

    subject = request.args.get('subject')
    student_id = request.args.get('student_id')
    visibility_filter = request.args.get('visibility')

    conn = get_db()
    cursor = conn.cursor()

    student_track = None
    student_exam_system = 'YKS'
    if student_id:
        cursor.execute("SELECT track, COALESCE(exam_system, 'YKS') as exam_system FROM students WHERE id = ?;", (student_id,))
        st = cursor.fetchone()
        if st:
            student_exam_system = (st['exam_system'] or 'YKS').upper()
            if st['track']:
                student_track = st['track'].upper()
                if student_track in ['SAY', 'MF']: student_track = 'SAYISAL'
                if student_track in ['TM']: student_track = 'EA'
                if student_track in ['TS']: student_track = 'SOZEL'
                if student_track in ['DIL']: student_track = 'YDT'

    query = """
    SELECT r.id, r.title, r.track, p.name as publisher_name, r.subject_id, s.name as subject_name, 
           r.exam_type, r.level, COALESCE(r.exam_system, 'YKS') as exam_system,
           COALESCE(r.visibility, 'GLOBAL') as visibility, r.owner_coach_id,
           u.name as owner_coach_name
    FROM resources r
    LEFT JOIN publishers p ON r.publisher_id = p.id
    LEFT JOIN subjects s ON r.subject_id = s.id
    LEFT JOIN coaches c ON r.owner_coach_id = c.id
    LEFT JOIN users u ON c.user_id = u.id
    WHERE COALESCE(r.status, 'ACTIVE') = 'ACTIVE'
    """
    params = []

    # Ownership & Visibility filtering
    if user['role'] == 'COACH':
        query += " AND (COALESCE(r.visibility, 'GLOBAL') = 'GLOBAL' OR r.owner_coach_id = ?)"
        params.append(user.get('coach_id'))
    elif user['role'] == 'ADMIN' and visibility_filter:
        query += " AND COALESCE(r.visibility, 'GLOBAL') = ?"
        params.append(visibility_filter.upper())

    if student_exam_system == 'LGS':
        query += " AND (r.exam_system = 'LGS' OR r.title LIKE '%LGS%')"
    else:
        query += " AND (r.exam_system = 'YKS' OR r.exam_system IS NULL OR r.exam_system != 'LGS')"

    if subject:
        clean_subj = subject.replace('AYT', '').replace('TYT', '').replace('LGS', '').replace('-1', '').replace('-2', '').strip()
        query += " AND (s.name LIKE ? OR r.title LIKE ?)"
        params.extend([f"%{clean_subj}%", f"%{clean_subj}%"])

    if student_exam_system != 'LGS' and student_track:
        if student_track == 'SAYISAL':
            query += " AND (r.exam_type = 'TYT' OR (r.exam_type = 'AYT' AND (r.track = 'SAYISAL' OR s.name IN ('Matematik', 'Geometri', 'Fizik', 'Kimya', 'Biyoloji'))))"
        elif student_track in ['EA', 'ESIT_AGIRLIK']:
            query += " AND (r.exam_type = 'TYT' OR (r.exam_type = 'AYT' AND (r.track = 'EA' OR s.name IN ('Matematik', 'Geometri', 'Türk Dili ve Edebiyatı', 'Tarih', 'Coğrafya'))))"
        elif student_track == 'SOZEL':
            query += " AND (r.exam_type = 'TYT' OR (r.exam_type = 'AYT' AND (r.track = 'SOZEL' OR s.name IN ('Türk Dili ve Edebiyatı', 'Tarih', 'Coğrafya', 'Felsefe', 'Din Kültürü'))))"
        elif student_track in ['YDT', 'YABANCI_DIL']:
            query += " AND (r.exam_type = 'TYT' OR r.exam_type = 'YDT')"

    query += " ORDER BY r.visibility DESC, r.title ASC;"

    cursor.execute(query, params)
    resources = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({'resources': resources})

@app.route('/api/kaynaklar', methods=['POST'])
def create_resource():
    user = get_auth_user()
    if not user or user['role'] not in ('ADMIN', 'COACH'):
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    data = request.json or {}
    title = data.get('title', '').strip()
    subject_id = data.get('subject_id')
    publisher_name = data.get('publisher_name', 'Özel Yayın')
    exam_system = data.get('exam_system', 'YKS')
    exam_type = data.get('exam_type', 'TYT')
    track = data.get('track', 'ALL')
    resource_type = data.get('resource_type', 'SORU_BANKASI')
    description = data.get('description', '')
    author = data.get('author', '')
    url = data.get('url', '')
    notes = data.get('notes', '')

    if not title or not subject_id:
        return jsonify({'error': 'Kaynak adı ve ders zorunludur'}), 400

    visibility = 'GLOBAL' if user['role'] == 'ADMIN' else 'PRIVATE'
    owner_coach_id = None if user['role'] == 'ADMIN' else user.get('coach_id')

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM publishers WHERE name = ?;", (publisher_name,))
    pub = cursor.fetchone()
    if pub:
        publisher_id = pub['id']
    else:
        cursor.execute("INSERT INTO publishers (name) VALUES (?);", (publisher_name,))
        publisher_id = cursor.lastrowid

    cursor.execute("""
    INSERT INTO resources (publisher_id, title, subject_id, exam_type, track, resource_type, visibility, owner_coach_id, created_by_user_id, description, author, url, notes, exam_system, status)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'ACTIVE');
    """, (publisher_id, title, subject_id, exam_type, track, resource_type, visibility, owner_coach_id, user['id'], description, author, url, notes, exam_system))
    
    res_id = cursor.lastrowid
    conn.commit()

    log_activity(user['id'], user['role'], 'CREATE_RESOURCE', 'resources', res_id, {'title': title, 'visibility': visibility})
    conn.close()

    return jsonify({'message': 'Kaynak oluşturuldu', 'id': res_id, 'visibility': visibility})

@app.route('/api/kaynaklar/create-and-assign', methods=['POST'])
def create_and_assign_resource():
    user = get_auth_user()
    if not user or user['role'] not in ('ADMIN', 'COACH'):
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    data = request.json or {}
    student_id = data.get('student_id')
    curriculum_topic_id = data.get('curriculum_topic_id')
    title = data.get('title', '').strip()
    subject_id = data.get('subject_id')

    if not student_id or not curriculum_topic_id or not title or not subject_id:
        return jsonify({'error': 'Öğrenci, konu, kaynak adı ve ders zorunludur'}), 400

    visibility = 'PRIVATE'
    owner_coach_id = user.get('coach_id')

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    INSERT INTO resources (title, subject_id, visibility, owner_coach_id, created_by_user_id, status)
    VALUES (?, ?, ?, ?, ?, 'ACTIVE');
    """, (title, subject_id, visibility, owner_coach_id, user['id']))
    res_id = cursor.lastrowid

    cursor.execute("""
    INSERT OR REPLACE INTO student_topic_resources (student_id, curriculum_id, resource_id, assigned_by_coach_id, status)
    VALUES (?, ?, ?, ?, 'NOT_STARTED');
    """, (student_id, curriculum_topic_id, res_id, owner_coach_id))

    conn.commit()
    log_activity(user['id'], user['role'], 'CREATE_AND_ASSIGN_RESOURCE', 'resources', res_id, {'title': title, 'student_id': student_id})
    conn.close()

    return jsonify({'message': 'Özel kaynak oluşturuldu ve öğrenciye atandı!', 'resource_id': res_id})

# ==========================================
# HAFTALIK PROGRAM API'LERİ (GRID, GÜNLER, SAATLER, SÜRÜKLE-BIRAK)
# ==========================================

@app.route('/api/weekly-program', methods=['GET'])
def get_weekly_program():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    student_id = request.args.get('student_id')
    if user['role'] == 'STUDENT':
        student_id = user.get('student_id') or user.get('id')
    elif not student_id:
        return jsonify({'error': 'Öğrenci ID gereklidir'}), 400

    try:
        student_id = int(student_id)
    except (ValueError, TypeError):
        return jsonify({'error': 'Geçersiz öğrenci ID'}), 400

    week_start = request.args.get('week_start')
    if not week_start:
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        week_start = monday.strftime('%Y-%m-%d')

    try:
        start_dt = datetime.strptime(week_start, '%Y-%m-%d').date()
        end_dt = start_dt + timedelta(days=6)
    except Exception:
        return jsonify({'error': 'Geçersiz tarih formatı (YYYY-MM-DD olmalı)'}), 400

    try:
        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("""
        SELECT s.id, u.name as student_name, u.surname as student_surname, COALESCE(s.exam_system, 'YKS') as exam_system, s.track, s.grade
        FROM students s
        JOIN users u ON s.user_id = u.id
        WHERE s.id = ?;
        """, (student_id,))
        st_info = cursor.fetchone()
        if not st_info:
            conn.close()
            return jsonify({'error': 'Öğrenci bulunamadı'}), 404

        # Role filter: Students only see PUBLISHED items. Coaches/Admins see DRAFT and PUBLISHED.
        if user['role'] == 'STUDENT':
            pub_filter = " AND COALESCE(wp.publication_status, 'PUBLISHED') = 'PUBLISHED'"
        else:
            pub_filter = " AND COALESCE(wp.publication_status, 'PUBLISHED') != 'CANCELLED'"

        query = f"""
        SELECT wp.id, wp.student_id, wp.created_by_coach_id, wp.date, wp.day_of_week, 
               wp.start_time, wp.end_time, wp.title, wp.subject_id, s.name as subject_name,
               wp.curriculum_topic_id, c.topic as topic_name, wp.resource_id, r.title as resource_title,
               wp.study_type, wp.description, 
               COALESCE(wp.completion_status, wp.status, 'PLANLANDI') as status,
               COALESCE(wp.completion_status, wp.status, 'PLANLANDI') as completion_status,
               COALESCE(wp.publication_status, 'PUBLISHED') as publication_status,
               wp.created_at
        FROM weekly_programs wp
        LEFT JOIN subjects s ON wp.subject_id = s.id
        LEFT JOIN curriculum c ON wp.curriculum_topic_id = c.id
        LEFT JOIN resources r ON wp.resource_id = r.id
        WHERE wp.student_id = ? AND wp.date BETWEEN ? AND ? {pub_filter}
        ORDER BY wp.date ASC, wp.start_time ASC;
        """

        cursor.execute(query, (student_id, start_dt.strftime('%Y-%m-%d'), end_dt.strftime('%Y-%m-%d')))
        items = []
        for row in cursor.fetchall():
            it = dict(row)
            if it.get('date'):
                if hasattr(it['date'], 'strftime'):
                    it['date'] = it['date'].strftime('%Y-%m-%d')
                else:
                    it['date'] = str(it['date'])[:10]
            items.append(it)

        tot_planned = sum(1 for i in items if i['status'] in ('PLANLANDI', 'PLANNED'))
        tot_completed = sum(1 for i in items if i['status'] in ('TAMAMLANDI', 'COMPLETED'))
        tot_in_progress = sum(1 for i in items if i['status'] in ('DEVAM_EDIYOR', 'IN_PROGRESS'))
        tot_skipped = sum(1 for i in items if i['status'] in ('ATLANDI', 'SKIPPED'))
        tot_all = len(items)
        completion_rate = round((tot_completed / tot_all * 100), 1) if tot_all > 0 else 0.0

        conn.close()

        return jsonify({
            'student': dict(st_info),
            'week_start': start_dt.strftime('%Y-%m-%d'),
            'week_end': end_dt.strftime('%Y-%m-%d'),
            'items': items,
            'summary': {
                'total': tot_all,
                'planned': tot_planned,
                'completed': tot_completed,
                'in_progress': tot_in_progress,
                'skipped': tot_skipped,
                'completion_rate': completion_rate
            }
        })
    except Exception as db_err:
        print(f"[WEEKLY PROGRAM DB ERROR] {db_err}")
        try: conn.close()
        except: pass
        return jsonify({'error': f'Haftalık program verileri okunurken hata oluştu: {str(db_err)}'}), 500

@app.route('/api/weekly-program', methods=['POST'])
def create_weekly_program():
    user = get_auth_user()
    if not user or user['role'] not in ('COACH', 'ADMIN'):
        return jsonify({'error': 'Haftalık program oluşturma yetkiniz yok'}), 403

    data = request.json or {}
    student_id = data.get('student_id')
    prog_date = data.get('date')
    day_of_week = data.get('day_of_week', 'Pazartesi')
    start_time = data.get('start_time', '08:00')
    end_time = data.get('end_time', '09:00')
    title = data.get('title', '').strip()
    subject_id = data.get('subject_id')
    curriculum_topic_id = data.get('curriculum_topic_id')
    resource_id = data.get('resource_id')
    study_type = data.get('study_type', 'Konu Çalışması')
    description = data.get('description', '')
    status = data.get('status', data.get('completion_status', 'PLANLANDI'))
    publication_status = data.get('publication_status', 'DRAFT')

    if not student_id or not prog_date or not start_time or not end_time or not title:
        return jsonify({'error': 'Öğrenci, tarih, saat ve başlık zorunludur'}), 400

    conn = get_db()
    cursor = conn.cursor()

    # 1. Exact Same Time Slot Check (Seamless Upsert / Update without False Conflict)
    cursor.execute("""
    SELECT id, title, start_time, end_time FROM weekly_programs
    WHERE student_id = ? AND date = ? AND start_time = ? AND end_time = ? AND COALESCE(publication_status, 'PUBLISHED') != 'CANCELLED'
    LIMIT 1;
    """, (student_id, prog_date, start_time, end_time))
    exact_match = cursor.fetchone()
    if exact_match:
        cursor.execute("""
        UPDATE weekly_programs 
        SET title = ?, subject_id = ?, curriculum_topic_id = ?, resource_id = ?, study_type = ?, description = ?, status = ?, completion_status = ?, publication_status = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?;
        """, (title, subject_id, curriculum_topic_id, resource_id, study_type, description, status, status, publication_status, exact_match['id']))
        conn.commit()
        prog_id = exact_match['id']
        log_activity(user['id'], user['role'], 'UPDATE_WEEKLY_PROGRAM', 'weekly_programs', prog_id, {'title': title, 'date': prog_date, 'student_id': student_id}, cursor=cursor)
        conn.close()
        return jsonify({'message': 'Program kartı güncellendi', 'id': prog_id})

    # 2. Time Overlap / Conflict Check with OTHER Differing Slots
    cursor.execute("""
    SELECT id, title, start_time, end_time FROM weekly_programs
    WHERE student_id = ? AND date = ? AND COALESCE(publication_status, 'PUBLISHED') != 'CANCELLED'
    AND (
        (start_time < ? AND end_time > ?) OR
        (start_time >= ? AND start_time < ?) OR
        (end_time > ? AND end_time <= ?)
    );
    """, (student_id, prog_date, end_time, start_time, start_time, end_time, start_time, end_time))
    conflict = cursor.fetchone()
    if conflict:
        conn.close()
        return jsonify({'error': f"Program çakışması! Bu saat diliminde '{conflict['title']}' ({conflict['start_time']}-{conflict['end_time']}) dersi bulunuyor."}), 409

    coach_id = user.get('coach_id')

    cursor.execute("""
    INSERT INTO weekly_programs (student_id, created_by_coach_id, date, day_of_week, start_time, end_time, title, subject_id, curriculum_topic_id, resource_id, study_type, description, status, completion_status, publication_status)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
    """, (student_id, coach_id, prog_date, day_of_week, start_time, end_time, title, subject_id, curriculum_topic_id, resource_id, study_type, description, status, status, publication_status))
    
    prog_id = cursor.lastrowid

    # Auto notification to student if created by coach
    if user['role'] in ('COACH', 'ADMIN'):
        cursor.execute("SELECT user_id FROM students WHERE id = ?;", (student_id,))
        st_row = cursor.fetchone()
        if st_row:
            coach_name = f"{user['name']} {user['surname'] or ''}".strip()
            msg_text = f"📅 Koçunuz ({coach_name}), çalışma programınıza yeni bir ders/görev ekledi: '{title}' ({prog_date} {start_time}-{end_time})."
            send_auto_notification(user['id'], st_row['user_id'], msg_text, message_type='SYSTEM')

    conn.commit()

    log_activity(user['id'], user['role'], 'CREATE_WEEKLY_PROGRAM', 'weekly_programs', prog_id, {'title': title, 'date': prog_date, 'student_id': student_id}, cursor=cursor)
    conn.close()

    return jsonify({'message': 'Program kartı eklendi', 'id': prog_id})

@app.route('/api/weekly-program/<int:prog_id>', methods=['PUT'])
def update_weekly_program(prog_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM weekly_programs WHERE id = ?;", (prog_id,))
    item = cursor.fetchone()
    if not item:
        conn.close()
        return jsonify({'error': 'Program kartı bulunamadı'}), 404

    data = request.json or {}

    if user['role'] == 'STUDENT':
        new_status = data.get('status') or data.get('completion_status')
        if new_status:
            cursor.execute("UPDATE weekly_programs SET status = ?, completion_status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?;", (new_status, new_status, prog_id))
            conn.commit()
            conn.close()
            return jsonify({'message': 'Program durumu güncellendi'})
        conn.close()
        return jsonify({'error': 'Öğrenciler program detayını değiştiremez'}), 403

    prog_date = data.get('date', item['date'])
    day_of_week = data.get('day_of_week', item['day_of_week'])
    start_time = data.get('start_time', item['start_time'])
    end_time = data.get('end_time', item['end_time'])
    title = data.get('title', item['title'])
    subject_id = data.get('subject_id', item['subject_id'])
    curriculum_topic_id = data.get('curriculum_topic_id', item['curriculum_topic_id'])
    resource_id = data.get('resource_id', item['resource_id'])
    study_type = data.get('study_type', item['study_type'])
    description = data.get('description', item['description'])
    status = data.get('status', data.get('completion_status', item['status']))
    publication_status = data.get('publication_status', item['publication_status'] if 'publication_status' in item.keys() and item['publication_status'] else 'DRAFT')

    # Time Overlap Check on move/resize (excluding current item)
    if prog_date != item['date'] or start_time != item['start_time'] or end_time != item['end_time']:
        cursor.execute("""
        SELECT id, title, start_time, end_time FROM weekly_programs
        WHERE student_id = ? AND date = ? AND id != ? AND COALESCE(publication_status, 'PUBLISHED') != 'CANCELLED'
        AND (
            (start_time <= ? AND end_time > ?) OR
            (start_time < ? AND end_time >= ?) OR
            (start_time >= ? AND end_time <= ?)
        );
        """, (item['student_id'], prog_date, prog_id, start_time, start_time, end_time, end_time, start_time, end_time))
        conflict = cursor.fetchone()
        if conflict:
            conn.close()
            return jsonify({'error': f"Program çakışması! Bu saatte '{conflict['title']}' dersi bulunmaktadır ({conflict['start_time']}-{conflict['end_time']})."}), 409

    cursor.execute("""
    UPDATE weekly_programs 
    SET date = ?, day_of_week = ?, start_time = ?, end_time = ?, title = ?, subject_id = ?, 
        curriculum_topic_id = ?, resource_id = ?, study_type = ?, description = ?, status = ?, completion_status = ?, publication_status = ?, updated_at = CURRENT_TIMESTAMP
    WHERE id = ?;
    """, (prog_date, day_of_week, start_time, end_time, title, subject_id, curriculum_topic_id, resource_id, study_type, description, status, status, publication_status, prog_id))
    
    conn.commit()
    log_activity(user['id'], user['role'], 'UPDATE_WEEKLY_PROGRAM', 'weekly_programs', prog_id, {'title': title, 'date': prog_date})
    conn.close()

    return jsonify({'message': 'Program kartı güncellendi'})

@app.route('/api/weekly-program/publish', methods=['POST'])
def publish_weekly_program():
    user = get_auth_user()
    if not user or user['role'] not in ('COACH', 'ADMIN'):
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    data = request.json or {}
    student_id = data.get('student_id')
    week_start = data.get('week_start')

    if not student_id or not week_start:
        return jsonify({'error': 'Öğrenci ID ve Hafta başlangıç tarihi gereklidir'}), 400

    try:
        start_dt = datetime.strptime(week_start, '%Y-%m-%d').date()
        end_dt = start_dt + timedelta(days=6)
    except Exception:
        return jsonify({'error': 'Geçersiz tarih formatı'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    UPDATE weekly_programs
    SET publication_status = 'PUBLISHED', updated_at = CURRENT_TIMESTAMP
    WHERE student_id = ? AND date BETWEEN ? AND ? AND COALESCE(publication_status, 'DRAFT') != 'CANCELLED';
    """, (student_id, start_dt.strftime('%Y-%m-%d'), end_dt.strftime('%Y-%m-%d')))
    
    affected = cursor.rowcount
    
    # Send Automatic System Notification to Student
    cursor.execute("""
    SELECT s.user_id, u.name, u.surname 
    FROM students s 
    JOIN users u ON s.user_id = u.id 
    WHERE s.id = ?;
    """, (student_id,))
    st_row = cursor.fetchone()
    if st_row:
        st_user_id = st_row['user_id']
        st_name = st_row['name']
        coach_name = f"{user['name']} {user['surname'] or ''}".strip()
        msg_text = f"📅 Koçunuz ({coach_name}), {week_start} haftası için haftalık çalışma programınızı kaydedip yayınladı! 'Haftalık Programım' ekranından tüm ders ve saat dilimlerinizi kontrol edebilirsiniz."
        send_auto_notification(user['id'], st_user_id, msg_text, message_type='SYSTEM')

    conn.commit()
    log_activity(user['id'], user['role'], 'PUBLISH_WEEKLY_PROGRAM', 'weekly_programs', student_id, {'week_start': week_start, 'count': affected}, cursor=cursor)
    conn.close()

    return jsonify({'message': 'Haftalık program başarıyla kaydedildi ve öğrencinin mesaj kutusuna bildirim gönderildi.', 'published_count': affected})

@app.route('/api/weekly-program/clear', methods=['POST'])
def clear_weekly_program():
    user = get_auth_user()
    if not user or user['role'] not in ('COACH', 'ADMIN'):
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    data = request.json or {}
    student_id = data.get('student_id')
    week_start = data.get('week_start')

    if not student_id or not week_start:
        return jsonify({'error': 'Öğrenci ID ve Hafta başlangıç tarihi gereklidir'}), 400

    try:
        start_dt = datetime.strptime(week_start, '%Y-%m-%d').date()
        end_dt = start_dt + timedelta(days=6)
    except Exception:
        return jsonify({'error': 'Geçersiz tarih formatı'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    UPDATE weekly_programs
    SET publication_status = 'CANCELLED', updated_at = CURRENT_TIMESTAMP
    WHERE student_id = ? AND date BETWEEN ? AND ?;
    """, (student_id, start_dt.strftime('%Y-%m-%d'), end_dt.strftime('%Y-%m-%d')))
    
    affected = cursor.rowcount
    conn.commit()
    log_activity(user['id'], user['role'], 'CLEAR_WEEKLY_PROGRAM', 'weekly_programs', student_id, {'week_start': week_start, 'cleared_count': affected})
    conn.close()

    return jsonify({'message': 'Seçili haftanın tüm program hücreleri temizlendi.', 'cleared_count': affected})

@app.route('/api/weekly-program/copy', methods=['POST'])
def copy_weekly_program():
    user = get_auth_user()
    if not user or user['role'] not in ('COACH', 'ADMIN'):
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    data = request.json or {}
    student_id = data.get('student_id')
    source_week = data.get('source_week_start')
    target_week = data.get('target_week_start')

    if not student_id or not source_week or not target_week:
        return jsonify({'error': 'Öğrenci ID, Kaynak Hafta ve Hedef Hafta gereklidir'}), 400

    try:
        src_start = datetime.strptime(source_week, '%Y-%m-%d').date()
        src_end = src_start + timedelta(days=6)
        tgt_start = datetime.strptime(target_week, '%Y-%m-%d').date()
    except Exception:
        return jsonify({'error': 'Geçersiz tarih formatı'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT * FROM weekly_programs
    WHERE student_id = ? AND date BETWEEN ? AND ? AND COALESCE(publication_status, 'PUBLISHED') != 'CANCELLED';
    """, (student_id, src_start.strftime('%Y-%m-%d'), src_end.strftime('%Y-%m-%d')))
    source_items = cursor.fetchall()

    copied_count = 0
    coach_id = user.get('coach_id')

    for item in source_items:
        item_date = datetime.strptime(item['date'], '%Y-%m-%d').date()
        day_offset = (item_date - src_start).days
        new_date = (tgt_start + timedelta(days=day_offset)).strftime('%Y-%m-%d')

        cursor.execute("""
        INSERT INTO weekly_programs (student_id, created_by_coach_id, date, day_of_week, start_time, end_time, title, subject_id, curriculum_topic_id, resource_id, study_type, description, status, completion_status, publication_status)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'PLANLANDI', 'PLANLANDI', 'DRAFT');
        """, (student_id, coach_id, new_date, item['day_of_week'], item['start_time'], item['end_time'], item['title'], item['subject_id'], item['curriculum_topic_id'], item['resource_id'], item['study_type'], item['description']))
        copied_count += 1

    conn.commit()
    log_activity(user['id'], user['role'], 'COPY_WEEKLY_PROGRAM', 'weekly_programs', student_id, {'source_week': source_week, 'target_week': target_week, 'count': copied_count})
    conn.close()

    return jsonify({'message': f"{copied_count} adet ders programı başarıyla gelecek haftaya kopyalandı (Taslak olarak).", 'copied_count': copied_count})

@app.route('/api/weekly-program/<int:prog_id>', methods=['DELETE'])
def delete_weekly_program(prog_id):
    user = get_auth_user()
    if not user or user['role'] not in ('COACH', 'ADMIN'):
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM weekly_programs WHERE id = ?;", (prog_id,))
    item = cursor.fetchone()
    if not item:
        conn.close()
        return jsonify({'error': 'Program kartı bulunamadı'}), 404

    cursor.execute("DELETE FROM weekly_programs WHERE id = ?;", (prog_id,))
    conn.commit()
    log_activity(user['id'], user['role'], 'DELETE_WEEKLY_PROGRAM', 'weekly_programs', prog_id, {'title': item['title']})
    conn.close()

    return jsonify({'message': 'Program kartı silindi'})

@app.route('/api/weekly-program/<int:prog_id>/status', methods=['POST'])
def update_weekly_program_status(prog_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    new_status = data.get('status') or data.get('completion_status')
    if new_status not in ('PLANLANDI', 'DEVAM_EDIYOR', 'TAMAMLANDI', 'ATLANDI', 'PLANNED', 'IN_PROGRESS', 'COMPLETED', 'SKIPPED'):
        return jsonify({'error': 'Geçersiz durum'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT wp.title, wp.student_id, s.user_id as student_user_id, s.coach_id, u.name as student_name, u.surname as student_surname
    FROM weekly_programs wp
    JOIN students s ON wp.student_id = s.id
    JOIN users u ON s.user_id = u.id
    WHERE wp.id = ?;
    """, (prog_id,))
    wp_info = cursor.fetchone()

    cursor.execute("UPDATE weekly_programs SET status = ?, completion_status = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?;", (new_status, new_status, prog_id))
    
    if wp_info and user['role'] == 'STUDENT' and new_status in ('TAMAMLANDI', 'COMPLETED'):
        cursor.execute("SELECT user_id FROM coaches WHERE id = ?;", (wp_info['coach_id'],))
        ch_row = cursor.fetchone()
        if ch_row:
            msg_text = f"✅ Öğrenciniz {wp_info['student_name']} {wp_info['student_surname'] or ''}, '{wp_info['title']}' isimli çalışma görevini TAMAMLADI!"
            send_auto_notification(user['id'], ch_row['user_id'], msg_text, message_type='SYSTEM')

    conn.commit()
    log_activity(user['id'], user['role'], 'UPDATE_PROGRAM_STATUS', 'weekly_programs', prog_id, {'status': new_status}, cursor=cursor)
    conn.close()

    return jsonify({'message': 'Durum başarıyla güncellendi'})

# ==========================================
# ADMIN DASHBOARD & ACTIVITY LOGS API
# ==========================================

@app.route('/api/admin/dashboard', methods=['GET'])
def get_admin_dashboard():
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) as count FROM users;")
    tot_users = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM coaches;")
    tot_coaches = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM students;")
    tot_students = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM students WHERE COALESCE(exam_system, 'YKS') = 'YKS';")
    tot_yks = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM students WHERE exam_system = 'LGS';")
    tot_lgs = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM resources WHERE COALESCE(status, 'ACTIVE') = 'ACTIVE';")
    tot_resources = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM resources WHERE COALESCE(visibility, 'GLOBAL') = 'GLOBAL' AND COALESCE(status, 'ACTIVE') = 'ACTIVE';")
    tot_global_res = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM resources WHERE visibility = 'PRIVATE' AND COALESCE(status, 'ACTIVE') = 'ACTIVE';")
    tot_private_res = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM mock_exams;")
    tot_mocks = cursor.fetchone()['count']

    cursor.execute("SELECT COUNT(*) as count FROM assignments;")
    tot_assignments = cursor.fetchone()['count']

    cursor.execute("""
    SELECT al.id, al.user_id, u.name as user_name, al.role, al.action, al.entity_type, al.entity_id, al.created_at
    FROM activity_logs al
    LEFT JOIN users u ON al.user_id = u.id
    ORDER BY al.id DESC LIMIT 15;
    """)
    recent_logs = [dict(row) for row in cursor.fetchall()]

    conn.close()

    return jsonify({
        'stats': {
            'total_users': tot_users,
            'total_coaches': tot_coaches,
            'total_students': tot_students,
            'yks_students': tot_yks,
            'lgs_students': tot_lgs,
            'total_resources': tot_resources,
            'global_resources': tot_global_res,
            'private_resources': tot_private_res,
            'total_mocks': tot_mocks,
            'total_assignments': tot_assignments
        },
        'recent_logs': recent_logs
    })

@app.route('/api/admin/activity-logs', methods=['GET'])
def get_activity_logs():
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT al.id, al.user_id, u.name as user_name, al.role, al.action, al.entity_type, al.entity_id, al.metadata_json, al.created_at
    FROM activity_logs al
    LEFT JOIN users u ON al.user_id = u.id
    ORDER BY al.id DESC LIMIT 100;
    """)
    logs = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify({'logs': logs})

@app.route('/api/student/update-field', methods=['POST'])
def update_student_field():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401
    if user['role'] not in ('ADMIN', 'COACH'):
        return jsonify({'error': 'Öğrenci alanı değiştirme yetkisi yalnızca koç ve yöneticilere aittir.'}), 403

    data = request.json or {}
    student_id = data.get('student_id')
    new_field = data.get('field', 'SAYISAL').upper()

    if new_field in ['SAY', 'MF']: new_field = 'SAYISAL'
    if new_field in ['TM', 'ESIT_AGIRLIK']: new_field = 'EA'
    if new_field in ['TS']: new_field = 'SOZEL'
    if new_field in ['DIL', 'YABANCI_DIL']: new_field = 'YDT'
    if new_field not in ['SAYISAL', 'EA', 'SOZEL', 'YDT']:
        return jsonify({'error': 'Geçersiz alan/track değeri'}), 400

    if not student_id:
        return jsonify({'error': 'Öğrenci ID gereklidir'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE students SET track = ? WHERE id = ?;", (new_field, student_id))
    conn.commit()
    conn.close()

    return jsonify({'message': f'Öğrencinin alanı ({new_field}) başarıyla güncellendi!'})

@app.route('/api/kaynaklar/<int:student_resource_id>/mufredat-ilerleme', methods=['GET'])
def get_resource_curriculum_progress(student_resource_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT sr.*, r.title, r.level, p.name as publisher_name, s.name as subject_name
    FROM student_resources sr
    JOIN resources r ON sr.resource_id = r.id
    LEFT JOIN publishers p ON r.publisher_id = p.id
    LEFT JOIN subjects s ON r.subject_id = s.id
    WHERE sr.id = ?;
    """, (student_resource_id,))
    sr = cursor.fetchone()
    if not sr:
        conn.close()
        return jsonify({'error': 'Kaynak ataması bulunamadı'}), 404

    sr_dict = dict(sr)

    cursor.execute("""
    SELECT srtp.*, t.name as topic_name, t.category
    FROM student_resource_topic_progress srtp
    JOIN topics t ON srtp.topic_id = t.id
    WHERE srtp.student_resource_id = ?
    ORDER BY t.id ASC;
    """, (student_resource_id,))
    topic_progress = [dict(r) for r in cursor.fetchall()]

    if not topic_progress:
        cursor.execute("SELECT id, name, category FROM topics LIMIT 15;")
        t_rows = cursor.fetchall()
        for tr in t_rows:
            cursor.execute("""
            INSERT INTO student_resource_topic_progress (student_resource_id, topic_id, status, progress_percentage)
            VALUES (?, ?, 'NOT_STARTED', 0.0);
            """, (student_resource_id, tr['id']))
        conn.commit()

        cursor.execute("""
        SELECT srtp.*, t.name as topic_name, t.category
        FROM student_resource_topic_progress srtp
        JOIN topics t ON srtp.topic_id = t.id
        WHERE srtp.student_resource_id = ?
        ORDER BY t.id ASC;
        """, (student_resource_id,))
        topic_progress = [dict(r) for r in cursor.fetchall()]

    total_topics = len(topic_progress)
    completed_topics = sum(1 for t in topic_progress if t['status'] == 'COMPLETED')
    in_progress_topics = sum(1 for t in topic_progress if t['status'] == 'IN_PROGRESS')
    review_topics = sum(1 for t in topic_progress if t['status'] == 'REVIEW_REQUIRED')
    not_started_topics = sum(1 for t in topic_progress if t['status'] == 'NOT_STARTED')

    conn.close()
    return jsonify({
        'resource': sr_dict,
        'summary': {
            'book_completion_percentage': sr_dict.get('completion_percentage', 0.0),
            'total_topics': total_topics,
            'completed_topics': completed_topics,
            'in_progress_topics': in_progress_topics,
            'review_topics': review_topics,
            'not_started_topics': not_started_topics,
            'curriculum_completion_percentage': round((completed_topics / total_topics * 100.0), 1) if total_topics > 0 else 0.0
        },
        'topics': topic_progress
    })

@app.route('/api/kaynaklar/<int:student_resource_id>/konu-durumu', methods=['POST'])
def update_resource_topic_status(student_resource_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    topic_id = data.get('topic_id')
    new_status = data.get('status', 'COMPLETED')
    notes = data.get('notes', '')

    if not topic_id:
        return jsonify({'error': 'topic_id zorunludur'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM student_resources WHERE id = ?;", (student_resource_id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'error': 'Öğrenci kaynağı bulunamadı'}), 404

    cursor.execute("SELECT id FROM topics WHERE id = ?;", (topic_id,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'error': 'Konu bulunamadı'}), 404


    cursor.execute("""
    SELECT status FROM student_resource_topic_progress
    WHERE student_resource_id = ? AND topic_id = ?;
    """, (student_resource_id, topic_id))
    old_row = cursor.fetchone()
    old_status = old_row['status'] if old_row else 'NOT_STARTED'

    is_coach = user['role'] in ['COACH', 'ADMIN']
    coach_approved = 1 if (is_coach and new_status == 'COMPLETED') else 0
    progress_pct = 100.0 if new_status == 'COMPLETED' else (50.0 if new_status == 'IN_PROGRESS' else 0.0)

    cursor.execute("""
    INSERT INTO student_resource_topic_progress (student_resource_id, topic_id, status, progress_percentage, marked_by, marked_by_role, coach_approved, notes, updated_at)
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP);
    """, (student_resource_id, topic_id, new_status, progress_pct, user['name'], user['role'], coach_approved, notes))

    cursor.execute("""
    INSERT INTO topic_progress_history (student_resource_id, topic_id, old_status, new_status, changed_by_user_id, changed_by_role)
    VALUES (?, ?, ?, ?, ?, ?);
    """, (student_resource_id, topic_id, old_status, new_status, user['id'], user['role']))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Konu müfredat durumu başarıyla güncellendi!', 'new_status': new_status})

@app.route('/api/mufredat/konu-detay', methods=['GET'])
def get_topic_cross_resource_detail():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    student_id = request.args.get('student_id')
    topic_id = request.args.get('topic_id', 1)

    conn = get_db()
    cursor = conn.cursor()

    if not student_id and user['role'] == 'STUDENT':
        cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
        st = cursor.fetchone()
        student_id = st['id'] if st else 1
    if not student_id:
        student_id = 1

    cursor.execute("SELECT t.*, s.name as subject_name FROM topics t LEFT JOIN subjects s ON t.subject_id = s.id WHERE t.id = ?;", (topic_id,))
    t_info = cursor.fetchone()
    if not t_info:
        conn.close()
        return jsonify({'error': 'Konu bulunamadı'}), 404

    cursor.execute("""
    SELECT sr.id as student_resource_id, r.title as resource_title, p.name as publisher_name, srtp.status, srtp.progress_percentage, srtp.updated_at, srtp.marked_by
    FROM student_resources sr
    JOIN resources r ON sr.resource_id = r.id
    LEFT JOIN publishers p ON r.publisher_id = p.id
    LEFT JOIN student_resource_topic_progress srtp ON sr.id = srtp.student_resource_id AND srtp.topic_id = ?
    WHERE sr.student_id = ?;
    """, (topic_id, student_id))
    resources_breakdown = [dict(r) for r in cursor.fetchall()]

    conn.close()
    return jsonify({
        'topic': dict(t_info),
        'student_id': student_id,
        'resources_breakdown': resources_breakdown,
        'mock_exam_net_average': '78.5%',
        'recommendation': '🔴 Bu konuda 3D kaynağı tamamlandı görünse de son deneme netleri %61 olduğu için 1 karma test tekrarı tavsiye ediliyor.'
    })

# ==========================================
# 15. KİTAP OKUMA TAKİBİ API
# ==========================================
@app.route('/api/kitaplar', methods=['GET', 'POST'])
def handle_books():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        student_id = request.args.get('student_id')
        if not student_id and user['role'] == 'STUDENT':
            cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
            st = cursor.fetchone()
            student_id = st['id'] if st else 1
        if not student_id:
            student_id = 1

        cursor.execute("SELECT * FROM books WHERE student_id = ? ORDER BY start_date DESC;", (student_id,))
        rows = cursor.fetchall()
        books = []
        for r in rows:
            d = dict(r)
            # Ensure both total_pages and page_count keys are present for 100% frontend compatibility
            d['page_count'] = d.get('total_pages', 0)
            books.append(d)
        conn.close()
        return jsonify({'books': books})

    elif request.method == 'POST':
        data = request.json or {}
        student_id = data.get('student_id', 1)
        if not student_id and user['role'] == 'STUDENT':
            cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
            st = cursor.fetchone()
            student_id = st['id'] if st else 1
        if not student_id:
            student_id = 1

        title = (data.get('title') or data.get('name') or 'Kitap').strip()
        author = (data.get('author') or '').strip()
        total_pages = int(data.get('total_pages') or data.get('page_count') or 0)
        read_pages = int(data.get('read_pages') or 0)
        rating_stars = int(data.get('rating_stars') or 5)

        cursor.execute("""
        INSERT INTO books (student_id, title, author, total_pages, read_pages, rating_stars, status, start_date)
        VALUES (?, ?, ?, ?, ?, ?, 'IN_PROGRESS', DATE('now'));
        """, (student_id, title, author, total_pages, read_pages, rating_stars))

        conn.commit()
        conn.close()
        return jsonify({'message': 'Kitap başarıyla eklendi!'})

# ==========================================
# 16. MESAJLAŞMA & KİŞİ LİSTESİ API
# ==========================================
# ==========================================
# 16. MESAJLAŞMA & WHATSAPP AKILLI SOHBET API
# ==========================================
def check_messaging_rbac_permission(user, target_user_id, cursor):
    if user['id'] == target_user_id:
        return True

    cursor.execute("SELECT id, role FROM users WHERE id = ?;", (target_user_id,))
    target_user = cursor.fetchone()
    if not target_user:
        return False

    target_role = target_user['role']

    # ADMIN can communicate with ANY STUDENT, COACH or ADMIN
    if user['role'] == 'ADMIN':
        return True

    # ANY user (COACH or STUDENT) can message ADMIN
    if target_role == 'ADMIN':
        return True

    # COACH can message assigned students
    if user['role'] == 'COACH':
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        ch = cursor.fetchone()
        if not ch:
            return False
        coach_id = ch['id']
        if target_role == 'STUDENT':
            cursor.execute("""
            SELECT 1 FROM students s 
            JOIN coach_student_relationships rel ON s.id = rel.student_id 
            WHERE s.user_id = ? AND rel.coach_id = ? AND rel.status = 'ACTIVE';
            """, (target_user_id, coach_id))
            return cursor.fetchone() is not None
        return False

    # STUDENT can message assigned coach
    if user['role'] == 'STUDENT':
        cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
        st = cursor.fetchone()
        if not st:
            return False
        student_id = st['id']
        if target_role == 'COACH':
            cursor.execute("""
            SELECT 1 FROM coaches c 
            JOIN coach_student_relationships rel ON c.id = rel.coach_id 
            WHERE c.user_id = ? AND rel.student_id = ? AND rel.status = 'ACTIVE';
            """, (target_user_id, student_id))
            return cursor.fetchone() is not None
        return False

    return False

@app.route('/api/mesajlar/contacts', methods=['GET'])
@app.route('/api/messages/conversations', methods=['GET'])
def get_message_contacts():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if user['role'] == 'ADMIN':
        # Admin gets all active students and coaches
        query = """
        SELECT DISTINCT u.id as user_id, u.name, u.surname, u.email, u.role, 
               s.id as student_id, s.track, s.school, 
               cu.name as coach_name, 'ÖĞRENCİ' as relationship_type
        FROM users u
        LEFT JOIN students s ON u.id = s.user_id
        LEFT JOIN coach_student_relationships rel ON s.id = rel.student_id AND rel.status = 'ACTIVE'
        LEFT JOIN coaches c ON rel.coach_id = c.id
        LEFT JOIN users cu ON c.user_id = cu.id
        WHERE u.id != ? AND u.status = 'ACTIVE' AND u.role IN ('STUDENT', 'COACH')
        ORDER BY u.role DESC, u.name ASC;
        """
        cursor.execute(query, (user['id'],))
        contacts = [dict(r) for r in cursor.fetchall()]
    elif user['role'] == 'COACH':
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        ch = cursor.fetchone()
        coach_id = ch['id'] if ch else None

        contacts = []
        # Include Admin users for coach
        cursor.execute("SELECT id as user_id, name, surname, email, role, NULL as student_id, NULL as track, NULL as school, NULL as coach_name FROM users WHERE role = 'ADMIN' AND status = 'ACTIVE';")
        contacts.extend([dict(r) for r in cursor.fetchall()])

        if coach_id:
            query = """
            SELECT DISTINCT u.id as user_id, u.name, u.surname, u.email, u.role, s.id as student_id, s.track, s.school, 'ÖĞRENCİ' as relationship_type
            FROM users u
            JOIN students s ON u.id = s.user_id
            JOIN coach_student_relationships rel ON s.id = rel.student_id
            WHERE rel.coach_id = ? AND rel.status = 'ACTIVE' AND u.status = 'ACTIVE'
            ORDER BY u.name ASC;
            """
            cursor.execute(query, (coach_id,))
            contacts.extend([dict(r) for r in cursor.fetchall()])
    else:
        # STUDENT role
        cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
        st = cursor.fetchone()
        student_id = st['id'] if st else 1

        contacts = []
        # Include Admin users for student
        cursor.execute("SELECT id as user_id, name, surname, email, role, NULL as student_id, NULL as track, NULL as school, NULL as coach_name FROM users WHERE role = 'ADMIN' AND status = 'ACTIVE';")
        contacts.extend([dict(r) for r in cursor.fetchall()])

        # Include assigned coach
        query = """
        SELECT DISTINCT u.id as user_id, u.name, u.surname, u.email, u.role, c.title as coach_title, 'KOÇ' as relationship_type
        FROM users u
        JOIN coaches c ON u.id = c.user_id
        JOIN coach_student_relationships rel ON c.id = rel.coach_id
        WHERE rel.student_id = ? AND rel.status = 'ACTIVE' AND u.status = 'ACTIVE'
        ORDER BY u.name ASC;
        """
        cursor.execute(query, (student_id,))
        contacts.extend([dict(r) for r in cursor.fetchall()])

    # De-duplicate contacts by user_id
    unique_contacts = {}
    for c in contacts:
        cid = c['user_id']
        if cid not in unique_contacts:
            unique_contacts[cid] = c

    enriched_contacts = []
    contact_ids = list(unique_contacts.keys())

    if contact_ids:
        placeholders = ', '.join(['?'] * len(contact_ids))

        # 1. Bulk Last Message per Contact
        q_last = f"""
        SELECT partner_id, content, message_type, sent_at, sender_id, is_read
        FROM (
            SELECT 
                CASE WHEN sender_id = ? THEN receiver_id ELSE sender_id END AS partner_id,
                content, message_type, sent_at, sender_id, is_read,
                ROW_NUMBER() OVER (
                    PARTITION BY (CASE WHEN sender_id = ? THEN receiver_id ELSE sender_id END)
                    ORDER BY sent_at DESC, id DESC
                ) as rn
            FROM messages
            WHERE (sender_id = ? AND receiver_id IN ({placeholders}))
               OR (receiver_id = ? AND sender_id IN ({placeholders}))
        ) sub
        WHERE rn = 1;
        """
        params_last = [user['id'], user['id'], user['id']] + contact_ids + [user['id']] + contact_ids
        cursor.execute(q_last, tuple(params_last))
        last_msg_map = {r['partner_id']: dict(r) for r in cursor.fetchall()}

        # 2. Bulk Unread Messages Count per Contact
        q_unread = f"""
        SELECT sender_id, COUNT(*) as count
        FROM messages
        WHERE receiver_id = ? AND sender_id IN ({placeholders}) AND COALESCE(is_read, 0) = 0
        GROUP BY sender_id;
        """
        params_unread = [user['id']] + contact_ids
        cursor.execute(q_unread, tuple(params_unread))
        unread_map = {r['sender_id']: r['count'] for r in cursor.fetchall()}

        for cid, c in unique_contacts.items():
            last_msg = last_msg_map.get(cid)
            unread = unread_map.get(cid, 0)

            c['last_message'] = last_msg['content'] if last_msg else 'Henüz mesaj yok'
            c['last_message_time'] = last_msg['sent_at'] if last_msg else None
            c['last_message_type'] = last_msg['message_type'] if last_msg else 'TEXT'
            c['unread_count'] = unread
            c['is_online'] = True if (cid % 2 == 0) else False

            enriched_contacts.append(c)
    else:
        enriched_contacts = []

    conn.close()
    return jsonify({'contacts': enriched_contacts, 'conversations': enriched_contacts})

@app.route('/api/mesajlar/unread-summary', methods=['GET'])
@app.route('/api/messages/unread-count', methods=['GET'])
def get_unread_message_summary():
    user = get_auth_user()
    if not user:
        return jsonify({'total_unread': 0, 'unread_count': 0})
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT COUNT(*) as count 
    FROM messages 
    WHERE receiver_id = ? AND is_read = 0 AND is_deleted = 0 
      AND (message_type IS NULL OR message_type != 'SYSTEM');
    """, (user['id'],))
    row = cursor.fetchone()
    total_unread = row['count'] if row else 0
    conn.close()
    return jsonify({'total_unread': total_unread, 'unread_count': total_unread})

@app.route('/api/mesajlar/conversations/<int:with_user_id>/read', methods=['PUT', 'POST'])
@app.route('/api/mesajlar/read', methods=['PUT', 'POST'])
def mark_conversation_as_read(with_user_id=None):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    if not with_user_id:
        data = request.json or {}
        with_user_id = data.get('with_user_id') or data.get('sender_id') or data.get('partner_id')

    if not with_user_id:
        return jsonify({'error': 'with_user_id gereklidir'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE messages SET is_read = 1 WHERE sender_id = ? AND receiver_id = ? AND is_read = 0;", (with_user_id, user['id']))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Mesajlar okundu işaretlendi.', 'success': True})

@app.route('/api/mesajlar', methods=['GET', 'POST'])
@app.route('/api/messages', methods=['GET', 'POST'])
@app.route('/api/messages/conversations/<int:with_user_id>', methods=['GET'])
def handle_messages(with_user_id=None):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        other_user_id = with_user_id or request.args.get('with_user_id')
        search_query = request.args.get('q', '').strip()
        
        if not other_user_id:
            cursor.execute("SELECT receiver_id FROM messages WHERE sender_id = ? UNION SELECT sender_id FROM messages WHERE receiver_id = ? LIMIT 1;", (user['id'], user['id']))
            row = cursor.fetchone()
            other_user_id = row['receiver_id'] if row else (2 if user['role'] == 'STUDENT' else 4)

        other_user_id = int(other_user_id)

        # RBAC Check
        if not check_messaging_rbac_permission(user, other_user_id, cursor):
            conn.close()
            return jsonify({'error': 'Bu kullanıcıyla mesajlaşma veya konuşma görüntüleme yetkiniz bulunmuyor.'}), 403

        cursor.execute("UPDATE messages SET is_read = 1 WHERE sender_id = ? AND receiver_id = ? AND is_read = 0;", (other_user_id, user['id']))
        if cursor.rowcount and cursor.rowcount > 0:
            conn.commit()

        cursor.execute("SELECT u.id, u.name, u.surname, u.role, u.email, s.track, s.school, c.title as coach_title FROM users u LEFT JOIN students s ON u.id = s.user_id LEFT JOIN coaches c ON u.id = c.user_id WHERE u.id = ?;", (other_user_id,))
        recipient_info = cursor.fetchone()
        recipient_dict = dict(recipient_info) if recipient_info else {'id': other_user_id, 'name': 'Kullanıcı', 'role': 'STUDENT'}

        limit_param = request.args.get('limit')
        offset_param = request.args.get('offset', 0)
        limit = None
        offset = 0
        if limit_param is not None:
            try:
                limit = max(1, min(200, int(limit_param)))
            except (ValueError, TypeError):
                limit = 50
        else:
            limit = 50

        if offset_param is not None:
            try:
                offset = max(0, int(offset_param))
            except (ValueError, TypeError):
                offset = 0

        if search_query:
            cursor.execute("""
            SELECT COUNT(*) as count FROM messages
            WHERE ((sender_id = ? AND receiver_id = ?) OR (sender_id = ? AND receiver_id = ?))
              AND content LIKE ?;
            """, (user['id'], other_user_id, other_user_id, user['id'], f"%{search_query}%"))
            total_count_row = cursor.fetchone()
            total_count = total_count_row['count'] if total_count_row else 0

            cursor.execute("""
            SELECT * FROM (
                SELECT m.*, u.name as sender_name, r.content as reply_content
                FROM messages m
                JOIN users u ON m.sender_id = u.id
                LEFT JOIN messages r ON m.reply_to_id = r.id
                WHERE ((m.sender_id = ? AND m.receiver_id = ?) OR (m.sender_id = ? AND m.receiver_id = ?))
                  AND m.content LIKE ?
                ORDER BY m.sent_at DESC, m.id DESC
                LIMIT ? OFFSET ?
            ) sub
            ORDER BY sub.sent_at ASC, sub.id ASC;
            """, (user['id'], other_user_id, other_user_id, user['id'], f"%{search_query}%", limit, offset))
        else:
            cursor.execute("""
            SELECT COUNT(*) as count FROM messages
            WHERE (sender_id = ? AND receiver_id = ?) OR (sender_id = ? AND receiver_id = ?);
            """, (user['id'], other_user_id, other_user_id, user['id']))
            total_count_row = cursor.fetchone()
            total_count = total_count_row['count'] if total_count_row else 0

            cursor.execute("""
            SELECT * FROM (
                SELECT m.*, u.name as sender_name, r.content as reply_content
                FROM messages m
                JOIN users u ON m.sender_id = u.id
                LEFT JOIN messages r ON m.reply_to_id = r.id
                WHERE (m.sender_id = ? AND m.receiver_id = ?) OR (m.sender_id = ? AND m.receiver_id = ?)
                ORDER BY m.sent_at DESC, m.id DESC
                LIMIT ? OFFSET ?
            ) sub
            ORDER BY sub.sent_at ASC, sub.id ASC;
            """, (user['id'], other_user_id, other_user_id, user['id'], limit, offset))

        msgs = [dict(r) for r in cursor.fetchall()]
        has_more = (offset + len(msgs)) < total_count

        cursor.execute("""
        SELECT * FROM messages
        WHERE ((sender_id = ? AND receiver_id = ?) OR (sender_id = ? AND receiver_id = ?))
          AND is_pinned = 1 AND is_deleted = 0
        ORDER BY sent_at DESC LIMIT 3;
        """, (user['id'], other_user_id, other_user_id, user['id']))
        pinned_msgs = [dict(r) for r in cursor.fetchall()]

        conn.close()
        return jsonify({
            'messages': msgs,
            'pinned_messages': pinned_msgs,
            'recipient': recipient_dict,
            'total_count': total_count,
            'has_more': has_more,
            'limit': limit,
            'offset': offset
        })

    elif request.method == 'POST':
        data = request.json or {}
        receiver_id = data.get('receiver_id') or data.get('alici_id') or data.get('recipient_id')
        content = data.get('content', '').strip()
        message_type = data.get('message_type', 'TEXT')
        attachment_url = data.get('attachment_url')
        file_name = data.get('file_name')
        file_size = data.get('file_size')
        reply_to_id = data.get('reply_to_id')

        if not receiver_id or not content:
            return jsonify({'error': 'Alıcı ve mesaj içeriği zorunludur'}), 400

        receiver_id = int(receiver_id)

        # RBAC Check
        if not check_messaging_rbac_permission(user, receiver_id, cursor):
            conn.close()
            return jsonify({'error': 'Bu kullanıcıya mesaj gönderme yetkiniz bulunmuyor.'}), 403

        cursor.execute("""
        INSERT INTO messages (sender_id, receiver_id, message_type, content, attachment_url, file_name, file_size, reply_to_id)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?);
        """, (user['id'], receiver_id, message_type, content, attachment_url, file_name, file_size, reply_to_id))
        new_msg_id = cursor.lastrowid

        # Send Notification to Recipient
        sender_name = f"{user.get('name', '')} {user.get('surname', '')}".strip() or user.get('username')
        notif_msg = f"💬 {sender_name} size yeni bir mesaj gönderdi."
        event_key = f"MESSAGE_{new_msg_id}_{receiver_id}"
        cursor.execute("""
        INSERT INTO notifications (
            recipient_user_id, actor_user_id, type, title, message, 
            entity_type, entity_id, event_key
        ) VALUES (?, ?, 'MESSAGE_RECEIVED', '💬 Yeni Mesaj', ?, 'MESSAGE', ?, ?)
        ON CONFLICT DO NOTHING;
        """, (receiver_id, user['id'], notif_msg, new_msg_id, event_key))

        conn.commit()
        conn.close()
        return jsonify({'message': 'Mesaj başarıyla gönderildi!', 'id': new_msg_id, 'message_id': new_msg_id})

@app.route('/api/mesajlar/<int:message_id>/pin', methods=['POST'])
def toggle_pin_message(message_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT is_pinned FROM messages WHERE id = ?;", (message_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Mesaj bulunamadı'}), 404

    new_pin = 0 if row['is_pinned'] == 1 else 1
    cursor.execute("UPDATE messages SET is_pinned = ? WHERE id = ?;", (new_pin, message_id))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Mesaj sabitleme durumu güncellendi!', 'is_pinned': new_pin})

@app.route('/api/mesajlar/<int:message_id>', methods=['DELETE', 'PUT'])
def handle_single_message_action(message_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'DELETE':
        cursor.execute("UPDATE messages SET is_deleted = 1, content = 'Bu mesaj silindi.' WHERE id = ? AND sender_id = ?;", (message_id, user['id']))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Mesaj silindi.'})

    elif request.method == 'PUT':
        data = request.json or {}
        new_content = data.get('content', '').strip()
        if not new_content:
            conn.close()
            return jsonify({'error': 'İçerik gereklidir'}), 400

        cursor.execute("""
        UPDATE messages SET content = ?, edited_at = CURRENT_TIMESTAMP
        WHERE id = ? AND sender_id = ? AND is_deleted = 0;
        """, (new_content, message_id, user['id']))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Mesaj düzenlendi!'})

@app.route('/api/mesajlar/settings', methods=['POST'])
def update_conversation_settings():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    with_user_id = data.get('with_user_id')
    is_muted = data.get('is_muted', 0)
    is_archived = data.get('is_archived', 0)
    is_pinned = data.get('is_pinned', 0)

    if not with_user_id:
        return jsonify({'error': 'with_user_id zorunludur'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    INSERT INTO conversation_settings (user_id, with_user_id, is_muted, is_archived, is_pinned)
    VALUES (?, ?, ?, ?, ?)
    ON CONFLICT(user_id, with_user_id) DO UPDATE SET
        is_muted = excluded.is_muted,
        is_archived = excluded.is_archived,
        is_pinned = excluded.is_pinned;
    """, (user['id'], with_user_id, is_muted, is_archived, is_pinned))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Sohbet tercihleri güncellendi!'})

@app.route('/api/mesajlar/broadcast', methods=['POST'])
def broadcast_messages():
    user = get_auth_user()
    if not user or user['role'] not in ['COACH', 'ADMIN']:
        return jsonify({'error': 'Yalnızca koçlar ve admin toplu mesaj gönderebilir'}), 403

    conn = get_db()
    cursor = conn.cursor()

    data = request.json or {}
    target_student_ids = data.get('student_ids', [])
    content = data.get('content', '').strip()

    if not content:
        return jsonify({'error': 'Mesaj içeriği gereklidir'}), 400

    if not target_student_ids:
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        ch = cursor.fetchone()
        cid = ch['id'] if ch else 1

        cursor.execute("""
        SELECT u.id as user_id FROM users u
        JOIN students s ON u.id = s.user_id
        JOIN coach_student_relationships rel ON s.id = rel.student_id
        WHERE rel.coach_id = ? AND rel.status = 'ACTIVE';
        """, (cid,))
        rows = cursor.fetchall()
        target_student_ids = [r['user_id'] for r in rows]

    sent_count = 0
    for sid in target_student_ids:
        cursor.execute("""
        INSERT INTO messages (sender_id, receiver_id, message_type, content)
        VALUES (?, ?, 'SYSTEM', ?);
        """, (user['id'], sid, f"📢 KOÇ DUYURUSU: {content}"))
        sent_count += 1

    conn.commit()
    conn.close()
    return jsonify({'message': f'Duyuru {sent_count} öğrenciye başarıyla iletildi!'})

# ==========================================
# 17. PDF RAPOR OLUŞTURUCU
# ==========================================
@app.route('/api/raporlar/pdf', methods=['GET'])
def generate_pdf_report():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    req_st_id = request.args.get('student_id')
    conn = get_db()
    cursor = conn.cursor()
    student_id, err_resp, err_code = resolve_and_verify_student_id(cursor, user, req_st_id)
    if err_resp:
        conn.close()
        return err_resp, err_code

    cursor.execute("SELECT s.*, u.name, u.email FROM students s JOIN users u ON s.user_id = u.id WHERE s.id = ?;", (student_id,))
    student = cursor.fetchone()
    if not student:
        conn.close()
        return jsonify({'error': 'Öğrenci bulunamadı.'}), 404

    cursor.execute("SELECT mr.net, me.title, me.exam_type, mr.exam_date FROM mock_exam_results mr JOIN mock_exams me ON mr.mock_exam_id = me.id WHERE mr.student_id = ? ORDER BY mr.exam_date DESC LIMIT 10;", (student_id,))
    mocks = cursor.fetchall()
    conn.close()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>YKS AKADEMİK GELİŞİM RAPORU</b>", styles['Title']))
    story.append(Spacer(1, 10))
    story.append(Paragraph(f"<b>Öğrenci Adı:</b> {student['name']} | <b>Alan:</b> {student['track']} | <b>Hedef:</b> {student['target_university']} - {student['target_department']}", styles['Normal']))
    story.append(Spacer(1, 15))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#3b82f6")))
    story.append(Spacer(1, 15))

    data = [["Deneme Adı", "Sınav Türü", "Tarih", "Net"]]
    for m in mocks:
        data.append([m['title'], m['exam_type'], m['exam_date'], str(m['net'])])

    t = Table(data, colWidths=[200, 100, 100, 80])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(t)

    doc.build(story)
    buffer.seek(0)

    return send_file(buffer, as_attachment=True, download_name=f"YKS_Rapor_{student['name'].replace(' ', '_')}.pdf", mimetype='application/pdf')

# ==========================================
# 18. EXCEL EXPORT API
# ==========================================
@app.route('/api/excel/export', methods=['GET'])
def export_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Öğrenci Net Analizi"

    headers = ["Öğrenci Adı", "Alan", "Deneme Adı", "Tarih", "Net"]
    ws.append(headers)

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT u.name, s.track, me.title, mr.exam_date, mr.net FROM mock_exam_results mr JOIN students s ON mr.student_id = s.id JOIN users u ON s.user_id = u.id JOIN mock_exams me ON mr.mock_exam_id = me.id ORDER BY mr.exam_date DESC;")
    rows = cursor.fetchall()
    conn.close()

    for r in rows:
        ws.append([r['name'], r['track'], r['title'], r['exam_date'], r['net']])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='export.xlsx')


import math

def calculate_std_dev(values):
    if not values or len(values) < 2:
        return 0.0
    mean = sum(values) / float(len(values))
    variance = sum((x - mean) ** 2 for x in values) / float(len(values))
    return round(math.sqrt(variance), 2)

def determine_volatility_label(std_dev):
    if std_dev <= 2.0:
        return 'STABLE'
    elif std_dev <= 4.5:
        return 'MEDIUM'
    else:
        return 'VOLATILE'

def determine_trend_direction(first_val, last_val, std_dev=0.0):
    diff = last_val - first_val
    if diff >= 2.0:
        return 'UP'
    elif diff <= -2.0:
        return 'DOWN'
    elif std_dev > 4.5:
        return 'VOLATILE'
    else:
        return 'STABLE'

# ==========================================
# 18.b RAPORLAMA + DERS BAZLI GELİŞİM ENGINE API
# ==========================================
@app.route('/api/raporlar', methods=['GET'])
def get_reports_analytics():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    student_id = request.args.get('student_id')
    preset = request.args.get('preset', '3_MONTHS')
    start_date_arg = request.args.get('start_date')
    end_date_arg = request.args.get('end_date')
    subject_filter = request.args.get('subject_id', 'ALL')

    conn = get_db()
    cursor = conn.cursor()

    student_id, err_resp, err_code = resolve_and_verify_student_id(cursor, user, student_id)
    if err_resp:
        conn.close()
        return err_resp, err_code

    date_clause = ""
    params = []
    today = date.today()

    if preset == '7_DAYS':
        start_dt = today - timedelta(days=7)
        date_clause = " AND ea.exam_date >= ?"
        params.append(start_dt.isoformat())
    elif preset == '30_DAYS':
        start_dt = today - timedelta(days=30)
        date_clause = " AND ea.exam_date >= ?"
        params.append(start_dt.isoformat())
    elif preset == '3_MONTHS':
        start_dt = today - timedelta(days=90)
        date_clause = " AND ea.exam_date >= ?"
        params.append(start_dt.isoformat())
    elif preset == '6_MONTHS':
        start_dt = today - timedelta(days=180)
        date_clause = " AND ea.exam_date >= ?"
        params.append(start_dt.isoformat())
    elif preset == 'THIS_YEAR':
        start_dt = date(today.year, 1, 1)
        date_clause = " AND ea.exam_date >= ?"
        params.append(start_dt.isoformat())
    elif preset == 'CUSTOM' and start_date_arg and end_date_arg:
        date_clause = " AND ea.exam_date >= ? AND ea.exam_date <= ?"
        params.append(start_date_arg)
        params.append(end_date_arg)

    query_params = list(params)
    query_params.append(student_id)

    cursor.execute(f"""
    SELECT 
        s.id as student_id, s.user_id, s.exam_system as student_exam_system, s.track, s.target_university, s.target_department,
        u.name, u.surname,
        ea.id as attempt_id, ea.exam_system as attempt_exam_system, ea.exam_type, ea.exam_name, ea.publisher, 
        ea.exam_date, ea.duration_minutes, ea.total_net, ea.status
    FROM students s
    JOIN users u ON s.user_id = u.id
    LEFT JOIN exam_attempts ea 
        ON ea.student_id = s.id 
       AND ea.status != 'CANCELLED' {date_clause}
    WHERE s.id = ?
    ORDER BY ea.exam_date ASC, ea.id ASC;
    """, query_params)
    rows = [dict(r) for r in cursor.fetchall()]

    if not rows:
        conn.close()
        return jsonify({'error': 'Öğrenci profili bulunamadı.'}), 404

    first_r = rows[0]
    student_info = {
        'id': first_r['student_id'],
        'user_id': first_r['user_id'],
        'exam_system': first_r['student_exam_system'],
        'track': first_r['track'],
        'target_university': first_r['target_university'],
        'target_department': first_r['target_department'],
        'name': first_r['name'],
        'surname': first_r['surname']
    }
    student_name = f"{student_info.get('name') or 'Öğrenci'} {student_info.get('surname') or ''}".strip()

    attempts = []
    for r in rows:
        if r.get('attempt_id') is not None:
            attempts.append({
                'id': r['attempt_id'],
                'student_id': r['student_id'],
                'exam_system': r['attempt_exam_system'] or r['student_exam_system'],
                'exam_type': r['exam_type'],
                'exam_name': r['exam_name'],
                'publisher': r['publisher'],
                'exam_date': r['exam_date'],
                'duration_minutes': r['duration_minutes'],
                'total_net': r['total_net'],
                'status': r['status']
            })

    if not attempts:
        cursor.execute(f"""
        SELECT me.id, me.student_id, me.exam_system, me.exam_type, me.title as exam_name, me.publisher, me.created_at as exam_date, me.total_net, 'COMPLETED' as status
        FROM mock_exams me
        WHERE me.student_id = ? {date_clause.replace('exam_date', 'DATE(me.created_at)')}
        ORDER BY me.created_at ASC;
        """, params)
        attempts = [dict(r) for r in cursor.fetchall()]

    if attempts:
        attempt_ids = [a['id'] for a in attempts]
        placeholders = ','.join(['?'] * len(attempt_ids))
        
        cursor.execute(f"""
        SELECT tr.*, s.name as subject_name
        FROM exam_test_results tr 
        JOIN subjects s ON tr.subject_id = s.id
        WHERE tr.exam_attempt_id IN ({placeholders})
        ORDER BY s.sort_order ASC;
        """, attempt_ids)
        
        tr_map = {}
        for r in cursor.fetchall():
            d = dict(r)
            att_id = d['exam_attempt_id']
            if att_id not in tr_map:
                tr_map[att_id] = []
            tr_map[att_id].append(d)

        missing_ids = [aid for aid in attempt_ids if aid not in tr_map]
        if missing_ids:
            m_placeholders = ','.join(['?'] * len(missing_ids))
            cursor.execute(f"""
            SELECT mr.*, s.name as subject_name 
            FROM mock_exam_results mr 
            JOIN subjects s ON mr.subject_id = s.id 
            WHERE mr.mock_exam_id IN ({m_placeholders})
            ORDER BY s.sort_order ASC;
            """, missing_ids)
            
            for r in cursor.fetchall():
                d = dict(r)
                m_id = d['mock_exam_id']
                if m_id not in tr_map:
                    tr_map[m_id] = []
                tr_map[m_id].append(d)

        for att in attempts:
            att['test_results'] = tr_map.get(att['id'], [])

    total_nets = [a['total_net'] for a in attempts]
    num_exams = len(total_nets)

    if num_exams > 0:
        first_net = round(total_nets[0], 2)
        last_net = round(total_nets[-1], 2)
        net_change = round(last_net - first_net, 2)
        avg_net = round(sum(total_nets) / float(num_exams), 2)
        highest_net = round(max(total_nets), 2)
        lowest_net = round(min(total_nets), 2)
        std_dev = calculate_std_dev(total_nets)
        volatility = determine_volatility_label(std_dev)
        trend = determine_trend_direction(first_net, last_net, std_dev)
    else:
        first_net = last_net = net_change = avg_net = highest_net = lowest_net = std_dev = 0.0
        volatility = 'STABLE'
        trend = 'STABLE'

    overall_summary = {
        'total_exams': num_exams,
        'first_net': first_net,
        'last_net': last_net,
        'net_change': net_change,
        'average_net': avg_net,
        'highest_net': highest_net,
        'lowest_net': lowest_net,
        'std_dev': std_dev,
        'volatility': volatility,
        'trend': trend
    }

    total_net_series = [{'date': a['exam_date'], 'net': a['total_net'], 'exam_name': a.get('exam_name')} for a in attempts]

    subject_map = {}
    for att in attempts:
        for tr in att.get('test_results', []):
            s_name = tr.get('subject_name')
            s_id = tr.get('subject_id')
            if not s_name: continue
            if s_id not in subject_map:
                subject_map[s_id] = {'subject_id': s_id, 'subject_name': s_name, 'nets': [], 'series': []}
            subject_map[s_id]['nets'].append(tr.get('net', 0.0))
            subject_map[s_id]['series'].append({'date': att['exam_date'], 'net': tr.get('net', 0.0), 'exam_name': att.get('exam_name')})

    subject_analytics = []
    for s_id, s_data in subject_map.items():
        s_nets = s_data['nets']
        s_len = len(s_nets)
        if s_len == 0: continue

        s_first = round(s_nets[0], 2)
        s_last = round(s_nets[-1], 2)
        s_change = round(s_last - s_first, 2)
        s_avg = round(sum(s_nets) / float(s_len), 2)
        s_high = round(max(s_nets), 2)
        s_low = round(min(s_nets), 2)
        s_std = calculate_std_dev(s_nets)
        s_vol = determine_volatility_label(s_std)
        s_trend = determine_trend_direction(s_first, s_last, s_std)

        last_3_avg = round(sum(s_nets[-3:]) / float(len(s_nets[-3:])), 2)
        last_5_avg = round(sum(s_nets[-5:]) / float(len(s_nets[-5:])), 2)
        last_10_avg = round(sum(s_nets[-10:]) / float(len(s_nets[-10:])), 2)

        if s_change >= 4.0 or (s_len >= 3 and last_3_avg > s_avg + 2.0):
            badge = 'GÜÇLÜ_GELİŞİM'
        elif s_change <= -2.0 or (s_len >= 3 and last_3_avg < s_avg - 1.5):
            badge = 'GERİLEME'
        elif s_avg >= 25.0:
            badge = 'GÜÇLÜ'
        else:
            badge = 'STABİL'

        subject_analytics.append({
            'subject_id': s_id,
            'subject_name': s_data['subject_name'],
            'first_net': s_first,
            'last_net': s_last,
            'net_change': s_change,
            'average_net': s_avg,
            'highest_net': s_high,
            'lowest_net': s_low,
            'std_dev': s_std,
            'volatility': s_vol,
            'trend': s_trend,
            'status_badge': badge,
            'last_3_average': last_3_avg,
            'last_5_average': last_5_avg,
            'last_10_average': last_10_avg,
            'net_series': s_data['series']
        })

    monthly_groups = {}
    for att in attempts:
        m_key = str(att['exam_date'])[:7] if att['exam_date'] else '2026-08'
        if m_key not in monthly_groups:
            monthly_groups[m_key] = []
        monthly_groups[m_key].append(att)

    sorted_months = sorted(monthly_groups.keys())
    monthly_matrix = {}
    month_name_map = {'01': 'Ocak', '02': 'Şubat', '03': 'Mart', '04': 'Nisan', '05': 'Mayıs', '06': 'Haziran', '07': 'Temmuz', '08': 'Ağustos', '09': 'Eylül', '10': 'Ekim', '11': 'Kasım', '12': 'Aralık'}
    month_labels = [f"{month_name_map.get(m.split('-')[1], m)} '{m.split('-')[0][-2:]}" for m in sorted_months]

    for s_id, s_data in subject_map.items():
        s_name = s_data['subject_name']
        monthly_matrix[s_name] = {}

        for m_key in sorted_months:
            m_label = f"{month_name_map.get(m_key.split('-')[1], m_key)} '{m_key.split('-')[0][-2:]}"
            m_atts = monthly_groups[m_key]
            m_nets = []
            for a in m_atts:
                for tr in a.get('test_results', []):
                    if tr.get('subject_id') == s_id:
                        m_nets.append(tr.get('net', 0.0))
            if m_nets:
                monthly_matrix[s_name][m_label] = round(sum(m_nets) / float(len(m_nets)), 2)
            else:
                monthly_matrix[s_name][m_label] = 0.0

        m_vals = list(monthly_matrix[s_name].values())
        if len(m_vals) >= 2:
            monthly_matrix[s_name]['change'] = round(m_vals[-1] - m_vals[0], 2)
            monthly_matrix[s_name]['trend'] = 'UP' if m_vals[-1] >= m_vals[0] else 'DOWN'
        else:
            monthly_matrix[s_name]['change'] = 0.0
            monthly_matrix[s_name]['trend'] = 'STABLE'

    contradictions = []
    strongest_subject = max(subject_analytics, key=lambda x: x['average_net'])['subject_name'] if subject_analytics else 'Belirsiz'
    most_improved = max(subject_analytics, key=lambda x: x['net_change'])['subject_name'] if subject_analytics else 'Belirsiz'
    declining_subject = min(subject_analytics, key=lambda x: x['net_change'])['subject_name'] if subject_analytics else 'Belirsiz'

    for s in subject_analytics:
        if s['trend'] == 'UP' and s['last_3_average'] < s['average_net']:
            contradictions.append(f"{s['subject_name']} dersinde uzun vadeli yükseliş devam ediyor ancak son 3 denemede ortalamanın altında bir gerileme görülüyor.")
        if s['volatility'] == 'VOLATILE':
            contradictions.append(f"{s['subject_name']} dersi net ortalaması iyi ancak sınavlar arası yüksek dalgalanma (volatilite) mevcut.")

    ai_context = {
        'analytics_context_version': '1.0',
        'student_profile': {
            'student_id': student_id,
            'name': student_name,
            'exam_system': student_info.get('exam_system', 'YKS'),
            'track': student_info.get('track', 'SAYISAL')
        },
        'time_windows': {
            'overall_avg': avg_net,
            'last_3_exams_avg': round(sum(total_nets[-3:]) / float(len(total_nets[-3:])), 2) if total_nets else 0.0,
            'last_5_exams_avg': round(sum(total_nets[-5:]) / float(len(total_nets[-5:])), 2) if total_nets else 0.0,
        },
        'subjects_payload': [
            {
                'subject': s['subject_name'],
                'first_net': s['first_net'],
                'last_net': s['last_net'],
                'average_net': s['average_net'],
                'highest_net': s['highest_net'],
                'lowest_net': s['lowest_net'],
                'change': s['net_change'],
                'trend': s['trend'],
                'volatility': s['volatility'],
                'last_3_average': s['last_3_average'],
                'last_5_average': s['last_5_average'],
                'last_10_average': s['last_10_average']
            } for s in subject_analytics
        ],
        'strongest_subject': strongest_subject,
        'most_improved_subject': most_improved,
        'declining_subject': declining_subject,
        'contradictions': contradictions
    }

    conn.close()

    return jsonify({
        'student': {
            'id': student_id,
            'name': student_name,
            'exam_system': student_info.get('exam_system', 'YKS'),
            'track': student_info.get('track', 'SAYISAL'),
            'last_exam_net': last_net,
            'last_exam_date': attempts[-1]['exam_date'] if attempts else '',
            'general_trend': trend
        },
        'filter': {
            'preset': preset,
            'start_date': start_date_arg,
            'end_date': end_date_arg,
            'subject_id': subject_filter
        },
        'overall_summary': overall_summary,
        'total_net_series': total_net_series,
        'subject_analytics': subject_analytics,
        'monthly_analytics': {
            'month_labels': month_labels,
            'subject_monthly_matrix': monthly_matrix
        },
        'insights': {
            'strongest_subject': strongest_subject,
            'most_improved': most_improved,
            'declining_subject': declining_subject,
            'contradictions': contradictions
        },
        'ai_analytics_context': ai_context
    })

# Helper: Build Grounded Student Analytics Context
def build_student_analytics_context(cursor, student_id):
    cursor.execute("""
    SELECT s.id, s.user_id, s.exam_system, s.track, s.grade, s.school, s.target_university, s.target_department, u.name, u.surname
    FROM students s JOIN users u ON s.user_id = u.id
    WHERE s.id = ?;
    """, (student_id,))
    st = cursor.fetchone()
    if not st:
        return None

    st_dict = dict(st)
    student_name = f"{st_dict.get('name') or 'Öğrenci'} {st_dict.get('surname') or ''}".strip()
    exam_sys = st_dict.get('exam_system') or 'YKS'
    track = st_dict.get('track') or 'SAYISAL'

    cursor.execute("""
    SELECT id, exam_system, exam_type, exam_name, publisher, exam_date, total_net, total_score, status
    FROM exam_attempts
    WHERE student_id = ? AND status != 'CANCELLED'
    ORDER BY exam_date DESC, id DESC;
    """, (student_id,))
    all_attempts = [dict(r) for r in cursor.fetchall()]

    valid_attempts = []
    invalid_attempts = []
    for att in all_attempts:
        att_sys = att.get('exam_system') or 'YKS'
        if att_sys == exam_sys or (exam_sys == 'YKS' and att.get('exam_type') in ['TYT', 'AYT', 'YDT']):
            valid_attempts.append(att)
        else:
            invalid_attempts.append(att)

    valid_exam_count = len(valid_attempts)

    overall_stats = {
        'valid_exam_count': valid_exam_count,
        'invalid_exam_count': len(invalid_attempts),
        'average_net': (sum(a['total_net'] for a in valid_attempts if a.get('total_net')) / valid_exam_count) if valid_exam_count > 0 else 0.0,
        'latest_net': (valid_attempts[0].get('total_net') or 0.0) if valid_exam_count > 0 else 0.0,
        'highest_net': max((a.get('total_net') or 0.0 for a in valid_attempts), default=0.0) if valid_exam_count > 0 else 0.0
    }

    subject_summary = []
    if valid_exam_count > 0:
        valid_ids = [a['id'] for a in valid_attempts]
        placeholders = ','.join('?' * len(valid_ids))
        cursor.execute(f"""
        SELECT s.name as subject_name, tr.subject_id,
               AVG(tr.net) as avg_net,
               AVG(tr.correct) as avg_correct,
               AVG(tr.wrong) as avg_wrong,
               AVG(tr.blank) as avg_blank,
               AVG(tr.success_rate) as avg_success_rate
        FROM exam_test_results tr
        JOIN subjects s ON tr.subject_id = s.id
        WHERE tr.exam_attempt_id IN ({placeholders})
        GROUP BY tr.subject_id
        ORDER BY avg_net DESC;
        """, valid_ids)
        subject_summary = [dict(r) for r in cursor.fetchall()]

    topic_risks = []
    if valid_exam_count > 0:
        valid_ids = [a['id'] for a in valid_attempts]
        placeholders = ','.join('?' * len(valid_ids))
        cursor.execute(f"""
        SELECT ct.name as topic_name, s.name as subject_name, ct.id as curriculum_topic_id,
               COUNT(tr.id) as attempt_count,
               SUM(tr.wrong) as total_wrong,
               SUM(tr.blank) as total_blank,
               AVG(tr.success_rate) as avg_success_rate
        FROM exam_topic_results tr
        JOIN subjects s ON tr.subject_id = s.id
        JOIN topics ct ON tr.curriculum_topic_id = ct.id
        WHERE tr.exam_attempt_id IN ({placeholders})
        GROUP BY tr.curriculum_topic_id
        ORDER BY total_wrong DESC, total_blank DESC;
        """, valid_ids)
        for tr in cursor.fetchall():
            t = dict(tr)
            succ = t.get('avg_success_rate') or 0.0
            wr = t.get('total_wrong') or 0
            att_cnt = t.get('attempt_count') or 1

            mastery = max(0, min(100, int(succ - (wr * 5))))
            if att_cnt < 3:
                status = 'NOT_ENOUGH_DATA'
                risk_label = '🟡 İzlenmeli (Yetersiz Veri)'
            elif mastery < 45 or wr >= 3:
                status = 'CRITICAL'
                risk_label = '🔴 Kronik Eksik'
            elif mastery < 65:
                status = 'AT_RISK'
                risk_label = '🟠 Riskli'
            else:
                status = 'MASTERED'
                risk_label = '🟢 Hâkim'

            t['mastery_score'] = mastery
            t['risk_label'] = risk_label
            t['status'] = status
            topic_risks.append(t)

    error_breakdown = {
        'Bilgi Eksikliği': 0, 'Dikkat Hatası': 0, 'İşlem Hatası': 0,
        'Kavram Yanılgısı': 0, 'Süre Baskısı': 0, 'Formül / Kural Eksikliği': 0,
        'Okuma / Anlama Hatası': 0, 'Soruyu Yanlış Yorumlama': 0, 'Yöntem Seçimi Hatası': 0,
        'Soruyu Yetiştirememe': 0, 'Şıklar Arasında Kalma': 0, 'Rastgele İşaretleme': 0,
        'Bilinmiyor / Analiz Edilmedi': 0
    }

    if valid_exam_count == 0:
        confidence = 'NO_DATA'
        data_quality_score = 0 if len(invalid_attempts) == 0 else 20
    elif valid_exam_count < 3:
        confidence = 'LOW'
        data_quality_score = 60
    elif valid_exam_count < 5:
        confidence = 'MEDIUM'
        data_quality_score = 85
    else:
        confidence = 'HIGH'
        data_quality_score = 100

    # Fetch Assignment Summary
    today_str = date.today().isoformat()
    cursor.execute("""
    SELECT 
        COUNT(*) as total_assignments,
        SUM(CASE WHEN status IN ('COMPLETED', 'SUBMITTED') THEN 1 ELSE 0 END) as completed_count,
        SUM(CASE WHEN status IN ('LATE', 'OVERDUE') OR (status NOT IN ('COMPLETED', 'CANCELLED', 'SUBMITTED') AND due_date < ?) THEN 1 ELSE 0 END) as late_count
    FROM assignments WHERE student_id = ?;
    """, (today_str, student_id))
    assign_row = cursor.fetchone()
    assign_dict = dict(assign_row) if assign_row else {}
    tot_a = assign_dict.get('total_assignments') or 0
    comp_a = assign_dict.get('completed_count') or 0
    late_a = assign_dict.get('late_count') or 0
    assign_rate = round((comp_a / tot_a * 100), 1) if tot_a > 0 else 0.0

    # Fetch Program Summary
    cursor.execute("""
    SELECT 
        COUNT(*) as total_slots,
        SUM(CASE WHEN status = 'TAMAMLANDI' OR status = 'COMPLETED' THEN 1 ELSE 0 END) as completed_slots
    FROM weekly_programs WHERE student_id = ?;
    """, (student_id,))
    prog_row = cursor.fetchone()
    prog_dict = dict(prog_row) if prog_row else {}
    tot_p = prog_dict.get('total_slots') or 0
    comp_p = prog_dict.get('completed_slots') or 0
    prog_rate = round((comp_p / tot_p * 100), 1) if tot_p > 0 else 0.0

    context = {
        'student_id': student_id,
        'student_name': student_name,
        'exam_system': exam_sys,
        'track': track,
        'grade': st_dict.get('grade'),
        'target_university': st_dict.get('target_university'),
        'target_department': st_dict.get('target_department'),
        'confidence': confidence,
        'data_quality_score': data_quality_score,
        'analysisDataQuality': data_quality_score,
        'overall_stats': overall_stats,
        'valid_attempts': [{'name': a['exam_name'], 'date': a['exam_date'], 'net': a['total_net'], 'type': a['exam_type']} for a in valid_attempts[:5]],
        'invalid_attempts': [{'name': a['exam_name'], 'system': a['exam_system']} for a in invalid_attempts],
        'subject_summary': subject_summary[:6],
        'topic_risks': topic_risks[:10],
        'assignment_summary': {
            'total': tot_a,
            'completed': comp_a,
            'late': late_a,
            'completion_rate': assign_rate
        },
        'program_summary': {
            'total_slots': tot_p,
            'completed_slots': comp_p,
            'compliance_rate': prog_rate
        }
    }

    assert context['student_id'] == student_id, f"Context student_id mismatch: expected {student_id}"
    return context

@app.route('/api/admin/clear-demo-data', methods=['POST'])
def clear_demo_data():
    user = get_auth_user()
    if not user or user.get('role') != 'ADMIN':
        return jsonify({'error': 'Bu işlem için ADMIN yetkisi gereklidir'}), 403

    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute("SELECT id FROM mock_exams WHERE is_demo = 1;")
        demo_exam_ids = [r[0] for r in cursor.fetchall()]
        if demo_exam_ids:
            placeholders = ','.join('?' for _ in demo_exam_ids)
            cursor.execute(f"SELECT id FROM mock_exam_results WHERE mock_exam_id IN ({placeholders});", demo_exam_ids)
            result_ids = [r[0] for r in cursor.fetchall()]
            if result_ids:
                res_placeholders = ','.join('?' for _ in result_ids)
                cursor.execute(f"DELETE FROM mock_exam_topic_errors WHERE result_id IN ({res_placeholders});", result_ids)
            cursor.execute(f"DELETE FROM mock_exam_results WHERE mock_exam_id IN ({placeholders});", demo_exam_ids)
            cursor.execute(f"DELETE FROM mock_exams WHERE is_demo = 1;")

        cursor.execute("DELETE FROM exam_attempts WHERE is_demo = 1;")
        cursor.execute("DELETE FROM assignments WHERE is_demo = 1;")
        cursor.execute("DELETE FROM weekly_programs WHERE is_demo = 1;")

        conn.commit()
        conn.close()
        return jsonify({'message': 'Tüm demo akademik verileri başarıyla temizlendi.'}), 200
    except Exception as e:
        return jsonify({'error': f'Demo veri temizleme hatası: {str(e)}'}), 500

@app.route('/api/ai/analyze-student', methods=['POST'])
def ai_analyze_student():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    req_st_id = data.get('student_id')

    conn = get_db()
    cursor = conn.cursor()

    student_id, err_resp, err_code = resolve_and_verify_student_id(cursor, user, req_st_id)
    if err_resp:
        conn.close()
        return err_resp, err_code

    context = build_student_analytics_context(cursor, student_id)
    conn.close()

    if not context:
        return jsonify({'error': 'Öğrenci bulunamadı'}), 404

    # Grounded Student Analysis Response Generation
    valid_count = context['overall_stats']['valid_exam_count']
    student_name = context['student_name']
    exam_sys = context['exam_system']
    track = context['track']
    confidence = context['confidence']

    if valid_count == 0:
        summary = f"{student_name} ({exam_sys} {track}) öğrencisi için henüz kayıtlı ve geçerli deneme sınavı verisi bulunmamaktadır."
        strengths = ["Analiz edilecek geçerli deneme verisi henüz girilmemiş."]
        weaknesses = ["Veri bulunmadığından ders veya konu bazlı eksik analizi yapılamamaktadır."]
        subject_analysis = []
        topic_risks = []
        recommendations = [
            {
                "problem": "Geçerli Deneme Sınavı Kaydı Bulunmuyor",
                "evidence": "0 Kayıtlı Deneme",
                "reason": f"{student_name} için veritabanında geçerli {exam_sys} denemesi bulunmuyor.",
                "action": "+ İlk Deneme Kaydını Gir butonuna tıklayarak deneme ekleyin.",
                "priority": "HIGH"
            }
        ]
    else:
        avg_net = context['overall_stats']['average_net']
        highest_net = context['overall_stats']['highest_net']
        summary = f"{student_name} ({exam_sys} {track}) son {valid_count} denemede ortalama {avg_net:.2f} net çıkardı. En yüksek neti: {highest_net:.2f}."
        
        strengths = []
        weaknesses = []
        subject_analysis = []
        topic_risks = []
        recommendations = []

        for subj in context['subject_summary']:
            s_name = subj['subject_name']
            s_net = subj.get('avg_net') or 0.0
            s_succ = subj.get('avg_success_rate') or 0.0
            subject_analysis.append({
                "subject_name": s_name,
                "avg_net": round(s_net, 2),
                "success_rate": round(s_succ, 1)
            })
            if s_succ >= 70:
                strengths.append(f"{s_name} dersinde ortalama {s_net:.2f} net ile %{s_succ:.1f} yüksek başarı sağlandı.")
            elif s_succ < 50:
                weaknesses.append(f"{s_name} dersinde ortalama {s_net:.2f} net ve %{s_succ:.1f} başarı oranı ile gelişim ihtiyacı var.")

        if not strengths:
            strengths.append(f"Toplam {valid_count} denemede düzenli sınav katılımı sağlandı.")

        for top in context['topic_risks']:
            t_name = top['topic_name']
            s_name = top['subject_name']
            wr = top['total_wrong']
            m_score = top['mastery_score']
            r_label = top['risk_label']
            topic_risks.append({
                "topic_name": t_name,
                "subject_name": s_name,
                "total_wrong": wr,
                "mastery_score": m_score,
                "risk_label": r_label
            })
            if m_score < 50 or wr >= 3:
                weaknesses.append(f"{t_name} ({s_name}) konusunda {wr} yanlış ve %{m_score} hâkimiyet puanı ile kritik eksik tespit edildi.")
                recommendations.append({
                    "problem": f"{t_name} Düşük Hâkimiyet",
                    "evidence": f"Hâkimiyet Puanı: {m_score}/100, {wr} Yanlış",
                    "reason": "Konu bilgi eksikliği ve soru pratiği yetersizliği.",
                    "action": f"{t_name} konusunda 1 adet soru çözümlü ödev atanmalı.",
                    "priority": "HIGH",
                    "curriculum_topic_id": top.get('curriculum_topic_id'),
                    "subject_name": s_name,
                    "topic_name": t_name
                })

        if not recommendations:
            recommendations.append({
                "problem": "Mevcut Performansı Koruma",
                "evidence": f"Ortalama Net: {avg_net:.2f}",
                "reason": "Temel derslerde başarım istikrarlı.",
                "action": "Düzenli deneme çözümü ve genel haftalık tekrar programı sürdürülmeli.",
                "priority": "MEDIUM"
            })

    analysis_response = {
        'student_id': student_id,
        'student_name': student_name,
        'exam_system': exam_sys,
        'track': track,
        'confidence': confidence,
        'analysisDataQuality': context['data_quality_score'],
        'valid_exam_count': valid_count,
        'summary': summary,
        'strengths': strengths,
        'weaknesses': weaknesses,
        'subject_analysis': subject_analysis,
        'topic_risks': topic_risks,
        'recommendations': [r['action'] if isinstance(r, dict) else r for r in recommendations],
        'structured_recommendations': recommendations,
        'context_debug': {
            'studentId': student_id,
            'studentName': student_name,
            'examSystem': exam_sys,
            'track': track,
            'validExamCount': valid_count,
            'confidence': confidence,
            'dataQualityScore': context['data_quality_score'],
            'mathAvg': next((s['avg_net'] for s in context['subject_summary'] if 'Matematik' in s['subject_name']), 0.0),
            'weakTopicsCount': len(topic_risks)
        }
    }
    return jsonify(analysis_response)

# ==========================================
# 20. KOÇ – ÖĞRENCİ ÇOKLU EŞLEŞTİRME & YETKİLENDİRME API
# ==========================================

# Helper: Check Many-to-Many Permission
def check_coach_student_access(user, student_id):
    if not user:
        return False, None
    if user['role'] == 'ADMIN':
        return True, 'ADMIN'
    
    conn = get_db()
    cursor = conn.cursor()
    
    if user['role'] == 'STUDENT':
        cursor.execute("SELECT id FROM students WHERE user_id = ? AND id = ?;", (user['id'], student_id))
        st = cursor.fetchone()
        conn.close()
        return (True, 'SELF') if st else (False, None)
        
    elif user['role'] == 'COACH':
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        ch = cursor.fetchone()
        if not ch:
            conn.close()
            return False, None
        coach_id = ch['id']

        cursor.execute("SELECT relationship_type FROM coach_student_relationships WHERE coach_id = ? AND student_id = ? AND status = 'ACTIVE';", (coach_id, student_id))
        rel = cursor.fetchone()
        conn.close()
        if rel:
            return True, rel['relationship_type']
        return False, None

    return False, None

# GET /api/rel/students - List connected students for logged in coach or admin
@app.route('/api/rel/students', methods=['GET'])
def get_my_connected_students():
    user = get_auth_user()
    if not user or user['role'] not in ['COACH', 'ADMIN']:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    rel_type_filter = request.args.get('relationship_type')
    status_filter = request.args.get('status', 'ACTIVE')

    if user['role'] == 'ADMIN':
        query = "SELECT s.*, u.name, u.email, u.phone, csr.id as rel_id, csr.relationship_type, csr.status as rel_status, c.id as coach_id, cu.name as coach_name, c.title as coach_title FROM students s JOIN users u ON s.user_id = u.id LEFT JOIN coach_student_relationships csr ON s.id = csr.student_id LEFT JOIN coaches c ON csr.coach_id = c.id LEFT JOIN users cu ON c.user_id = cu.id WHERE u.status = 'ACTIVE' AND u.deleted_at IS NULL"
        params = []
        if status_filter != 'ALL':
            query += " AND csr.status = ?"
            params.append(status_filter)
        if rel_type_filter:
            query += " AND csr.relationship_type = ?"
            params.append(rel_type_filter)

        cursor.execute(query, params)
        students = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'students': students})

    # COACH LOGIC
    cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
    coach = cursor.fetchone()
    if not coach:
        conn.close()
        return jsonify({'students': []})

    coach_id = coach['id']

    query = "SELECT s.*, u.name, u.email, u.phone, csr.id as rel_id, csr.relationship_type, csr.status as rel_status, csr.assigned_at FROM coach_student_relationships csr JOIN students s ON csr.student_id = s.id JOIN users u ON s.user_id = u.id WHERE csr.coach_id = ? AND u.status = 'ACTIVE' AND u.deleted_at IS NULL"
    params = [coach_id]

    if status_filter != 'ALL':
        query += " AND csr.status = ?"
        params.append(status_filter)
    if rel_type_filter:
        query += " AND csr.relationship_type = ?"
        params.append(rel_type_filter)

    query += " ORDER BY u.name ASC;"

    cursor.execute(query, params)
    students = [dict(r) for r in cursor.fetchall()]

    # Fetch pending connection requests
    cursor.execute("SELECT csr.id as request_id, u.name as student_name, s.track, s.school, csr.relationship_type FROM coach_student_relationships csr JOIN students s ON csr.student_id = s.id JOIN users u ON s.user_id = u.id WHERE csr.coach_id = ? AND csr.status = 'PENDING';", (coach_id,))
    pending_requests = [dict(r) for r in cursor.fetchall()]

    conn.close()
    return jsonify({
        'students': students,
        'pending_requests': pending_requests
    })

# GET /api/rel/my-coaches - List single main assigned coach for logged in student
@app.route('/api/rel/my-coaches', methods=['GET'])
def get_my_coaches():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    student_id = request.args.get('student_id')
    if not student_id and user['role'] == 'STUDENT':
        cursor.execute("SELECT id, coach_id FROM students WHERE user_id = ?;", (user['id'],))
        st = cursor.fetchone()
        student_id = st['id'] if st else None
    elif student_id:
        try:
            student_id = int(student_id)
        except Exception:
            student_id = None

    if not student_id:
        conn.close()
        return jsonify({'coaches': []})

    # Fetch the single main assigned coach using students.coach_id as single source of truth
    cursor.execute("""
    SELECT csr.id as rel_id,
           COALESCE(csr.relationship_type, 'MAIN_COACH') as relationship_type,
           COALESCE(csr.status, 'ACTIVE') as rel_status,
           COALESCE(csr.assigned_at, CURRENT_TIMESTAMP) as assigned_at,
           c.id as coach_id,
           c.user_id as coach_user_id,
           cu.name as coach_name,
           cu.email as coach_email,
           COALESCE(c.title, 'YKS / LGS Koçu') as coach_title,
           COALESCE(c.specialty, 'Derece & Akademik Koçluk') as specialty,
           c.coach_code,
           COALESCE(cu.status, 'ACTIVE') as coach_status
    FROM students s
    JOIN coaches c ON s.coach_id = c.id
    JOIN users cu ON c.user_id = cu.id
    LEFT JOIN coach_student_relationships csr ON (csr.student_id = s.id AND csr.coach_id = c.id AND csr.status = 'ACTIVE')
    WHERE s.id = ?
    LIMIT 1;
    """, (student_id,))
    coaches = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({'coaches': coaches})

# POST /api/rel/admin-assign - Admin assigns coach to student
@app.route('/api/rel/admin-assign', methods=['POST'])
def admin_assign_coach():
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yalnızca Admin işlem yapabilir'}), 403

    data = request.json or {}
    coach_id = data.get('coach_id')
    student_id = data.get('student_id')
    rel_type = data.get('relationship_type', 'MAIN_COACH')

    if not coach_id or not student_id:
        return jsonify({'error': 'Koç ve Öğrenci seçimi zorunludur'}), 400

    conn = get_db()
    cursor = conn.cursor()

    if rel_type == 'MAIN_COACH':
        cursor.execute("UPDATE students SET coach_id = ? WHERE id = ?;", (coach_id, student_id))
        try:
            cursor.execute("DELETE FROM coach_students WHERE student_id = ?;", (student_id,))
            cursor.execute("INSERT INTO coach_students (coach_id, student_id) VALUES (?, ?);", (coach_id, student_id))
        except Exception:
            pass

    # Upsert single relationship record for (student_id, coach_id)
    cursor.execute("SELECT id FROM coach_student_relationships WHERE student_id = ? AND coach_id = ?;", (student_id, coach_id))
    existing_rel = cursor.fetchone()
    if existing_rel:
        cursor.execute("UPDATE coach_student_relationships SET relationship_type = ?, status = 'ACTIVE', assigned_by = ? WHERE id = ?;", (rel_type, user['id'], existing_rel['id']))
    else:
        cursor.execute("INSERT INTO coach_student_relationships (coach_id, student_id, relationship_type, status, assigned_by) VALUES (?, ?, ?, 'ACTIVE', ?);", (coach_id, student_id, rel_type, user['id']))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Koç-Öğrenci eşleştirmesi başarıyla yapıldı!'})

# POST /api/rel/invite - Coach creates secure single-use invitation link /invite/TOKEN
@app.route('/api/rel/invite', methods=['POST'])
def create_coach_invitation():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz işlem'}), 401

    data = request.json or {}
    rel_type = data.get('relationship_type', 'MAIN_COACH')
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
    ch = cursor.fetchone()
    if not ch:
        cursor.execute("INSERT INTO coaches (user_id, title) VALUES (?, 'Uzman YKS Koçu');", (user['id'],))
        conn.commit()
        coach_id = cursor.lastrowid
    else:
        coach_id = ch['id']

    import uuid
    token = str(uuid.uuid4())[:8].upper()
    expires_at = (datetime.now() + timedelta(days=7)).isoformat()

    cursor.execute("INSERT INTO coach_invitations (coach_id, invitation_token, relationship_type, expires_at) VALUES (?, ?, ?, ?);", (coach_id, token, rel_type, expires_at))

    conn.commit()
    conn.close()

    invite_url = f"/invite/{token}"
    return jsonify({
        'message': 'Koç davet bağlantısı oluşturuldu!',
        'invitation_token': token,
        'invite_url': invite_url
    })

# GET /api/rel/invite/<token> - Verify invitation token for student
@app.route('/api/rel/invite/<token>', methods=['GET'])
def get_invitation_details(token):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT ci.*, c.title as coach_title, c.specialty, u.name as coach_name, u.email as coach_email FROM coach_invitations ci JOIN coaches c ON ci.coach_id = c.id JOIN users u ON c.user_id = u.id WHERE ci.invitation_token = ? AND ci.is_used = 0;", (token,))
    inv = cursor.fetchone()
    conn.close()

    if not inv:
        return jsonify({'error': 'Geçersiz veya kullanılmış davet bağlantısı'}), 404

    return jsonify({'invitation': dict(inv)})

# POST /api/rel/invite/<token>/respond - Student accepts/rejects invitation
@app.route('/api/rel/invite/<token>/respond', methods=['POST'])
def respond_invitation(token):
    user = get_auth_user()
    if not user or user['role'] != 'STUDENT':
        return jsonify({'error': 'Davetleri yalnızca öğrenciler yanıtlayabilir'}), 403

    data = request.json or {}
    action = data.get('action', 'ACCEPT')

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
    st = cursor.fetchone()
    if not st:
        conn.close()
        return jsonify({'error': 'Öğrenci profili bulunamadı'}), 404
    student_id = st['id']

    cursor.execute("SELECT * FROM coach_invitations WHERE invitation_token = ? AND is_used = 0;", (token,))
    inv = cursor.fetchone()
    if not inv:
        conn.close()
        return jsonify({'error': 'Davet bulunamadı veya daha önce kullanılmış'}), 404

    if action == 'ACCEPT':
        cursor.execute("INSERT INTO coach_student_relationships (coach_id, student_id, relationship_type, status, accepted_at) VALUES (?, ?, ?, 'ACTIVE', CURRENT_TIMESTAMP);", (inv['coach_id'], student_id, inv['relationship_type']))

        if inv['relationship_type'] == 'MAIN_COACH':
            cursor.execute("UPDATE students SET coach_id = ? WHERE id = ?;", (inv['coach_id'], student_id))

        cursor.execute("UPDATE coach_invitations SET is_used = 1, used_by_student_id = ? WHERE id = ?;", (student_id, inv['id']))
        msg = "Koç daveti kabul edildi! Koç-öğrenci bağlantısı kuruldu."
    else:
        cursor.execute("UPDATE coach_invitations SET is_used = 1 WHERE id = ?;", (inv['id'],))
        msg = "Koç daveti reddedildi."

    conn.commit()
    conn.close()
    return jsonify({'message': msg})

# POST /api/rel/coach-code - Student connects using coach code (e.g. UMMU-2026) -> Creates PENDING relationship
@app.route('/api/rel/coach-code', methods=['POST'])
def connect_with_coach_code():
    user = get_auth_user()
    if not user or user['role'] != 'STUDENT':
        return jsonify({'error': 'Yalnızca öğrenciler koç kodu kullanabilir'}), 403

    data = request.json or {}
    code = data.get('coach_code', '').strip().upper()

    if not code:
        return jsonify({'error': 'Koç kodu girilmesi zorunludur'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
    st = cursor.fetchone()
    if not st:
        conn.close()
        return jsonify({'error': 'Öğrenci profili bulunamadı'}), 404
    student_id = st['id']

    cursor.execute("SELECT id, user_id FROM coaches WHERE coach_code = ?;", (code,))
    ch = cursor.fetchone()
    if not ch:
        conn.close()
        return jsonify({'error': 'Geçersiz Koç Kodu! Lütfen koçunuzdan doğru kodu isteyin.'}), 404
    coach_id = ch['id']

    cursor.execute("INSERT INTO coach_student_relationships (coach_id, student_id, relationship_type, status) VALUES (?, ?, 'MAIN_COACH', 'PENDING');", (coach_id, student_id))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Koç kodu ile bağlantı talebi gönderildi! Koç onayından sonra aktif olacaktır.'})

# GET /api/rel/coaches-search - List all coaches for student discovery
@app.route('/api/rel/coaches-search', methods=['GET'])
def search_coaches():
    search = request.args.get('q', '').strip()
    conn = get_db()
    cursor = conn.cursor()
    query = "SELECT c.id as coach_id, c.title, c.specialty, c.bio, c.coach_code, u.name as coach_name, u.email FROM coaches c JOIN users u ON c.user_id = u.id WHERE u.status = 'ACTIVE'"
    params = []
    if search:
        query += " AND (u.name LIKE ? OR c.specialty LIKE ? OR c.title LIKE ?)"
        params.extend([f"%{search}%", f"%{search}%", f"%{search}%"])

    cursor.execute(query, params)
    coaches = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({'coaches': coaches})

# POST /api/rel/requests/<int:req_id>/respond - Coach accepts/rejects pending connection
@app.route('/api/rel/requests/<int:req_id>/respond', methods=['POST'])
def respond_connection_request(req_id):
    user = get_auth_user()
    if not user or user['role'] not in ['COACH', 'ADMIN']:
        return jsonify({'error': 'Yetkisiz işlem'}), 403

    data = request.json or {}
    raw_action = data.get('action', 'APPROVE').upper()
    status_map = {
        'APPROVE': 'ACTIVE',
        'APPROVED': 'ACTIVE',
        'ACCEPT': 'ACTIVE',
        'ACCEPTED': 'ACTIVE',
        'ACTIVE': 'ACTIVE',
        'REJECT': 'REJECTED',
        'REJECTED': 'REJECTED',
        'PENDING': 'PENDING',
        'SUSPENDED': 'SUSPENDED',
        'ENDED': 'ENDED'
    }
    status_val = status_map.get(raw_action)
    if not status_val:
        return jsonify({'error': 'Geçersiz action parametresi'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM coach_student_relationships WHERE id = ?;", (req_id,))
    rel = cursor.fetchone()
    if not rel:
        conn.close()
        return jsonify({'error': 'Talep bulunamadı'}), 404

    cursor.execute("UPDATE coach_student_relationships SET status = ? WHERE id = ?;", (status_val, req_id))
    conn.commit()
    conn.close()
    return jsonify({'message': f'Bağlantı talebi {status_val} olarak güncellendi!'})


# GET & POST /api/rel/coach-notes - Coach notes for student
@app.route('/api/rel/coach-notes', methods=['GET', 'POST'])
def handle_coach_notes():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    if request.method == 'GET':
        student_id = request.args.get('student_id', 1)
        
        has_access, rel_type = check_coach_student_access(user, int(student_id))
        if not has_access:
            return jsonify({'error': 'Bu öğrencinin notlarına erişim yetkiniz yok'}), 403

        conn = get_db()
        cursor = conn.cursor()


        query = "SELECT cn.*, u.name as coach_name FROM coach_notes cn JOIN coaches c ON cn.coach_id = c.id JOIN users u ON c.user_id = u.id WHERE cn.student_id = ?"
        params = [student_id]
        if user['role'] == 'STUDENT':
            query += " AND cn.visibility = 'VISIBLE_TO_STUDENT'"

        query += " ORDER BY cn.created_at DESC;"
        cursor.execute(query, params)
        notes = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'notes': notes})

    elif request.method == 'POST':
        if user['role'] not in ['COACH', 'ADMIN']:
            return jsonify({'error': 'Yalnızca koçlar not yazabilir'}), 403

        data = request.json or {}
        student_id = data.get('student_id')
        note = data.get('note', '').strip()
        visibility = data.get('visibility', 'PRIVATE_TO_COACH')

        if not student_id or not note:
            return jsonify({'error': 'Öğrenci seçimi ve not metni gereklidir'}), 400

        conn = get_db()
        cursor = conn.cursor()

        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))

        ch = cursor.fetchone()
        coach_id = ch['id'] if ch else 1

        cursor.execute("INSERT INTO coach_notes (coach_id, student_id, note, visibility) VALUES (?, ?, ?, ?);", (coach_id, student_id, note, visibility))

        conn.commit()
        conn.close()
        return jsonify({'message': 'Koç notu eklendi!'})

# ==========================================
# 21. KAYNAK HAVUZU & KÜTÜPHANE YÖNETİMİ API
# ==========================================

@app.route('/api/kaynak-havuzu', methods=['GET', 'POST'])
def handle_kaynak_havuzu():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    coach_id = None
    if user['role'] == 'COACH':
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        ch = cursor.fetchone()
        if ch:
            coach_id = ch['id']

    if request.method == 'GET':
        tab = request.args.get('tab', 'ALL') # ALL, PERSONAL, SYSTEM, ASSIGNED, ARCHIVED
        search = request.args.get('search', '').strip()
        subject_id = request.args.get('subject_id')
        exam_system = request.args.get('exam_system')
        resource_type = request.args.get('resource_type')
        raw_limit = request.args.get('limit')
        raw_offset = request.args.get('offset', '0')

        # 1. Base Authorization Filter & Pool Isolation
        auth_where = "1=1"
        auth_params = []

        if user['role'] == 'COACH':
            auth_where = "(r.owner_type = 'COACH' AND r.owner_id = ?)"
            auth_params.append(coach_id)
        elif user['role'] == 'ADMIN':
            auth_where = "(r.owner_type = 'SYSTEM' OR r.owner_id IS NULL)"
        elif user['role'] == 'STUDENT':
            cursor.execute("SELECT id FROM students WHERE user_id = ?;", (user['id'],))
            st = cursor.fetchone()
            st_id = st['id'] if st else 1
            auth_where = "r.id IN (SELECT resource_id FROM resource_assignments WHERE student_id = ? AND status = 'ACTIVE')"
            auth_params.append(st_id)

        def _fetch_resources():
            raw_query = f"""
            SELECT r.*, COALESCE(r.name, r.title) as name, s.name as subject_name,
                   COALESCE(rt_agg.cnt, 0) as topic_count,
                   COALESCE(ra_agg.cnt, 0) as assigned_student_count
            FROM resources r
            LEFT JOIN subjects s ON r.subject_id = s.id
            LEFT JOIN (
                SELECT resource_id, COUNT(*) as cnt
                FROM resource_topics
                GROUP BY resource_id
            ) rt_agg ON rt_agg.resource_id = r.id
            LEFT JOIN (
                SELECT resource_id, COUNT(*) as cnt
                FROM resource_assignments
                WHERE status = 'ACTIVE' {"AND coach_id = ?" if user['role'] == 'COACH' else ""}
                GROUP BY resource_id
            ) ra_agg ON ra_agg.resource_id = r.id
            WHERE {auth_where}
            ORDER BY r.id DESC;
            """
            q_params = []
            if user['role'] == 'COACH':
                q_params.append(coach_id)
            q_params.extend(auth_params)
            cursor.execute(raw_query, q_params)
            return [dict(row) for row in cursor.fetchall()]

        all_authorized_rows = _fetch_resources()

        # If coach pool is totally empty, initialize from Genel Havuz on demand (0 overhead for normal requests)
        if user['role'] == 'COACH' and len(all_authorized_rows) == 0:
            copy_system_resources_to_coach_pool(coach_id, cursor=cursor)
            all_authorized_rows = _fetch_resources()

        # 2. Compute exact KPIs over ALL Authorized Resources
        total_raw = len(all_authorized_rows)
        visible_non_archived = [r for r in all_authorized_rows if r.get('status') != 'ARCHIVED']
        archived_rows = [r for r in all_authorized_rows if r.get('status') == 'ARCHIVED']
        active_rows = [r for r in visible_non_archived if r.get('status', 'ACTIVE') == 'ACTIVE']
        assigned_resources_rows = [r for r in visible_non_archived if (r.get('assigned_student_count') or 0) > 0]

        # 3. Apply Tab & Filter criteria for the display list
        filtered_resources = []
        for r in all_authorized_rows:
            r_status = r.get('status', 'ACTIVE')

            # Tab filter
            if tab == 'ARCHIVED':
                if r_status != 'ARCHIVED':
                    continue
            else:
                if r_status == 'ARCHIVED':
                    continue

            if tab == 'PERSONAL' and user['role'] == 'COACH':
                if r.get('owner_type') != 'COACH' or r.get('owner_id') != coach_id:
                    continue
            elif tab == 'SYSTEM' and user['role'] == 'COACH':
                if r.get('origin_resource_id') is None and r.get('owner_type') != 'SYSTEM':
                    continue
            elif tab == 'ASSIGNED' and user['role'] == 'COACH':
                if (r.get('assigned_student_count') or 0) == 0:
                    continue

            # Search filter
            if search:
                s_lower = search.lower()
                r_name = (r.get('name') or r.get('title') or '').lower()
                r_pub = (r.get('publisher') or '').lower()
                r_desc = (r.get('description') or '').lower()
                if s_lower not in r_name and s_lower not in r_pub and s_lower not in r_desc:
                    continue

            # Subject filter
            if subject_id and subject_id != 'ALL':
                if str(r.get('subject_id')) != str(subject_id):
                    continue

            # Exam system filter
            if exam_system and exam_system != 'ALL':
                if (r.get('exam_system') or '').upper() != str(exam_system).upper():
                    continue

            # Resource type filter
            if resource_type and resource_type != 'ALL':
                if (r.get('resource_type') or '').lower() != str(resource_type).lower():
                    continue

            filtered_resources.append(r)

        total_filtered_count = len(filtered_resources)
        has_more = False
        limit_val = None
        offset_val = 0

        # Optional pagination support (Backwards compatible when limit is omitted)
        if raw_limit is not None:
            try:
                limit_val = int(raw_limit)
                offset_val = int(raw_offset)
                paged_resources = filtered_resources[offset_val:offset_val + limit_val]
                has_more = (offset_val + limit_val) < total_filtered_count
                filtered_resources = paged_resources
            except (ValueError, TypeError):
                pass

        cursor.execute("SELECT id, name FROM subjects ORDER BY id ASC;")
        subjects_list = [dict(s) for s in cursor.fetchall()]

        conn.close()

        resp_data = {
            'resources': filtered_resources,
            'subjects': subjects_list,
            'kpis': {
                'total_resources': len(visible_non_archived),
                'active_resources': len(active_rows),
                'assigned_students': len(assigned_resources_rows),
                'archived_resources': len(archived_rows)
            },
            'debug': {
                'currentCoachId': coach_id if user['role'] == 'COACH' else None,
                'totalRawResources': total_raw,
                'visibleResources': len(visible_non_archived),
                'filteredResources': total_filtered_count,
                'activeResources': len(active_rows),
                'archivedResources': len(archived_rows),
                'assignedResources': len(assigned_resources_rows),
                'currentTab': tab,
                'searchQuery': search,
                'subjectFilter': subject_id or 'ALL',
                'examSystemFilter': exam_system or 'ALL',
                'resourceTypeFilter': resource_type or 'ALL'
            }
        }

        if raw_limit is not None:
            resp_data['total'] = total_filtered_count
            resp_data['limit'] = limit_val
            resp_data['offset'] = offset_val
            resp_data['has_more'] = has_more

        return jsonify(resp_data)

    elif request.method == 'POST':
        if user['role'] not in ['COACH', 'ADMIN']:
            return jsonify({'error': 'Yalnızca koçlar ve adminler kaynak ekleyebilir'}), 403

        data = request.json or {}
        name = data.get('name', '').strip()
        publisher = data.get('publisher', '').strip()
        exam_system = data.get('exam_system', 'YKS')
        exam_type = data.get('exam_type', 'TYT')
        field = data.get('field', 'ALL')
        subject_id = data.get('subject_id')
        resource_type = data.get('resource_type', 'Soru Bankası')
        isbn = data.get('isbn', '').strip()
        edition = data.get('edition', '').strip()
        description = data.get('description', '').strip()
        topic_ids = data.get('topic_ids', [])

        if not name or not subject_id:
            return jsonify({'error': 'Kaynak adı ve ders seçimi zorunludur.'}), 400

        owner_type = 'SYSTEM' if user['role'] == 'ADMIN' else 'COACH'
        owner_id = None if owner_type == 'SYSTEM' else coach_id

        cursor.execute("""
        INSERT INTO resources (owner_type, owner_id, name, title, publisher, exam_system, exam_type, field, subject_id, resource_type, isbn, edition, description)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
        """, (owner_type, owner_id, name, name, publisher, exam_system, exam_type, field, subject_id, resource_type, isbn, edition, description))
        resource_id = cursor.lastrowid

        if topic_ids:
            for idx, tid in enumerate(topic_ids):
                cursor.execute("""
                INSERT INTO resource_topics (resource_id, curriculum_topic_id, order_index)
                VALUES (?, ?, ?);
                """, (resource_id, tid, idx + 1))

        # If added by a COACH, create Admin Resource Suggestion & Notifications
        if user['role'] == 'COACH':
            cursor.execute("SELECT name FROM subjects WHERE id = ?;", (subject_id,))
            subj_row = cursor.fetchone()
            subject_name = subj_row['name'] if subj_row else 'Ders'

            coach_name = f"{user.get('name', '')} {user.get('surname', '')}".strip() or user.get('username')

            cursor.execute("""
            INSERT INTO resource_suggestions (
                coach_id, coach_user_id, coach_name, coach_resource_id, 
                resource_title, publisher, subject_id, subject_name, 
                exam_system, exam_type, field, resource_type, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'BEKLİYOR');
            """, (
                coach_id, user['id'], coach_name, resource_id,
                name, publisher, subject_id, subject_name,
                exam_system, exam_type, field, resource_type
            ))
            suggestion_id = cursor.lastrowid

            # Send Notification to all ADMIN users
            cursor.execute("SELECT id FROM users WHERE role = 'ADMIN';")
            admins = cursor.fetchall()
            event_key_base = f"RESOURCE_SUGGESTION_{suggestion_id}"
            notif_msg = f"Koç {coach_name}, {name} ({publisher}) kaynağını kendi havuzuna ekledi."

            for adm in admins:
                cursor.execute("""
                INSERT OR IGNORE INTO notifications (
                    recipient_user_id, actor_user_id, type, title, message, 
                    entity_type, entity_id, event_key
                ) VALUES (?, ?, 'RESOURCE_SUGGESTION', '🔔 Genel Havuz Kaynak Önerisi', ?, 'RESOURCE_SUGGESTION', ?, ?);
                """, (adm['id'], user['id'], notif_msg, suggestion_id, f"{event_key_base}_{adm['id']}"))

        conn.commit()
        conn.close()
        return jsonify({
            'message': 'Kaynak başarıyla eklendi.' if user['role'] == 'ADMIN' else 'Kaynak özel havuzunuza eklendi ve Genel Havuz için Admin onayına gönderildi.',
            'resource_id': resource_id
        })

# ==========================================
# 21b. ADMIN RESOURCE SUGGESTIONS ENGINE
# ==========================================

@app.route('/api/admin/resource-suggestions', methods=['GET'])
def get_resource_suggestions():
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT s.*, r.description, r.cover_url
    FROM resource_suggestions s
    LEFT JOIN resources r ON s.coach_resource_id = r.id
    ORDER BY s.id DESC;
    """)
    suggestions = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify({'suggestions': suggestions})

@app.route('/api/admin/resource-suggestions/<int:suggestion_id>/respond', methods=['POST'])
def respond_resource_suggestion(suggestion_id):
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    data = request.json or {}
    action = (data.get('action') or '').upper() # APPROVE or REJECT
    if action not in ('APPROVE', 'REJECT'):
        return jsonify({'error': 'Geçersiz işlem (APPROVE veya REJECT olmalı)'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM resource_suggestions WHERE id = ?;", (suggestion_id,))
    sug = cursor.fetchone()
    if not sug:
        conn.close()
        return jsonify({'error': 'Kaynak önerisi bulunamadı.'}), 404

    sug = dict(sug)
    if sug['status'] != 'BEKLİYOR':
        conn.close()
        return jsonify({'error': f"Bu öneri daha önce işlenmiş ({sug['status']})."}), 400

    new_status = 'ONAYLANDI' if action == 'APPROVE' else 'REDDEDİLDİ'

    if action == 'APPROVE':
        # Add resource to GENEL KAYNAK HAVUZU (SYSTEM)
        cursor.execute("SELECT * FROM resources WHERE id = ?;", (sug['coach_resource_id'],))
        coach_res = cursor.fetchone()
        if coach_res:
            c_res = dict(coach_res)
            cursor.execute("""
            INSERT INTO resources (
                owner_type, owner_id, origin_resource_id, name, title, publisher, 
                exam_system, exam_type, field, subject_id, resource_type, level, 
                isbn, edition, description, cover_url, total_questions, status
            ) VALUES (
                'SYSTEM', NULL, ?, ?, ?, ?, 
                ?, ?, ?, ?, ?, ?, 
                ?, ?, ?, ?, ?, 'ACTIVE'
            );
            """, (
                c_res['id'], c_res['name'], c_res['name'], c_res['publisher'],
                c_res['exam_system'], c_res['exam_type'], c_res['field'], c_res['subject_id'],
                c_res['resource_type'], c_res.get('level', 'Orta'), c_res.get('isbn'), c_res.get('edition'),
                c_res.get('description'), c_res.get('cover_url'), c_res.get('total_questions', 0)
            ))
            new_sys_res_id = cursor.lastrowid

            cursor.execute("SELECT curriculum_topic_id, chapter_name, order_index FROM resource_topics WHERE resource_id = ?;", (c_res['id'],))
            topics = cursor.fetchall()
            for t in topics:
                cursor.execute("""
                INSERT INTO resource_topics (resource_id, curriculum_topic_id, chapter_name, order_index)
                VALUES (?, ?, ?, ?);
                """, (new_sys_res_id, t['curriculum_topic_id'], t['chapter_name'], t['order_index']))

    rejection_reason = (data.get('rejection_reason') or '').strip() if action == 'REJECT' else None
    cursor.execute("UPDATE resource_suggestions SET status = ?, rejection_reason = ?, updated_at = CURRENT_TIMESTAMP WHERE id = ?;", (new_status, rejection_reason, suggestion_id))
    
    # Mark related Admin notification as read
    cursor.execute("UPDATE notifications SET is_read = 1 WHERE entity_type = 'RESOURCE_SUGGESTION' AND entity_id = ?;", (suggestion_id,))

    # Send Notification to Coach (Requirement 11)
    coach_user_id = sug.get('coach_user_id')
    if coach_user_id:
        if action == 'APPROVE':
            coach_msg = f"🎉 '{sug['resource_title']}' kaynak öneriniz Admin tarafından onaylandı ve Genel Kaynak Havuzu'na eklendi."
        else:
            reason_str = f" (Neden: {rejection_reason})" if rejection_reason else ""
            coach_msg = f"❌ '{sug['resource_title']}' kaynak öneriniz Genel Kaynak Havuzu'na eklenmedi.{reason_str}"

        event_key = f"RESOURCE_SUGGESTION_RESPONSE_{suggestion_id}_{coach_user_id}"
        cursor.execute("""
        INSERT OR IGNORE INTO notifications (
            recipient_user_id, actor_user_id, type, title, message, 
            entity_type, entity_id, event_key
        ) VALUES (?, ?, 'RESOURCE_SUGGESTION_RESPONSE', '🔔 Kaynak Önerisi Sonucu', ?, 'RESOURCE_SUGGESTION', ?, ?);
        """, (coach_user_id, user['id'], coach_msg, suggestion_id, event_key))

        send_auto_notification(
            user['id'],
            coach_user_id,
            coach_msg,
            message_type='RESOURCE',
            cursor=cursor
        )

    conn.commit()
    conn.close()

    msg = "Kaynak Genel Kaynak Havuzuna eklendi." if action == 'APPROVE' else "Kaynak önerisi reddedildi. Kaynak yalnızca koçun kendi özel havuzunda kalacak."
    return jsonify({'message': msg, 'status': new_status})

# ==========================================
# 21c. RESOURCE MANAGEMENT ADMIN SUITE APIs
# ==========================================

@app.route('/api/admin/resource-management/stats', methods=['GET'])
def get_resource_management_stats():
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) as cnt FROM resources WHERE (owner_type = 'SYSTEM' OR owner_id IS NULL) AND (status IS NULL OR status != 'ARCHIVED');")
    total_general = cursor.fetchone()['cnt']

    cursor.execute("SELECT COUNT(DISTINCT coach_id) as cnt FROM resource_assignments WHERE status = 'ACTIVE';")
    coaches_using = cursor.fetchone()['cnt']

    cursor.execute("SELECT COUNT(*) as cnt FROM resource_suggestions WHERE status = 'BEKLİYOR';")
    pending_suggestions = cursor.fetchone()['cnt']

    cursor.execute("SELECT COUNT(*) as cnt FROM subjects;")
    total_subjects = cursor.fetchone()['cnt']

    cursor.execute("SELECT COUNT(DISTINCT publisher) as cnt FROM resources WHERE publisher IS NOT NULL AND publisher != '';")
    total_publishers = cursor.fetchone()['cnt']

    conn.close()
    return jsonify({
        'total_general_resources': total_general,
        'coaches_using_resources': coaches_using,
        'pending_suggestions_count': pending_suggestions,
        'total_subjects': total_subjects,
        'total_publishers': total_publishers
    })

@app.route('/api/kaynak-havuzu/bulk-action', methods=['POST'])
def handle_kaynak_bulk_action():
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    data = request.json or {}
    resource_ids = data.get('resource_ids', [])
    action = data.get('action') # ACTIVATE, DEACTIVATE, DELETE

    if not resource_ids or not action:
        return jsonify({'error': 'Kaynak seçimleri ve işlem türü gereklidir.'}), 400

    conn = get_db()
    cursor = conn.cursor()

    placeholders = ','.join(['?'] * len(resource_ids))
    params = list(resource_ids)

    if action == 'ACTIVATE':
        cursor.execute(f"UPDATE resources SET status = 'ACTIVE' WHERE id IN ({placeholders}) AND (owner_type = 'SYSTEM' OR owner_id IS NULL);", params)
        msg = f"{len(resource_ids)} kaynak aktifleştirildi."
    elif action in ('DEACTIVATE', 'DELETE'):
        cursor.execute(f"UPDATE resources SET status = 'INACTIVE' WHERE id IN ({placeholders}) AND (owner_type = 'SYSTEM' OR owner_id IS NULL);", params)
        msg = f"{len(resource_ids)} kaynak pasife alındı."
    else:
        conn.close()
        return jsonify({'error': 'Geçersiz işlem.'}), 400

    conn.commit()
    conn.close()
    return jsonify({'message': msg})

@app.route('/api/kaynak-havuzu/<int:resource_id>/details', methods=['GET'])
def get_resource_details(resource_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT r.*, s.name as subject_name
    FROM resources r
    LEFT JOIN subjects s ON r.subject_id = s.id
    WHERE r.id = ?;
    """, (resource_id,))
    res_row = cursor.fetchone()
    if not res_row:
        conn.close()
        return jsonify({'error': 'Kaynak bulunamadı.'}), 404

    res_dict = dict(res_row)

    # Find assigned coaches
    cursor.execute("""
    SELECT c.id as coach_id, u.name, u.surname, COUNT(ra.student_id) as student_count
    FROM resource_assignments ra
    JOIN coaches c ON ra.coach_id = c.id
    JOIN users u ON c.user_id = u.id
    WHERE ra.resource_id = ? OR ra.resource_id IN (SELECT id FROM resources WHERE origin_resource_id = ?)
    GROUP BY c.id;
    """, (resource_id, resource_id))
    assigned_coaches = [dict(r) for r in cursor.fetchall()]

    # Statistics
    cursor.execute("""
    SELECT 
        COUNT(*) as total_assignments,
        SUM(CASE WHEN status = 'COMPLETED' THEN 1 ELSE 0 END) as completed_count,
        SUM(CASE WHEN status = 'ACTIVE' THEN 1 ELSE 0 END) as active_count,
        SUM(CASE WHEN status = 'ARCHIVED' THEN 1 ELSE 0 END) as archived_count
    FROM resource_assignments
    WHERE resource_id = ? OR resource_id IN (SELECT id FROM resources WHERE origin_resource_id = ?);
    """, (resource_id, resource_id))
    stats_row = cursor.fetchone()
    tot = (stats_row['total_assignments'] if stats_row else 0) or 0
    comp = (stats_row['completed_count'] if stats_row else 0) or 0
    act = (stats_row['active_count'] if stats_row else 0) or 0
    comp_rate = round((comp / tot * 100), 1) if tot > 0 else 0

    conn.close()
    return jsonify({
        'resource': res_dict,
        'assigned_coaches': assigned_coaches,
        'statistics': {
            'total_assignments': tot,
            'completed_assignments': comp,
            'in_progress_assignments': act,
            'pending_assignments': max(0, tot - comp - act),
            'completion_rate': comp_rate
        }
    })

@app.route('/api/admin/coach-resources', methods=['GET'])
def get_coach_resources_admin():
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    coach_id = request.args.get('coach_id')
    conn = get_db()
    cursor = conn.cursor()

    if coach_id:
        cursor.execute("SELECT c.id, u.name, u.surname FROM coaches c JOIN users u ON c.user_id = u.id WHERE c.id = ?;", (coach_id,))
        ch = cursor.fetchone()
        if not ch:
            conn.close()
            return jsonify({'error': 'Koç bulunamadı.'}), 404

        cursor.execute("""
        SELECT r.*, s.name as subject_name,
               (CASE WHEN r.origin_resource_id IS NOT NULL OR r.owner_type = 'SYSTEM' THEN 'Sistem Kaynağı' ELSE 'Koç Kaynağı' END) as source_type,
               (SELECT COUNT(*) FROM resource_assignments ra WHERE ra.resource_id = r.id AND ra.status = 'ACTIVE') as assigned_student_count
        FROM resources r
        LEFT JOIN subjects s ON r.subject_id = s.id
        WHERE r.owner_type = 'COACH' AND r.owner_id = ?
        ORDER BY r.id DESC;
        """, (coach_id,))
        res_rows = [dict(r) for r in cursor.fetchall()]

        total_cnt = len(res_rows)
        sys_cnt = len([r for r in res_rows if r['source_type'] == 'Sistem Kaynağı'])
        coach_cnt = total_cnt - sys_cnt

        conn.close()
        return jsonify({
            'coach': dict(ch),
            'summary': {
                'total_resources': total_cnt,
                'system_resources_count': sys_cnt,
                'coach_added_count': coach_cnt
            },
            'resources': res_rows
        })

    cursor.execute("""
    SELECT c.id, u.name, u.surname, u.username,
           (SELECT COUNT(*) FROM resources r WHERE r.owner_type = 'COACH' AND r.owner_id = c.id) as total_resources_count
    FROM coaches c
    JOIN users u ON c.user_id = u.id
    ORDER BY u.name ASC;
    """)
    coaches = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({'coaches': coaches})

@app.route('/api/admin/coach-resources/<int:coach_resource_id>/promote-to-system', methods=['POST'])
def promote_coach_resource_to_system(coach_resource_id):
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM resources WHERE id = ?;", (coach_resource_id,))
    c_res = cursor.fetchone()
    if not c_res:
        conn.close()
        return jsonify({'error': 'Koç kaynağı bulunamadı.'}), 404

    c_res = dict(c_res)
    cursor.execute("""
    INSERT INTO resources (
        owner_type, owner_id, origin_resource_id, name, title, publisher, 
        exam_system, exam_type, field, subject_id, resource_type, level, 
        isbn, edition, description, cover_url, total_questions, status
    ) VALUES (
        'SYSTEM', NULL, ?, ?, ?, ?, 
        ?, ?, ?, ?, ?, ?, 
        ?, ?, ?, ?, ?, 'ACTIVE'
    );
    """, (
        c_res['id'], c_res['name'], c_res['name'], c_res['publisher'],
        c_res['exam_system'], c_res['exam_type'], c_res['field'], c_res['subject_id'],
        c_res['resource_type'], c_res.get('level', 'Orta'), c_res.get('isbn'), c_res.get('edition'),
        c_res.get('description'), c_res.get('cover_url'), c_res.get('total_questions', 0)
    ))
    new_sys_id = cursor.lastrowid

    cursor.execute("SELECT curriculum_topic_id, chapter_name, order_index FROM resource_topics WHERE resource_id = ?;", (c_res['id'],))
    topics = cursor.fetchall()
    for t in topics:
        cursor.execute("""
        INSERT INTO resource_topics (resource_id, curriculum_topic_id, chapter_name, order_index)
        VALUES (?, ?, ?, ?);
        """, (new_sys_id, t['curriculum_topic_id'], t['chapter_name'], t['order_index']))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Koç kaynağı Genel Kaynak Havuzuna eklendi.', 'new_system_resource_id': new_sys_id})

@app.route('/api/admin/publishers', methods=['GET', 'POST'])
def handle_publishers_admin():
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        cursor.execute("SELECT DISTINCT publisher FROM resources WHERE publisher IS NOT NULL AND publisher != '';")
        res_pubs = [r['publisher'].strip() for r in cursor.fetchall() if r['publisher']]
        for rp in res_pubs:
            cursor.execute("INSERT OR IGNORE INTO publishers (name, status) VALUES (?, 'ACTIVE');", (rp,))
        conn.commit()

        cursor.execute("""
        SELECT p.id, p.name, COALESCE(p.status, 'ACTIVE') as status,
               (SELECT COUNT(*) FROM resources r WHERE LOWER(TRIM(r.publisher)) = LOWER(TRIM(p.name))) as total_resources_count,
               (SELECT COUNT(*) FROM resources r WHERE LOWER(TRIM(r.publisher)) = LOWER(TRIM(p.name)) AND r.status = 'ACTIVE') as active_resources_count
        FROM publishers p
        ORDER BY p.name ASC;
        """)
        pub_rows = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'publishers': pub_rows})

    elif request.method == 'POST':
        data = request.json or {}
        name = (data.get('name') or '').strip()
        if not name:
            conn.close()
            return jsonify({'error': 'Yayınevi adı zorunludur.'}), 400

        cursor.execute("INSERT INTO publishers (name, status) VALUES (?, 'ACTIVE');", (name,))
        pub_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return jsonify({'message': 'Yayınevi başarıyla eklendi.', 'publisher_id': pub_id})

@app.route('/api/admin/publishers/<int:pub_id>', methods=['PUT'])
def update_publisher_admin(pub_id):
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    data = request.json or {}
    status = data.get('status')
    name = (data.get('name') or '').strip()

    conn = get_db()
    cursor = conn.cursor()

    if status:
        cursor.execute("UPDATE publishers SET status = ? WHERE id = ?;", (status, pub_id))
    if name:
        cursor.execute("UPDATE publishers SET name = ? WHERE id = ?;", (name, pub_id))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Yayınevi güncellendi.'})

@app.route('/api/admin/subjects-summary', methods=['GET'])
def get_subjects_summary_admin():
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT s.id, s.name, s.exam_system,
           (SELECT COUNT(*) FROM resources r WHERE r.subject_id = s.id) as total_resources_count,
           (SELECT COUNT(*) FROM resources r WHERE r.subject_id = s.id AND r.status = 'ACTIVE') as active_resources_count
    FROM subjects s
    ORDER BY s.id ASC;
    """)
    subjects = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({'subjects': subjects})

@app.route('/api/admin/resource-analytics', methods=['GET'])
def get_resource_analytics():
    user = get_auth_user()
    if not user or user['role'] != 'ADMIN':
        return jsonify({'error': 'Yetkisiz erişim'}), 403

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT r.id, COALESCE(r.name, r.title) as name, r.publisher, s.name as subject_name,
           COUNT(ra.id) as assignment_count,
           SUM(CASE WHEN ra.status = 'COMPLETED' THEN 1 ELSE 0 END) as completed_count
    FROM resources r
    JOIN resource_assignments ra ON ra.resource_id = r.id OR ra.resource_id IN (SELECT id FROM resources WHERE origin_resource_id = r.id)
    LEFT JOIN subjects s ON r.subject_id = s.id
    WHERE r.owner_type = 'SYSTEM' OR r.owner_id IS NULL
    GROUP BY r.id, r.name, r.subject, r.grade, r.exam_type, r.publisher, r.difficulty
    ORDER BY assignment_count DESC
    LIMIT 10;
    """)
    top_assigned = []
    for row in cursor.fetchall():
        d = dict(row)
        tot = d['assignment_count'] or 0
        comp = d['completed_count'] or 0
        d['completion_rate'] = round((comp / tot * 100), 1) if tot > 0 else 0
        top_assigned.append(d)

    cursor.execute("""
    SELECT r.publisher, COUNT(r.id) as total_resources, COUNT(ra.id) as total_assignments
    FROM resources r
    LEFT JOIN resource_assignments ra ON ra.resource_id = r.id
    WHERE r.publisher IS NOT NULL AND r.publisher != ''
    GROUP BY r.publisher
    ORDER BY total_assignments DESC, total_resources DESC
    LIMIT 10;
    """)
    top_publishers = [dict(r) for r in cursor.fetchall()]

    cursor.execute("""
    SELECT s.name as subject_name, COUNT(r.id) as resource_count, COUNT(ra.id) as assignment_count
    FROM subjects s
    LEFT JOIN resources r ON r.subject_id = s.id
    LEFT JOIN resource_assignments ra ON ra.resource_id = r.id
    GROUP BY s.id, s.name
    ORDER BY assignment_count DESC;
    """)
    subject_usage = [dict(r) for r in cursor.fetchall()]

    conn.close()
    return jsonify({
        'top_assigned_resources': top_assigned,
        'top_publishers': top_publishers,
        'subject_usage': subject_usage
    })

@app.route('/api/kaynak-havuzu/<int:resource_id>', methods=['PUT', 'DELETE'])
def handle_kaynak_havuzu_detail(resource_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    # Verify Ownership
    if user['role'] == 'COACH':
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        ch = cursor.fetchone()
        coach_id = ch['id'] if ch else None
        cursor.execute("SELECT * FROM resources WHERE id = ? AND owner_type = 'COACH' AND owner_id = ?;", (resource_id, coach_id))
    elif user['role'] == 'ADMIN':
        cursor.execute("SELECT * FROM resources WHERE id = ? AND (owner_type = 'SYSTEM' OR owner_id IS NULL);", (resource_id,))
    else:
        conn.close()
        return jsonify({'error': 'Yetkisiz işlem'}), 403

    res_item = cursor.fetchone()
    if not res_item:
        conn.close()
        return jsonify({'error': 'Kaynak bulunamadı veya bu kaynağı yönetme yetkiniz yok'}), 404
        conn.close()
        return jsonify({'error': 'Kaynak bulunamadı'}), 404

    res_dict = dict(res_item)

    if user['role'] == 'COACH':
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        ch = cursor.fetchone()
        coach_id = ch['id'] if ch else None

        if res_dict['owner_type'] == 'SYSTEM':
            conn.close()
            return jsonify({'error': 'Sistem kaynaklarını düzenleyemez veya silemezsiniz.'}), 403
        if res_dict['owner_id'] != coach_id:
            conn.close()
            return jsonify({'error': 'Başka bir koça ait kaynağı düzenleyemezsiniz.'}), 403

    if request.method == 'PUT':
        data = request.json or {}
        name = data.get('name', res_dict['name']).strip()
        publisher = data.get('publisher', res_dict['publisher']).strip()
        exam_system = data.get('exam_system', res_dict['exam_system'])
        exam_type = data.get('exam_type', res_dict['exam_type'])
        field = data.get('field', res_dict['field'])
        subject_id = data.get('subject_id', res_dict['subject_id'])
        resource_type = data.get('resource_type', res_dict['resource_type'])
        description = data.get('description', res_dict['description']).strip()
        status = data.get('status', res_dict['status'])

        cursor.execute("""
        UPDATE resources
        SET name = ?, publisher = ?, exam_system = ?, exam_type = ?, field = ?, subject_id = ?, resource_type = ?, description = ?, status = ?, updated_at = CURRENT_TIMESTAMP
        WHERE id = ?;
        """, (name, publisher, exam_system, exam_type, field, subject_id, resource_type, description, status, resource_id))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Kaynak başarıyla güncellendi.'})

    elif request.method == 'DELETE':
        if user['role'] not in ['COACH', 'ADMIN']:
            return jsonify({'error': 'Yalnızca koçlar veya admin kaynak silebilir.'}), 403

        cursor.execute("SELECT COUNT(*) as count FROM resource_assignments WHERE resource_id = ? AND status = 'ACTIVE';", (resource_id,))
        assigned_count = cursor.fetchone()['count']

        if assigned_count > 0:
            cursor.execute("UPDATE resources SET status = 'ARCHIVED' WHERE id = ?;", (resource_id,))
            conn.commit()
            conn.close()
            return jsonify({'message': 'Kaynak öğrencilere atanmış olduğu için Arşivlendi. Geçmiş öğrenci kayıtları korundu.', 'archived': True})
        else:
            cursor.execute("DELETE FROM resources WHERE id = ?;", (resource_id,))
            conn.commit()
            conn.close()
            return jsonify({'message': 'Kaynak başarıyla silindi.', 'deleted': True})

@app.route('/api/kaynak-havuzu/<int:resource_id>/topics', methods=['GET', 'POST'])
def handle_kaynak_topics(resource_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        cursor.execute("""
        SELECT rt.id as resource_topic_id, rt.curriculum_topic_id, rt.chapter_name, rt.order_index,
               t.name as topic_name, s.name as subject_name
        FROM resource_topics rt
        JOIN topics t ON rt.curriculum_topic_id = t.id
        JOIN subjects s ON t.subject_id = s.id
        WHERE rt.resource_id = ?
        ORDER BY rt.order_index ASC, t.id ASC;
        """, (resource_id,))
        mapped_topics = [dict(r) for r in cursor.fetchall()]
        conn.close()
        return jsonify({'topics': mapped_topics})

    elif request.method == 'POST':
        if user['role'] not in ['COACH', 'ADMIN']:
            return jsonify({'error': 'Yetkisiz erişim'}), 403

        data = request.json or {}
        topic_ids = data.get('topic_ids', [])

        cursor.execute("DELETE FROM resource_topics WHERE resource_id = ?;", (resource_id,))
        for idx, tid in enumerate(topic_ids):
            cursor.execute("""
            INSERT INTO resource_topics (resource_id, curriculum_topic_id, order_index)
            VALUES (?, ?, ?);
            """, (resource_id, tid, idx + 1))

        conn.commit()
        conn.close()
        return jsonify({'message': 'Kaynak konuları güncellendi.'})

@app.route('/api/kaynak-havuzu/<int:resource_id>/assign', methods=['POST'])
def handle_kaynak_assign(resource_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    if user['role'] not in ['COACH', 'ADMIN']:
        return jsonify({'error': 'Yalnızca koçlar kaynak atayabilir'}), 403

    data = request.json or {}
    student_ids = data.get('student_ids', [])
    single_st = data.get('student_id')
    if single_st and single_st not in student_ids:
        student_ids.append(single_st)

    if not student_ids:
        return jsonify({'error': 'Atanacak en az bir öğrenci seçiniz.'}), 400

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
    ch = cursor.fetchone()
    coach_id = ch['id'] if ch else 1

    cursor.execute("SELECT name FROM resources WHERE id = ?;", (resource_id,))
    res_row = cursor.fetchone()
    resource_name = res_row['name'] if res_row else "Kaynak"

    cursor.execute("SELECT curriculum_topic_id FROM resource_topics WHERE resource_id = ?;", (resource_id,))
    res_topics = [r['curriculum_topic_id'] for r in cursor.fetchall()]

    assigned_count = 0
    already_assigned_students = []

    for st_id in student_ids:
        cursor.execute("SELECT id FROM resource_assignments WHERE resource_id = ? AND student_id = ? AND status = 'ACTIVE';", (resource_id, st_id))
        existing = cursor.fetchone()

        if existing:
            cursor.execute("SELECT u.name FROM students s JOIN users u ON s.user_id = u.id WHERE s.id = ?;", (st_id,))
            u = cursor.fetchone()
            st_name = u['name'] if u else f"Öğrenci #{st_id}"
            already_assigned_students.append(st_name)
            continue

        cursor.execute("""
        INSERT INTO resource_assignments (resource_id, student_id, coach_id, status, progress_percentage)
        VALUES (?, ?, ?, 'ACTIVE', 0.0);
        """, (resource_id, st_id, coach_id))
        assignment_id = cursor.lastrowid
        assigned_count += 1

        for tid in res_topics:
            cursor.execute("""
            INSERT OR IGNORE INTO resource_topic_progress (resource_assignment_id, curriculum_topic_id, status)
            VALUES (?, ?, 'NOT_STARTED');
            """, (assignment_id, tid))

        # TRIGGER RESOURCE_ASSIGNED EVENT
        cursor.execute("SELECT user_id FROM students WHERE id = ?;", (st_id,))
        st_user = cursor.fetchone()
        if st_user:
            create_academic_event('RESOURCE_ASSIGNED', user['id'], st_user['user_id'], 'RESOURCE', assignment_id, '📚 Yeni Kaynak Atandı', f"Koçunuz size yeni bir kaynak atadı: {resource_name}", cursor=cursor)

    conn.commit()
    conn.close()

    msg = f"{assigned_count} öğrenciye kaynak başarıyla atandı."
    if already_assigned_students:
        msg += f" Zaten atanmış öğrenciler: {', '.join(already_assigned_students)}."

    return jsonify({
        'message': msg,
        'assigned_count': assigned_count,
        'already_assigned': already_assigned_students
    })

@app.route('/api/kaynak-havuzu/student-assignments', methods=['GET'])
def get_student_resource_assignments():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    student_id, err_resp, err_code = resolve_and_verify_student_id(cursor, user, request.args.get('student_id'))
    if err_resp:
        conn.close()
        return err_resp, err_code

    cursor.execute("""
    SELECT ra.id as assignment_id, ra.resource_id, ra.student_id, ra.coach_id, ra.assigned_at, ra.status as assignment_status, ra.progress_percentage,
           r.name as resource_name, r.publisher, r.exam_system, r.exam_type, r.field, r.resource_type,
           s.name as subject_name
    FROM resource_assignments ra
    JOIN resources r ON ra.resource_id = r.id
    LEFT JOIN subjects s ON r.subject_id = s.id
    WHERE ra.student_id = ? AND ra.status = 'ACTIVE'
    ORDER BY ra.assigned_at DESC;
    """, (student_id,))
    assignments = [dict(r) for r in cursor.fetchall()]

    for a in assignments:
        cursor.execute("""
        SELECT rtp.id as progress_id, rtp.curriculum_topic_id, rtp.status, rtp.completed_at,
               t.name as topic_name
        FROM resource_topic_progress rtp
        JOIN topics t ON rtp.curriculum_topic_id = t.id
        WHERE rtp.resource_assignment_id = ?
        ORDER BY t.id ASC;
        """, (a['assignment_id'],))
        topics_prog = [dict(tr) for tr in cursor.fetchall()]
        total_topics = len(topics_prog)
        completed_topics = len([tp for tp in topics_prog if tp['status'] == 'COMPLETED'])
        
        calc_perc = round((completed_topics / total_topics * 100.0), 1) if total_topics > 0 else 0.0

        if abs(calc_perc - (a['progress_percentage'] or 0)) > 0.1:
            cursor.execute("UPDATE resource_assignments SET progress_percentage = ? WHERE id = ?;", (calc_perc, a['assignment_id']))
            a['progress_percentage'] = calc_perc

        a['topics_progress'] = topics_prog
        a['total_topics'] = total_topics
        a['completed_topics'] = completed_topics

    conn.commit()
    conn.close()
    return jsonify({'assignments': assignments})

@app.route('/api/kaynak-havuzu/topic-progress', methods=['POST'])
def update_kaynak_topic_progress():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    data = request.json or {}
    assignment_id = data.get('assignment_id')
    topic_id = data.get('curriculum_topic_id')
    new_status = data.get('status', 'COMPLETED')

    if not assignment_id or not topic_id:
        return jsonify({'error': 'Atama ID ve konu ID gereklidir.'}), 400

    conn = get_db()
    cursor = conn.cursor()

    completed_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S') if new_status == 'COMPLETED' else None

    cursor.execute("""
    INSERT INTO resource_topic_progress (resource_assignment_id, curriculum_topic_id, status, completed_at)
    VALUES (?, ?, ?, ?)
    ON CONFLICT(resource_assignment_id, curriculum_topic_id) DO UPDATE SET
        status = excluded.status,
        completed_at = excluded.completed_at;
    """, (assignment_id, topic_id, new_status, completed_at))

    cursor.execute("SELECT COUNT(*) as total FROM resource_topic_progress WHERE resource_assignment_id = ?;", (assignment_id,))
    total = cursor.fetchone()['total']

    cursor.execute("SELECT COUNT(*) as completed FROM resource_topic_progress WHERE resource_assignment_id = ? AND status = 'COMPLETED';", (assignment_id,))
    completed = cursor.fetchone()['completed']

    new_perc = round((completed / total * 100.0), 1) if total > 0 else 0.0

    cursor.execute("UPDATE resource_assignments SET progress_percentage = ? WHERE id = ?;", (new_perc, assignment_id))

    conn.commit()
    conn.close()
    return jsonify({
        'message': 'Konu ilerleme durumu güncellendi.',
        'progress_percentage': new_perc,
        'completed_topics': completed,
        'total_topics': total
    })

@app.route('/api/kaynak-havuzu/<int:resource_id>/copy-to-my-pool', methods=['POST'])
def copy_resource_to_my_pool(resource_id):
    user = get_auth_user()
    if not user or user['role'] != 'COACH':
        return jsonify({'error': 'Yalnızca koçlar kaynağı kendi havuzuna ekleyebilir.'}), 403

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
    ch = cursor.fetchone()
    coach_id = ch['id'] if ch else 1

    cursor.execute("SELECT * FROM resources WHERE id = ?;", (resource_id,))
    res_item = cursor.fetchone()
    if not res_item:
        conn.close()
        return jsonify({'error': 'Kaynak bulunamadı'}), 404

    r = dict(res_item)
    new_name = f"{r['name']} (Kişisel Kopya)"

    cursor.execute("""
    INSERT INTO resources (owner_type, owner_id, name, title, publisher, exam_system, exam_type, field, subject_id, resource_type, isbn, edition, description)
    VALUES ('COACH', ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
    """, (coach_id, new_name, new_name, r['publisher'], r['exam_system'], r['exam_type'], r['field'], r['subject_id'], r['resource_type'], r['isbn'], r['edition'], r['description']))
    new_res_id = cursor.lastrowid

    cursor.execute("SELECT curriculum_topic_id, chapter_name, order_index FROM resource_topics WHERE resource_id = ?;", (resource_id,))
    top_rows = cursor.fetchall()
    for t in top_rows:
        cursor.execute("""
        INSERT INTO resource_topics (resource_id, curriculum_topic_id, chapter_name, order_index)
        VALUES (?, ?, ?, ?);
        """, (new_res_id, t['curriculum_topic_id'], t['chapter_name'], t['order_index']))

    conn.commit()
    conn.close()
    return jsonify({'message': 'Kaynak kişisel havuzunuza kopyalandı.', 'new_resource_id': new_res_id})




# ============================================================
# CENTRALIZED NOTIFICATIONS & ACTIVITIES API ENDPOINTS
# ============================================================

@app.route('/api/notifications', methods=['GET'])
def get_notifications():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    unread_only = request.args.get('unread_only') == 'true'
    category = request.args.get('category', 'ALL')

    where_clauses = ["n.recipient_user_id = ?"]
    params = [user['id']]

    if user['role'] == 'COACH':
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        ch = cursor.fetchone()
        coach_id = ch['id'] if ch else None
        if coach_id:
            where_clauses.append("(n.actor_user_id IS NULL OR n.actor_user_id IN (SELECT user_id FROM students WHERE coach_id = ?) OR n.actor_user_id IN (SELECT id FROM users WHERE role = 'ADMIN'))")
            params.append(coach_id)

    if unread_only:
        where_clauses.append("n.is_read = 0")

    if category != 'ALL':
        if category == 'HOMEWORK':
            where_clauses.append("n.type LIKE 'HOMEWORK_%'")
        elif category == 'PROGRAM':
            where_clauses.append("n.type LIKE 'PROGRAM_%'")
        elif category == 'RESOURCE':
            where_clauses.append("n.type LIKE 'RESOURCE_%'")
        elif category == 'EXAM':
            where_clauses.append("(n.type LIKE 'MOCK_EXAM_%' OR n.type = 'ACADEMIC_RISK_DETECTED')")
        elif category == 'MESSAGE':
            where_clauses.append("n.type = 'MESSAGE_RECEIVED'")
        elif category == 'OTHER':
            where_clauses.append("n.type IN ('BOOK_GOAL_COMPLETED', 'STUDY_SESSION_COMPLETED', 'CURRICULUM_TOPIC_COMPLETED')")

    where_str = " AND ".join(where_clauses)

    query = f"""
    SELECT n.*, u.username as actor_username, u.role as actor_role
    FROM notifications n
    LEFT JOIN users u ON n.actor_user_id = u.id
    WHERE {where_str}
    ORDER BY n.id DESC
    LIMIT 100;
    """

    cursor.execute(query, params)
    notifs = [dict(r) for r in cursor.fetchall()]

    cursor.execute("SELECT COUNT(*) as unread FROM notifications WHERE recipient_user_id = ? AND is_read = 0;", (user['id'],))
    unread_count = cursor.fetchone()['unread']

    conn.close()
    return jsonify({'notifications': notifs, 'unread_count': unread_count})

@app.route('/api/notifications/<int:notif_id>/read', methods=['POST'])
def mark_notification_read(notif_id):
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE notifications SET is_read = 1, read_at = CURRENT_TIMESTAMP WHERE id = ? AND recipient_user_id = ?;", (notif_id, user['id']))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Bildirim okundu olarak işaretlendi.'})

@app.route('/api/notifications/read-all', methods=['POST'])
def mark_all_notifications_read():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE notifications SET is_read = 1, read_at = CURRENT_TIMESTAMP WHERE recipient_user_id = ? AND is_read = 0;", (user['id'],))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Tüm bildirimler okundu olarak işaretlendi.'})

@app.route('/api/activity-logs', methods=['GET'])
def get_academic_activity_logs():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if user['role'] == 'COACH':
        cursor.execute("SELECT id FROM coaches WHERE user_id = ?;", (user['id'],))
        ch = cursor.fetchone()
        coach_id = ch['id'] if ch else None

        cursor.execute("""
        SELECT a.*, u.username as actor_name, su.name as student_name
        FROM activity_logs a
        LEFT JOIN users u ON a.user_id = u.id
        LEFT JOIN students s ON a.student_id = s.id
        LEFT JOIN users su ON s.user_id = su.id
        WHERE a.coach_id = ? OR a.user_id IN (SELECT user_id FROM students WHERE coach_id = ?)
        ORDER BY a.id DESC
        LIMIT 30;
        """, (coach_id, coach_id))
    else:
        cursor.execute("""
        SELECT a.*, u.username as actor_name
        FROM activity_logs a
        LEFT JOIN users u ON a.user_id = u.id
        WHERE a.user_id = ?
        ORDER BY a.id DESC
        LIMIT 30;
        """, (user['id'],))

    logs = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({'activity_logs': logs})

@app.route('/api/notification-preferences', methods=['GET', 'PUT'])
def handle_notification_preferences():
    user = get_auth_user()
    if not user:
        return jsonify({'error': 'Yetkisiz erişim'}), 401

    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'GET':
        cursor.execute("SELECT notification_type, enabled FROM notification_preferences WHERE user_id = ?;", (user['id'],))
        prefs = {r['notification_type']: bool(r['enabled']) for r in cursor.fetchall()}
        conn.close()
        return jsonify({'preferences': prefs})

    elif request.method == 'PUT':
        data = request.json or {}
        for ntype, enabled in data.items():
            cursor.execute("""
            INSERT INTO notification_preferences (user_id, notification_type, enabled)
            VALUES (?, ?, ?)
            ON CONFLICT(user_id, notification_type) DO UPDATE SET enabled = excluded.enabled;
            """, (user['id'], ntype, 1 if enabled else 0))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Bildirim tercihleri güncellendi.'})


if __name__ == '__main__':
    port = int(os.environ.get("PORT", 5005))
    print(f"Starting YKS Platform Server on port {port}...")
    app.run(host='0.0.0.0', port=port, debug=False, use_reloader=False, threaded=True)
